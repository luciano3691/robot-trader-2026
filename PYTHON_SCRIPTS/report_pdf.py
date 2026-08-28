# -*- coding: utf-8 -*-
"""
RT2026 — Generatore PDF report screener  (v2 — 25/08/2026)
Genera un PDF A4 verticale dal foglio 'Top N per Score' dell'Excel.
Aggiunge:
  - nota ticker stabili (frequenza da ticker_frequency.json)
  - pagina 2: Scheda Ordine stampabile (Task 4)
"""
import io, os, base64 as _b64lib, urllib.request
from datetime import datetime

try:
    from assets import FUERTE_LOGO_B64 as _LOGO_B64
except ImportError:
    _LOGO_B64 = None

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ── Brand FVC ────────────────────────────────────────────────────────────────
_BLU  = (44,  82, 130)    # #2C5282
_DARK = (15,  23,  42)    # #0f172a
_ROW  = (241, 245, 249)   # righe alternate #F1F5F9
_GRN  = (39, 103,  73)    # verde stabilità
_GOLD = (180, 120,  30)   # oro FVC

_LOGO_URL  = 'https://www.fuerteventurecapital.com/assets/fuerte_venture_logo_0526.png'
_LOGO_PATH = '/tmp/fvc_logo_report.png'


def _get_logo():
    if not os.path.exists(_LOGO_PATH):
        if _LOGO_B64:
            try:
                with open(_LOGO_PATH, 'wb') as _f:
                    _f.write(_b64lib.b64decode(_LOGO_B64.strip()))
            except Exception:
                pass
        if not os.path.exists(_LOGO_PATH):
            try:
                urllib.request.urlretrieve(_LOGO_URL, _LOGO_PATH)
            except Exception:
                return None
    return _LOGO_PATH if os.path.exists(_LOGO_PATH) else None


def _stars_to_str(val):
    """Converte '★★★★☆' in '4/5' evitando problemi font Unicode."""
    if val is None:
        return 'N/A'
    s = str(val)
    if 'N/A' in s or s.strip() == '':
        return 'N/A'
    filled = s.count('★')
    return f'{filled}/5' if filled else 'N/A'


def _fmt_float(val):
    if val is None:
        return ''
    try:
        return f'{float(val):.2f}'
    except Exception:
        return str(val)


def _truncate(val, maxlen):
    s = str(val) if val is not None else ''
    return s[:maxlen - 1] + '.' if len(s) > maxlen else s


# ── Frequenza ticker ─────────────────────────────────────────────────────────
_BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
_FREQ_FILE = os.path.join(_BASE_DIR, 'ticker_frequency.json')

def _load_freq_map(asset_class, piano):
    """Ritorna ({ticker_upper: count}, total_days) dal file ticker_frequency.json."""
    import json
    try:
        if os.path.exists(_FREQ_FILE):
            with open(_FREQ_FILE, encoding='utf-8') as f:
                data = json.load(f)
            plan_data = data.get(asset_class, {}).get(piano, {})
            total_days = plan_data.get('_meta', {}).get('total_days', 1)
            freq_map = {
                tk.upper(): info.get('count', 0)
                for tk, info in plan_data.items()
                if tk != '_meta'
            }
            return freq_map, max(total_days, 1)
    except Exception:
        pass
    return {}, 1


# ── Configurazione colonne ────────────────────────────────────────────────────
# Ogni voce: (keyword_header, larghezza_mm, align, formatter|None)
# keyword_header speciale: '#' = contatore riga; 'FREQ' = frequenza ticker
# Layout landscape A4 → usable width = 297-12-12 = 273mm

_AZIONI_COLS = [
    ('#',                7,  'C', None),
    ('Ticker',          14,  'L', None),
    ('Nome',            36,  'L', None),
    ('Valuta',           9,  'C', None),
    ('Mercato',         18,  'L', None),
    ('Indice',          17,  'L', None),
    ('Prezzo',          15,  'R', _fmt_float),
    ('Var 1D %',        12,  'R', None),
    ('Score',           11,  'C', None),
    ('Perf 1M %',       11,  'R', None),
    ('Perf 3M %',       11,  'R', None),
    ('Perf 6M %',       11,  'R', None),
    ('Perf YTD %',      11,  'R', None),
    ('Perf 1Y %',       11,  'R', None),
    ('P/B',             10,  'R', _fmt_float),
    ('ROE',             10,  'R', None),
    ('EV/FCF',          11,  'R', _fmt_float),
    ('Net Debt/EBITDA', 14,  'R', _fmt_float),
    ('Market Cap',      14,  'R', None),
    ('Settore',         20,  'L', None),
]  # total 273mm — universale BASIC/PRO/VALUE

_ETF_COLS = [
    ('#',            12,  'C', None),
    ('Ticker',       35,  'L', None),
    ('Nome',        112,  'L', None),
    ('Score',        23,  'C', None),
    ('FREQ',         18,  'C', None),
    ('Perf 1Y %',    27,  'R', None),
    ('Perf 6M %',    24,  'R', None),
    ('Perf 3M %',    22,  'R', None),
]  # total 273mm

_FONDI_COLS = [
    ('#',            12,  'C', None),
    ('Ticker',       28,  'L', None),
    ('Nome',         95,  'L', None),
    ('Score',        25,  'C', None),
    ('FREQ',         18,  'C', None),
    ('Stelle MS',    25,  'C', _stars_to_str),
    ('Perf 1Y %',    28,  'R', None),
    ('Perf 6M %',    21,  'R', None),
    ('Perf 3M %',    21,  'R', None),
]  # total 273mm

_FONDI_EU_COLS = [
    ('#',            12,  'C', None),
    ('Nome',        106,  'L', None),
    ('Gestore',      47,  'L', None),
    ('Score',        23,  'C', None),
    ('FREQ',         18,  'C', None),
    ('Stelle MS',    23,  'C', _stars_to_str),
    ('Perf 1Y %',    23,  'R', None),
    ('Perf 6M %',    21,  'R', None),
]  # total 273mm

_PLAN_COLS = {
    'AZIONI':   {'BASIC': _AZIONI_COLS,  'PRO': _AZIONI_COLS,  'VALUE': _AZIONI_COLS},
    'ETF':      {'BASIC': _ETF_COLS,     'PRO': _ETF_COLS,     'VALUE': _ETF_COLS},
    'FONDI':    {'BASIC': _FONDI_COLS,   'PRO': _FONDI_COLS,   'VALUE': _FONDI_COLS},
    'FONDI_EU': {'BASIC': _FONDI_EU_COLS,'PRO': _FONDI_EU_COLS,'VALUE': _FONDI_EU_COLS},
}

_ASSET_LABEL = {
    'AZIONI':   'Azioni Globali',
    'ETF':      'ETF',
    'FONDI':    'Fondi USA',
    'FONDI_EU': 'Fondi UCITS Europa',
}

_COL_LABEL = {
    'Stelle MS':        'Stelle MS',
    'Perf 1Y %':        '1A%',
    'Perf 6M %':        '6M%',
    'Perf 3M %':        '3M%',
    'Perf 1M %':        '1M%',
    'Perf YTD %':       'YTD%',
    'Var 1D %':         'Var1D%',
    'EV/FCF':           'EV/FCF',
    'Net Debt/EBITDA':  'ND/EBITDA',
    'Market Cap':       'Mkt Cap',
    'Valuta':           'Val.',
    'FREQ':             'Freq',
}

# Troncamento: solo eccezioni — per le altre colonne si usa width/1.6 automaticamente
_TRUNCATE_LEN = {
    'Gestore': 28,
}

# Colonne ordine per pagina Scheda Ordine (larghezze per asset class)
# 'SEL' = checkbox stampabile (□); 'QTA' = campo vuoto quantita; 'IMP' = calcolo importo
_ORDER_COLS_AZIONI = [
    ('SEL',      8,  'C'),    # checkbox □
    ('#',        7,  'C'),
    ('Ticker',  20,  'L'),
    ('Nome',    48,  'L'),
    ('Mercato', 18,  'L'),
    ('Prezzo',  18,  'R'),
    ('Valuta',  12,  'C'),
    ('QTA',     26,  'C'),    # blank per quantita
    ('IMP',     29,  'R'),    # blank per importo
]  # total = 186mm

_ORDER_COLS_FUND = [
    ('SEL',      8,  'C'),
    ('#',        7,  'C'),
    ('Ticker',  22,  'L'),
    ('Nome',    73,  'L'),
    ('Prezzo',  18,  'R'),
    ('Valuta',  12,  'C'),
    ('QTA',     22,  'C'),
    ('IMP',     24,  'R'),
]  # total = 186mm


def _find_col_idx(headers, keyword):
    """Trova indice colonna per keyword parziale, case-insensitive."""
    if keyword in ('#', 'FREQ'):
        return keyword
    kw = keyword.lower().replace(' ', '').replace('/', '').replace('%', '')
    for i, h in enumerate(headers):
        if h is None:
            continue
        hn = str(h).lower().replace(' ', '').replace('/', '').replace('%', '')
        if kw in hn or hn in kw:
            return i
    return None


# ── Classe PDF ────────────────────────────────────────────────────────────────

class _ReportPDF(FPDF if FPDF else object):
    def __init__(self, title1, title2, date_str, logo_path):
        super().__init__(orientation='L', unit='mm', format='A4')
        self._t1   = title1
        self._t2   = title2
        self._date = date_str
        self._logo = logo_path
        self.set_margins(12, 44, 12)   # top=44 per ospitare logo quadrato 32mm
        self.set_auto_page_break(auto=True, margin=12)

    def header(self):
        if self._logo:
            self.image(self._logo, x=12, y=5, w=32)   # 32mm → finisce a y=37
        self.set_xy(48, 8)
        self.set_font('Helvetica', 'B', 14)
        self.set_text_color(*_BLU)
        self.cell(0, 7, self._t1)
        self.set_xy(48, 18)
        self.set_font('Helvetica', '', 9)
        self.set_text_color(90, 90, 90)
        self.cell(0, 5, f'{self._t2}  |  Report del {self._date}')
        self.set_draw_color(*_BLU)
        self.set_line_width(0.4)
        self.line(12, 39, 285, 39)
        self.set_xy(12, 43)

    def footer(self):
        self.set_y(-11)
        self.set_font('Helvetica', 'I', 6.5)
        self.set_text_color(150, 150, 150)
        self.cell(
            0, 4,
            'Robot Trader 2026 - Fuerte Venture Capital SL - CIF B23881691 - '
            'trader.fuerteventurecapital.com  |  '
            'Documento ad uso esclusivo del destinatario. Non costituisce consulenza finanziaria.',
            align='C',
        )
        self.ln(3)
        self.cell(0, 3, f'Pagina {self.page_no()}', align='R')


# ── Pagina Scheda Ordine ──────────────────────────────────────────────────────

def _pagina_ordini(pdf, data_rows, sheet_headers, asset_class, piano, date_str, logo_path):
    """
    Pagina 2: Scheda Ordine.
    Riproduce la logica della pagina web /ordine-bancario:
    tutti i Top N sono pre-caricati, il cliente spunta quelli che VUOLE
    (checkbox SEL) e barra quelli che ESCLUDE.
    """
    pdf.add_page()

    # Logo + titolo
    if logo_path:
        pdf.image(logo_path, x=12, y=8, w=26)
    pdf.set_xy(42, 8)
    pdf.set_font('Helvetica', 'B', 12)
    pdf.set_text_color(*_BLU)
    pdf.cell(0, 6, 'Robot Trader 2026 - Ordine Acquisto Titoli')
    pdf.set_xy(42, 15)
    pdf.set_font('Helvetica', '', 9)
    pdf.set_text_color(90, 90, 90)
    asset_lbl = _ASSET_LABEL.get(asset_class, asset_class)
    pdf.cell(0, 5, f'{asset_lbl}  -  Piano {piano}  |  Data: {date_str}')
    pdf.set_draw_color(*_BLU)
    pdf.set_line_width(0.4)
    pdf.line(12, 23, 198, 23)
    pdf.set_xy(12, 27)

    # Istruzione operativa
    pdf.set_fill_color(235, 244, 255)
    pdf.set_draw_color(*_BLU)
    pdf.set_line_width(0.3)
    pdf.rect(12, pdf.get_y(), 186, 14, 'FD')
    pdf.set_xy(14, pdf.get_y() + 1.5)
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_text_color(*_BLU)
    pdf.cell(0, 4, 'ISTRUZIONI: Spunta (V) nella colonna SEL i titoli che vuoi acquistare. '
                   'Barra (X) quelli da escludere.')
    pdf.set_xy(14, pdf.get_y() + 1)
    pdf.set_font('Helvetica', '', 7.5)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(0, 4, 'Inserisci la Quantita per ogni titolo selezionato. '
                   'Per ordini digitali: trader.fuerteventurecapital.com > Ordine Bancario')
    pdf.ln(8)

    # Seleziona configurazione colonne
    use_fund_cols = asset_class in ('ETF', 'FONDI', 'FONDI_EU')
    ord_cols = _ORDER_COLS_FUND if use_fund_cols else _ORDER_COLS_AZIONI

    # Risolvi indici colonne nel foglio
    idx = {}
    for kw in ('Ticker', 'Nome', 'Mercato', 'Prezzo', 'Valuta'):
        idx[kw] = _find_col_idx(sheet_headers, kw)

    row_h  = 7.5
    head_h = 8.0

    # Intestazione colonne
    pdf.set_fill_color(*_BLU)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 7.5)
    col_labels = {
        'SEL': 'SEL', '#': '#', 'Ticker': 'Ticker', 'Nome': 'Denominazione',
        'Mercato': 'Mercato', 'Prezzo': 'Prezzo Rif.', 'Valuta': 'Val.',
        'QTA': 'Quantita', 'IMP': 'Importo EUR',
    }
    for (name, w, al) in ord_cols:
        pdf.cell(w, head_h, col_labels.get(name, name), align='C', fill=True)
    pdf.ln()

    # Righe dati
    pdf.set_line_width(0.2)
    for ri, raw_row in enumerate(data_rows):
        bg = _ROW if ri % 2 == 1 else (255, 255, 255)
        pdf.set_fill_color(*bg)

        row_y = pdf.get_y()

        for (name, w, al) in ord_cols:
            if name == 'SEL':
                # Disegna checkbox vuoto (□)
                cx = pdf.get_x() + (w - 4) / 2
                cy = row_y + (row_h - 4) / 2
                pdf.set_fill_color(*bg)
                pdf.cell(w, row_h, '', fill=True)
                pdf.set_draw_color(80, 80, 80)
                pdf.rect(cx, cy, 4, 4)
                pdf.set_draw_color(200, 200, 200)

            elif name == '#':
                pdf.set_fill_color(*bg)
                pdf.set_text_color(120, 120, 120)
                pdf.set_font('Helvetica', '', 7.5)
                pdf.cell(w, row_h, str(ri + 1), align='C', fill=True)

            elif name == 'QTA':
                pdf.set_fill_color(*bg)
                pdf.set_draw_color(160, 160, 160)
                pdf.cell(w, row_h, '', fill=True, border=1)
                pdf.set_draw_color(200, 200, 200)

            elif name == 'IMP':
                pdf.set_fill_color(*bg)
                pdf.set_draw_color(160, 160, 160)
                pdf.cell(w, row_h, '', fill=True, border=1)
                pdf.set_draw_color(200, 200, 200)

            else:
                pdf.set_fill_color(*bg)
                col_idx = idx.get(name)
                raw = raw_row[col_idx] if (col_idx is not None and col_idx < len(raw_row)) else None

                if name == 'Ticker':
                    val_str = str(raw).strip() if raw else ''
                    pdf.set_text_color(*_BLU)
                    pdf.set_font('Helvetica', 'B', 8)
                elif name == 'Prezzo':
                    val_str = _fmt_float(raw)
                    pdf.set_text_color(*_DARK)
                    pdf.set_font('Helvetica', 'B', 8)
                elif name == 'Nome':
                    val_str = _truncate(str(raw).strip() if raw else '', 28)
                    pdf.set_text_color(*_DARK)
                    pdf.set_font('Helvetica', '', 8)
                else:
                    val_str = str(raw).strip() if raw else ''
                    if val_str.lower() in ('nan', 'none', ''):
                        val_str = ''
                    val_str = _truncate(val_str, 18)
                    pdf.set_text_color(*_DARK)
                    pdf.set_font('Helvetica', '', 8)

                pdf.cell(w, row_h, val_str, align=al, fill=True)

        pdf.ln()

    # Riga TOTALE
    pdf.ln(1)
    pdf.set_fill_color(*_DARK)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    pdf.set_draw_color(160, 160, 160)
    for (name, w, al) in ord_cols:
        if name == 'IMP':
            pdf.cell(w, 8, 'TOTALE:', fill=True, border=1, align='R')
        elif name == 'QTA':
            pdf.cell(w, 8, '', fill=True, border=1)
        elif name in ('#', 'SEL', 'Valuta', 'Mercato'):
            pdf.cell(w, 8, '', fill=True)
        else:
            pdf.cell(w, 8, '', fill=True)
    pdf.ln()

    # Spazio firme
    pdf.ln(6)
    pdf.set_line_width(0.3)
    pdf.set_draw_color(120, 120, 120)
    pdf.set_font('Helvetica', '', 8)
    pdf.set_text_color(60, 60, 60)
    pdf.cell(88, 5, 'Luogo e Data: ___________________________', align='L')
    pdf.cell(0,  5, 'Firma del Cliente: ________________________________', align='R')
    pdf.ln(10)
    pdf.cell(0, 5,
             'Banca / Intermediario: ______________________________   '
             'Conto Titoli n.: ________________________   '
             'Gestore: _____________________', align='L')
    pdf.ln(9)

    # Disclaimer MiFID
    pdf.set_font('Helvetica', 'I', 6.5)
    pdf.set_text_color(150, 150, 150)
    pdf.multi_cell(
        0, 3.5,
        'AVVERTENZE MiFID II: Le presenti istruzioni sono predisposte autonomamente dal cliente a titolo '
        'personale e non costituiscono consulenza finanziaria, gestione patrimoniale ne raccomandazione '
        'di investimento. Robot Trader 2026 fornisce esclusivamente un servizio di screening '
        'quantitativo informativo. I prezzi indicati sono di riferimento alla data del report e non '
        'costituiscono prezzi garantiti di esecuzione. Il cliente e l\'unico responsabile delle '
        'istruzioni impartite al proprio intermediario bancario. '
        'Per ordini digitali completi: trader.fuerteventurecapital.com/ordine-bancario'
    )


# ── Pagina Legenda ────────────────────────────────────────────────────────────

def _legend_section_header(pdf, title):
    pdf.set_fill_color(*_BLU)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 8)
    pdf.cell(273, 7, f'  {title}', fill=True, align='L')
    pdf.ln()


def _legend_row(pdf, label, desc, row_idx):
    W_LABEL, W_DESC, LH = 58, 215, 3.5
    pdf.set_font('Helvetica', '', 6.5)
    words = desc.split()
    lines, cur = 1, ''
    for word in words:
        test = f'{cur} {word}'.strip()
        if pdf.get_string_width(test) > W_DESC - 3:
            lines += 1
            cur = word
        else:
            cur = test
    h = max(lines * LH + 3, 7.5)

    bg = _ROW if row_idx % 2 else (255, 255, 255)
    y0 = pdf.get_y()
    pdf.set_fill_color(*bg)
    pdf.rect(12, y0, W_LABEL + W_DESC, h, 'F')

    pdf.set_xy(13, y0 + 1.5)
    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_text_color(*_BLU)
    pdf.multi_cell(W_LABEL - 2, LH, label)

    pdf.set_xy(12 + W_LABEL + 1, y0 + 1.5)
    pdf.set_font('Helvetica', '', 6.5)
    pdf.set_text_color(*_DARK)
    pdf.multi_cell(W_DESC - 2, LH, desc)

    pdf.set_xy(12, y0 + h)


def _pagina_legenda(pdf, piano, asset_class):
    """
    Pagina 1 del report: Guida alla lettura.
    Contiene: descrizione Score, indicatori, colonne.
    Escluse: pesi Score, elenco fogli Excel.
    """
    _legend_section_header(pdf, 'LO SCORE  (0 - 100)')
    _legend_row(pdf, "Cos'e' lo Score",
        "Un punteggio da 0 a 100 che sintetizza la qualita' dell'azione per questo piano. "
        "PIU' ALTO = MIGLIORE. Le azioni sono ordinate dal punteggio piu' alto al piu' basso.", 0)
    _legend_row(pdf, 'Come si calcola',
        "Per ogni indicatore, l'azione riceve un percentile 0-100 rispetto a tutte le altre selezionate. "
        "I percentili vengono pesati e sommati per ottenere lo Score finale.", 1)
    pdf.ln(3)

    _legend_section_header(pdf, 'INDICATORI - COLONNE')
    indicators = [
        ('Ticker',
         "Codice di borsa dell'azione (es. AAPL, NESN.SW, MC.PA)."),
        ('Nome',
         "Ragione sociale dell'azienda quotata."),
        ('Val. / Mercato / Indice',
         "Val. = divisa del prezzo (USD, EUR, GBP...) | Mercato = borsa di quotazione | "
         "Indice = indice di appartenenza (S&P 500, NASDAQ, FTSE 100, DAX 40...)."),
        ('Prezzo  /  Var 1D %',
         "Prezzo corrente in valuta locale | Var. percentuale rispetto alla chiusura del giorno precedente."),
        ('Score  (0 - 100)',
         "Punteggio composito piu' alto = migliore. L'azione in cima e' la migliore del piano per questo screener."),
        ('1M% / 3M% / 6M% / YTD% / 1A%',
         "Performance del prezzo: ultimo mese / trimestre / semestre / da inizio anno / ultimo anno. Prezzi puri, senza aggiustamento dividendi."),
        ('P/B  (Price / Book)',
         "Prezzo / Patrimonio netto. Sottovalutata se P/B < 1. MEGLIO SE BASSO."),
        ('ROE  (Return on Equity)',
         "Utile netto / Patrimonio. Es. ROE = 8% = EUR 8 guadagnati per ogni EUR 100 di patrimonio. MEGLIO SE ALTO."),
        ('EV / FCF',
         "Enterprise Value / Free Cash Flow. Quanti anni di cassa vale l'azienda intera. MEGLIO SE BASSO."),
        ('ND / EBITDA  (Net Debt / EBITDA)',
         "Leva finanziaria: anni di EBITDA necessari a ripagare il debito. Negativo = cassa netta (ottimo). MEGLIO SE BASSO."),
        ('Market Cap',
         "Capitalizzazione = Prezzo x Azioni in circolazione. M = milioni | B = miliardi | T = trilioni (valuta locale)."),
        ('Settore',
         "Classificazione GICS: Technology, Healthcare, Energy, Industrials, Consumer Discretionary, "
         "Communication Services, Basic Materials, Financials, Real Estate, Utilities."),
    ]
    for i, (lbl, dsc) in enumerate(indicators):
        _legend_row(pdf, lbl, dsc, i)


# ── Funzione pubblica ─────────────────────────────────────────────────────────

def genera_report_pdf(excel_path, piano, asset_class,
                      n_selezionati=0, n_analizzati=0):
    """
    Genera PDF A4 verticale dal foglio 'Top N per Score' dell'Excel.
    Pagina 1: tabella Top N con colonna frequenza.
    Pagina 2: Scheda Ordine stampabile (solo AZIONI).

    Parametri:
        excel_path    percorso al file .xlsx generato dallo screener
        piano         'BASIC' | 'PRO' | 'VALUE'
        asset_class   'AZIONI' | 'ETF' | 'FONDI' | 'FONDI_EU'
        n_selezionati numero titoli passati i filtri
        n_analizzati  numero titoli analizzati totali

    Ritorna bytes del PDF, o None in caso di errore.
    """
    if FPDF is None:
        print('[report_pdf] fpdf2 non installato')
        return None
    if load_workbook is None:
        print('[report_pdf] openpyxl non installato')
        return None

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(f'[report_pdf] errore apertura {excel_path}: {e}')
        return None

    top_sheet = next((s for s in wb.sheetnames if 'Top' in s and 'Score' in s), None)
    if not top_sheet:
        print(f'[report_pdf] foglio Top non trovato in {excel_path}')
        wb.close()
        return None

    ws       = wb[top_sheet]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(all_rows) < 3:
        return None

    sheet_headers = list(all_rows[0])
    data_rows = [r for r in all_rows[2:] if any(v is not None for v in r)]
    if not data_rows:
        return None

    cols_config = _PLAN_COLS.get(asset_class, {}).get(piano)
    if not cols_config:
        print(f'[report_pdf] nessuna config colonne per {asset_class}/{piano}')
        return None

    # Carica mappa frequenza ticker
    freq_map, total_days = _load_freq_map(asset_class, piano)

    # Risolvi indici colonne nel foglio
    resolved = []
    for (kw, w, al, fmt) in cols_config:
        if kw in ('#', 'FREQ'):
            resolved.append((kw, kw, w, al, fmt))
        else:
            idx = _find_col_idx(sheet_headers, kw)
            if idx is None:
                print(f'[report_pdf] colonna "{kw}" non trovata — skip')
                continue
            resolved.append((kw, idx, w, al, fmt))

    if not resolved:
        return None

    # Titoli
    asset_lbl  = _ASSET_LABEL.get(asset_class, asset_class)
    top_count  = len(data_rows)
    title1     = f'Robot Trader 2026 - {asset_lbl}'
    if n_selezionati:
        title2 = f'Piano {piano}  |  Top {top_count} su {n_selezionati} selezionati'
    else:
        title2 = f'Piano {piano}  |  Top {top_count} per Score'
    date_str   = datetime.now().strftime('%d/%m/%Y')

    logo = _get_logo()
    pdf  = _ReportPDF(title1, title2, date_str, logo)

    # ── Pagina 1: Legenda ─────────────────────────────────────────────────────
    pdf.add_page()
    _pagina_legenda(pdf, piano, asset_class)

    # ── Pagina 2: Tabella dati ────────────────────────────────────────────────
    pdf.add_page()

    row_h  = 6.5
    head_h = 7.5

    # ── Header tabella ────────────────────────────────────────────────────────
    pdf.set_fill_color(*_BLU)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 7.5)
    for (name, idx, w, al, _fmt) in resolved:
        label = _COL_LABEL.get(name, name)
        pdf.cell(w, head_h, label, align='C', fill=True)
    pdf.ln()

    # ── Righe dati ────────────────────────────────────────────────────────────
    # Precalcola indice ticker per lookup frequenza
    ticker_idx = _find_col_idx(sheet_headers, 'Ticker')
    if ticker_idx == '#' or ticker_idx == 'FREQ':
        ticker_idx = None

    pdf.set_font('Helvetica', '', 7.5)
    for ri, raw_row in enumerate(data_rows):
        if ri % 2 == 1:
            pdf.set_fill_color(*_ROW)
        else:
            pdf.set_fill_color(255, 255, 255)

        # Ottieni ticker per frequenza
        row_ticker = ''
        if ticker_idx is not None and ticker_idx < len(raw_row):
            row_ticker = str(raw_row[ticker_idx]).strip().upper() if raw_row[ticker_idx] else ''

        for (name, idx, w, al, fmt) in resolved:
            if idx == '#':
                val_str = str(ri + 1)
            elif idx == 'FREQ':
                cnt = freq_map.get(row_ticker, 0)
                pct = round(cnt / total_days * 100) if cnt > 0 else 0
                if pct >= 50:
                    pdf.set_text_color(*_GRN)
                    pdf.set_font('Helvetica', 'B', 7.5)
                elif pct >= 25:
                    pdf.set_text_color(*_GOLD)
                    pdf.set_font('Helvetica', 'B', 7.5)
                else:
                    pdf.set_text_color(160, 160, 160)
                    pdf.set_font('Helvetica', '', 7)
                val_str = f'{pct}%' if cnt > 0 else '-'
                pdf.cell(w, row_h, val_str, align='C', fill=True)
                pdf.set_font('Helvetica', '', 7.5)
                pdf.set_text_color(*_DARK)
                continue
            else:
                raw = raw_row[idx] if idx < len(raw_row) else None
                if fmt:
                    val_str = fmt(raw)
                elif raw is None:
                    val_str = ''
                elif isinstance(raw, float):
                    val_str = f'{raw:.2f}'
                elif isinstance(raw, int):
                    val_str = str(raw)
                else:
                    val_str = str(raw)

            # Troncamento: esplicito o automatico da larghezza colonna (~1.6mm/char a 7.5pt)
            maxlen = _TRUNCATE_LEN.get(name, max(int(w / 1.6), 4))
            val_str = _truncate(val_str, maxlen)

            # Score in blu grassetto
            if name == 'Score':
                pdf.set_font('Helvetica', 'B', 7.5)
                pdf.set_text_color(*_BLU)
                pdf.cell(w, row_h, val_str, align=al, fill=True)
                pdf.set_font('Helvetica', '', 7.5)
                pdf.set_text_color(*_DARK)
            else:
                pdf.set_text_color(*_DARK)
                pdf.cell(w, row_h, val_str, align=al, fill=True)

        pdf.ln()

    # ── Nota footer tabella ───────────────────────────────────────────────────
    pdf.ln(3)
    pdf.set_font('Helvetica', 'I', 7)
    pdf.set_text_color(110, 110, 110)
    parts = []
    if n_selezionati:
        parts.append(f'{n_selezionati} strumenti selezionati dai filtri')
    if n_analizzati:
        parts.append(f'su {n_analizzati:,} analizzati'.replace(',', '.'))
    parts.append(f'mostrati i primi {top_count} per Score')
    pdf.cell(0, 5, '  |  '.join(parts), align='L')
    pdf.ln(4)

    # ── Nota ticker stabili ───────────────────────────────────────────────────
    stabili = [
        (tk, cnt, round(cnt / total_days * 100))
        for tk, cnt in freq_map.items()
        if round(cnt / total_days * 100) >= 50
    ]
    stabili.sort(key=lambda x: -x[2])
    if stabili:
        pdf.set_font('Helvetica', 'B', 7)
        pdf.set_text_color(*_GRN)
        pdf.cell(36, 4, 'Titoli stabili (>=50%):', align='L')
        pdf.set_font('Helvetica', '', 7)
        pdf.set_text_color(*_DARK)
        pdf.multi_cell(0, 4, '  '.join(f'{tk}({pct}%)' for tk, _cnt, pct in stabili[:15]), align='L')

    buf = io.BytesIO()
    pdf.output(buf)
    return buf.getvalue()
