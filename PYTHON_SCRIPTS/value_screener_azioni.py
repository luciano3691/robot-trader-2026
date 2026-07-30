# -*- coding: utf-8 -*-
"""
VALUE SCREENER AZIONI - Robot Trader 2026
FIXED: Formato percentuale italiano + Parametri dinamici da parametri.json
Author: Fuerte Venture Capital SL
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
from datetime import datetime
import time
import json
import smtplib
from email.mime.text import MIMEText
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout


def _is_network_error(detail):
    if not detail:
        return False
    d = detail.lower()
    return any(k in d for k in ['curl', 'resolve host', 'connection', 'timeout', 'network', 'errno 11001', 'recv failure'])


def _send_alert_email_azioni(total, network_errors, non_validi_total):
    """Invia email di alert all'admin se troppi errori di rete durante lo screening azioni."""
    try:
        with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json'), encoding='utf-8') as f:
            cfg_email = json.load(f).get('email', {})
        host   = cfg_email.get('smtp_server', 'smtp.gmail.com')
        port   = cfg_email.get('smtp_port', 587)
        login  = cfg_email.get('smtp_login', '') or cfg_email.get('sender', '')
        pwd    = cfg_email.get('app_password', '')
        sender = cfg_email.get('sender', login)
        if not login or not pwd:
            print(f"[ALERT] SMTP non configurato — alert non inviato", flush=True)
            return
        soglia_pct = round(network_errors / total * 100, 1) if total else 0
        body = (
            f"ALERT — Robot Trader 2026 — Screener AZIONI\n\n"
            f"Errori di rete rilevati durante lo screening:\n"
            f"  Errori rete: {network_errors} su {total} ({soglia_pct}%)\n"
            f"  Errori totali: {non_validi_total}\n\n"
            f"Causa probabile: DNS non risolveva query2.finance.yahoo.com al momento del run.\n\n"
            f"AZIONE RICHIESTA — rilanciare lo screener:\n"
            f"  cd C:\\Users\\lucia\\Desktop\\Robot Trader 2026\\PYTHON_SCRIPTS\n"
            f"  python orchestrator.py AZIONI\n\n"
            f"— Robot Trader 2026"
        )
        msg = MIMEText(body, 'plain', 'utf-8')
        msg['Subject'] = f"[ALERT] Screener AZIONI: {network_errors} errori rete su {total} ({soglia_pct}%)"
        msg['From']    = sender
        msg['To']      = login
        with smtplib.SMTP(host, port, timeout=15) as srv:
            srv.ehlo(); srv.starttls()
            srv.login(login, pwd)
            srv.sendmail(sender, [login], msg.as_string())
        print(f"[ALERT] Email alert inviata a {login}", flush=True)
    except Exception as e:
        print(f"[ALERT] Errore invio alert email: {e}", flush=True)

try:
    import yfinance as yf
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError as e:
    print(f"❌ ERRORE: Libreria mancante - {e}")
    sys.exit(1)

from screener_utils import batch_percentile_score
from ticker_lists_5000 import (
    ALL_AZIONI,
    USA_SP500, USA_MIDCAP, USA_SMALLCAP_SP600, USA_RUSSELL2000,
    UK_FTSE100, UK_FTSE250,
    GERMANY_DAX, GERMANY_MDAX,
    FRANCE_CAC40, FRANCE_MIDCAP,
    ITALY_MIB, SPAIN_IBEX, SWISS_SMI, NETHERLANDS_AEX,
    SWEDEN_OMXS, NORWAY_OBX, DENMARK_OMXC, FINLAND_OMXH,
    BELGIUM_BEL20, AUSTRIA_ATX, PORTUGAL_PSI,
    JAPAN_NIKKEI, JAPAN_TOPIX_EXT,
    HK_HANGSENG, KOREA_KOSPI, BRAZIL_IBOVESPA,
    INDIA_NIFTY, TAIWAN_TWSE,
    AUSTRALIA_ASX, AUSTRALIA_ASX_EXT,
    CANADA_TSX, CANADA_TSX_EXT,
)

# Dizionario inverso ticker → nome indice (primo match vince)
_INDEX_MAP = [
    (USA_SP500,          'S&P 500'),
    (USA_MIDCAP,         'S&P 400'),
    (USA_SMALLCAP_SP600, 'S&P 600'),
    (USA_RUSSELL2000,    'Russell 2000'),
    (UK_FTSE100,         'FTSE 100'),
    (UK_FTSE250,         'FTSE 250'),
    (GERMANY_DAX,        'DAX 40'),
    (GERMANY_MDAX,       'MDAX'),
    (FRANCE_CAC40,       'CAC 40'),
    (FRANCE_MIDCAP,      'SBF MidCap'),
    (ITALY_MIB,          'FTSE MIB'),
    (SPAIN_IBEX,         'IBEX 35'),
    (SWISS_SMI,          'SMI'),
    (NETHERLANDS_AEX,    'AEX'),
    (SWEDEN_OMXS,        'OMXS 30'),
    (NORWAY_OBX,         'OBX'),
    (DENMARK_OMXC,       'OMXC 25'),
    (FINLAND_OMXH,       'OMXH 25'),
    (BELGIUM_BEL20,      'BEL 20'),
    (AUSTRIA_ATX,        'ATX'),
    (PORTUGAL_PSI,       'PSI'),
    (JAPAN_NIKKEI,       'Nikkei 225'),
    (JAPAN_TOPIX_EXT,    'TOPIX'),
    (HK_HANGSENG,        'Hang Seng'),
    (KOREA_KOSPI,        'KOSPI'),
    (BRAZIL_IBOVESPA,    'IBOVESPA'),
    (INDIA_NIFTY,        'NIFTY 500'),
    (TAIWAN_TWSE,        'TWSE'),
    (AUSTRALIA_ASX,      'ASX 200'),
    (AUSTRALIA_ASX_EXT,  'ASX Ext'),
    (CANADA_TSX,         'TSX'),
    (CANADA_TSX_EXT,     'TSX Ext'),
]
TICKER_TO_INDICE = {}
for _tlist, _name in _INDEX_MAP:
    for _t in _tlist:
        if _t not in TICKER_TO_INDICE:
            TICKER_TO_INDICE[_t] = _name

# DIRECTORY OUTPUT
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(os.path.dirname(BASE_DIR), "REPORTS_DAILY")
PARAMETRI_FILE = os.path.join(BASE_DIR, "parametri.json")
os.makedirs(REPORTS_DIR, exist_ok=True)

# CARICA PARAMETRI DA JSON
def _get_param(params: dict, key: str, default):
    """Legge un parametro che può essere {'value': X} oppure X direttamente."""
    val = params.get(key, default)
    if isinstance(val, dict):
        return val.get('value', default)
    return val if val is not None else default


def load_filters():
    """Carica filtri da parametri.json"""
    try:
        with open(PARAMETRI_FILE, 'r', encoding='utf-8') as f:
            parametri = json.load(f)
        azioni_params = parametri.get('azioni', {})
        return {
            'ev_fcf_max':          _get_param(azioni_params, 'ev_fcf_max',          12),
            'price_book_max':      _get_param(azioni_params, 'price_book_max',      1.2),
            'roe_min':             _get_param(azioni_params, 'roe_min',             0),
            'net_debt_ebitda_max': _get_param(azioni_params, 'net_debt_ebitda_max', 2.5),
        }
    except Exception as e:
        print(f"⚠️ ERRORE caricamento parametri: {e}")
        print("⚠️ Uso valori di default")
        return {
            'ev_fcf_max': 12,
            'price_book_max': 1.2,
            'roe_min': 0,
            'net_debt_ebitda_max': 2.5
        }

HARD_FILTERS = load_filters()
EXCLUDED_SECTORS = ['Financial Services']
MARKET_CAP_MIN_USD = 100_000_000  # 100M USD

# ---------------------------------------------------------------------------
# Universo BASIC: indici principali globali (~1000 ticker)
# ---------------------------------------------------------------------------
BASIC_UNIVERSE = frozenset(
    USA_SP500 + UK_FTSE100 + GERMANY_DAX + FRANCE_CAC40 +
    ITALY_MIB + SPAIN_IBEX + SWISS_SMI + NETHERLANDS_AEX + JAPAN_NIKKEI
)

# ---------------------------------------------------------------------------
# Configurazioni per piano (filtri + ampiezza output)
# ---------------------------------------------------------------------------
SERVIZI_FILE = os.path.join(BASE_DIR, 'servizi_config.json')

def _build_azioni_plan_configs():
    """Legge i parametri filtro da servizi_config.json (modificabili dalla dashboard).
    Mappatura nomi: price_book_max→pb_max, net_debt_ebitda_max→nd_max.
    roe_min in servizi è in % intera (es. 1 = 1%) → converti in decimale (/100).
    universe e top_n restano hardcoded.
    """
    _fixed    = {
        'BASIC': {'universe': None,           'top_n': 20},
        'PRO':   {'universe': None,           'top_n': 50},
        'VALUE': {'universe': None,           'top_n': 50},
    }
    _fallback = {
        'BASIC': {'ev_fcf_max': 12.0, 'pb_max': 1.2, 'roe_min': 0.0,   'nd_max': 2.5},
        'PRO':   {'ev_fcf_max': 12.0, 'pb_max': 1.2, 'roe_min': 0.0,   'nd_max': 2.5},
        'VALUE': {'ev_fcf_max': 15.0, 'pb_max': 1.5, 'roe_min': -0.05, 'nd_max': 3.0},
    }
    try:
        with open(SERVIZI_FILE, encoding='utf-8') as f:
            sv = json.load(f).get('azioni', {})
        configs = {}
        for plan, key in [('BASIC','basic'), ('PRO','pro'), ('VALUE','value')]:
            p = sv.get(key, {}).get('parametri', {})
            cfg = dict(_fallback[plan])
            if p:
                if 'ev_fcf_max'          in p: cfg['ev_fcf_max'] = float(p['ev_fcf_max'])
                if 'price_book_max'      in p: cfg['pb_max']     = float(p['price_book_max'])
                if 'roe_min'             in p: cfg['roe_min']    = float(p['roe_min']) / 100
                if 'net_debt_ebitda_max' in p: cfg['nd_max']     = float(p['net_debt_ebitda_max'])
            cfg.update(_fixed[plan])
            configs[plan] = cfg
        return configs
    except Exception as e:
        print(f"⚠️ servizi_config.json non leggibile ({e}) — uso valori di default")
        return {pl: {**_fallback[pl], **_fixed[pl]} for pl in _fallback}

PLAN_CONFIGS = _build_azioni_plan_configs()

# Tassi approssimativi verso USD — aggiornare trimestralmente
CURRENCY_TO_USD = {
    'USD': 1.0,
    'EUR': 1.10,
    'GBP': 1.27,
    'GBp': 0.0127,   # pence → USD (÷100 poi ×1.27)
    'CHF': 1.13,
    'JPY': 0.0065,
    'HKD': 0.128,
    'SEK': 0.094,
    'NOK': 0.090,
    'DKK': 0.148,
    'CAD': 0.735,
    'AUD': 0.645,
    'KRW': 0.00073,
    'CNY': 0.138,
    'INR': 0.012,
}

def convert_market_cap_to_usd(market_cap, currency):
    """Converte market cap in USD per confronto uniforme tra mercati"""
    if market_cap is None:
        return None
    rate = CURRENCY_TO_USD.get(currency, 1.0)
    return market_cap * rate

def format_percent_ita(value, decimals=1):
    """Formatta percentuale con virgola italiana"""
    if value is None or (isinstance(value, float) and (value != value)):
        return 'N/A'
    return f"{value*100:.{decimals}f}%".replace('.', ',')

def _fmt_market_cap(mc):
    """Formatta Market Cap come stringa leggibile (es. 310M, 1.2B, 2.5T)"""
    if mc is None:
        return 'N/A'
    try:
        mc = float(mc)
    except (TypeError, ValueError):
        return 'N/A'
    if mc >= 1e12:
        return f"{mc/1e12:.1f}T"
    if mc >= 1e9:
        return f"{mc/1e9:.1f}B"
    if mc >= 1e6:
        return f"{mc/1e6:.0f}M"
    return f"{int(mc):,}"

print("="*70)
print("VALUE SCREENER AZIONI - Deep Value / Undervalued")
print("="*70)
print(f"Ticker da analizzare: {len(ALL_AZIONI)}")
print(f"Avvio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print("="*70)

def get_market_from_ticker(ticker):
    """Determina mercato da ticker"""
    if ticker.endswith('.DE'):
        return 'DAX 40'
    elif ticker.endswith('.PA'):
        return 'CAC 40'
    elif ticker.endswith('.L'):
        return 'FTSE 100'
    elif ticker.endswith('.MI'):
        return 'FTSE MIB'
    elif ticker.endswith('.MC'):
        return 'IBEX 35'
    elif ticker.endswith('.AS'):
        return 'AEX (Amsterdam)'
    elif ticker.endswith('.SW') or ticker.endswith('.VX'):
        return 'SMI (Svizzera)'
    elif ticker.endswith('.ST'):
        return 'OMX Stockholm'
    elif ticker.endswith('.OL'):
        return 'Oslo Bors'
    elif ticker.endswith('.CO'):
        return 'OMX Copenhagen'
    elif ticker.endswith('.HE'):
        return 'OMX Helsinki'
    elif ticker.endswith('.T'):
        return 'Tokyo (TSE)'
    elif ticker.endswith('.HK'):
        return 'Hong Kong (HKEX)'
    elif ticker.endswith('.AX'):
        return 'ASX (Australia)'
    elif ticker.endswith('.TO'):
        return 'TSX (Canada)'
    elif ticker.endswith('.BR'):
        return 'BEL 20 (Belgio)'
    elif ticker.endswith('.VI'):
        return 'ATX (Austria)'
    elif ticker.endswith('.LS'):
        return 'PSI (Portogallo)'
    elif ticker.endswith('.KS'):
        return 'KOSPI (Corea)'
    elif ticker.endswith('.SA'):
        return 'IBOVESPA (Brasile)'
    elif ticker.endswith('.NS'):
        return 'NSE (India)'
    elif ticker.endswith('.TW'):
        return 'TWSE (Taiwan)'
    else:
        return 'USA (S&P/NASDAQ)'

def get_stock_data(ticker):
    """Scarica dati stock da yfinance - versione completa"""
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        
        # --- EV e FCF con fallback a cascata ---
        ev  = info.get('enterpriseValue')
        fcf = info.get('freeCashflow')

        # Fallback 1: OCF + CAPEX da info (disponibile anche per ticker non-US)
        if not fcf:
            ocf_i  = info.get('operatingCashflow') or info.get('operatingCashFlow')
            cap_i  = info.get('capitalExpenditures')
            if ocf_i is not None and cap_i is not None:
                fcf = float(ocf_i) + float(cap_i)   # capex già negativo

        # Fallback 2: cashflow statement annuale
        if not fcf:
            try:
                cf = stock.cashflow
                if cf is not None and not cf.empty:
                    for rn in cf.index:
                        if 'free cash flow' in rn.lower():
                            v = cf.loc[rn].iloc[0]
                            if pd.notna(v):
                                fcf = float(v)
                            break
                    if not fcf:
                        ocf, capex = None, None
                        for rn in cf.index:
                            if 'operating cash flow' in rn.lower():
                                v = cf.loc[rn].iloc[0]
                                if pd.notna(v): ocf = float(v)
                            if 'capital expenditure' in rn.lower():
                                v = cf.loc[rn].iloc[0]
                                if pd.notna(v): capex = float(v)
                        if ocf is not None and capex is not None:
                            fcf = ocf + capex
            except Exception:
                pass

        # Fallback 3: cashflow trimestrale annualizzato
        if not fcf:
            try:
                qcf = stock.quarterly_cashflow
                if qcf is not None and not qcf.empty:
                    ocf, capex = None, None
                    for rn in qcf.index:
                        if 'operating cash flow' in rn.lower():
                            v = qcf.loc[rn].iloc[:4].sum()
                            if pd.notna(v): ocf = float(v)
                        if 'capital expenditure' in rn.lower():
                            v = qcf.loc[rn].iloc[:4].sum()
                            if pd.notna(v): capex = float(v)
                    if ocf is not None and capex is not None:
                        fcf = ocf + capex
            except Exception:
                pass

        # Fallback EV: Market Cap + Debt - Cash
        if not ev:
            mc = info.get('marketCap')
            td = info.get('totalDebt', 0) or 0
            tc = info.get('totalCash', 0) or 0
            if mc:
                ev = mc + td - tc

        # EV/FCF — calcolato anche se negativo (FCF < 0 → ev_fcf negativo → "Scartate" non "Dati Mancanti")
        if ev is not None and fcf is not None and fcf != 0:
            ev_fcf = round(ev / fcf, 2)
        else:
            ev_fcf = None
        
        # --- Net Debt / EBITDA (CORRETTO: NON è debtToEquity!) ---
        total_debt = info.get('totalDebt')
        total_cash = info.get('totalCash')
        ebitda = info.get('ebitda')
        
        net_debt = None
        nd_ebitda = None
        if total_debt is not None and total_cash is not None:
            net_debt = total_debt - total_cash
            if ebitda and ebitda > 0:
                nd_ebitda = round(net_debt / ebitda, 2)
        
        # P/B — Layer 1: yfinance info(), Layer 2: balance sheet
        pb = info.get('priceToBook')
        if pb is None or pb <= 0:
            try:
                bs = stock.balance_sheet
                if bs is not None and not bs.empty:
                    eq_row = next((r for r in bs.index if any(k in r.lower() for k in
                                   ['stockholders equity', 'total equity', 'shareholders equity'])), None)
                    if eq_row:
                        equity = float(bs.loc[eq_row].iloc[0])
                        mc = info.get('marketCap')
                        if mc and equity > 0:
                            pb = round(mc / equity, 4)
            except Exception:
                pass

        # ROE — Layer 1: yfinance info(), Layer 2: financials + balance sheet
        roe = info.get('returnOnEquity')
        if roe is None:
            try:
                inc = stock.financials
                bs_roe = stock.balance_sheet
                if inc is not None and not inc.empty and bs_roe is not None and not bs_roe.empty:
                    ni_row = next((r for r in inc.index if 'net income' in r.lower()), None)
                    eq_row = next((r for r in bs_roe.index if any(k in r.lower() for k in
                                   ['stockholders equity', 'total equity', 'shareholders equity'])), None)
                    if ni_row and eq_row:
                        ni = float(inc.loc[ni_row].iloc[0])
                        eq = float(bs_roe.loc[eq_row].iloc[0])
                        if eq != 0:
                            roe = round(ni / eq, 4)
            except Exception:
                pass

        # Fix GBp (pence) → GBP (sterline) per titoli London Stock Exchange
        currency = info.get('currency', 'USD')
        prezzo = info.get('currentPrice', info.get('regularMarketPrice'))
        if currency == 'GBp' and prezzo is not None:
            prezzo = prezzo / 100  # pence → sterline per display corretto
            valuta = 'GBP'
        else:
            valuta = currency

        # Var_1D_%: fonte primaria = regularMarketChangePercent (prezzo puro, no div adjustment)
        # Necessario per .KS / .T e altri mercati dove auto_adjust altera la variazione giornaliera
        perf_1m = perf_3m = perf_6m = perf_ytd = perf_1y = var_1d = None
        var_1d_raw = info.get('regularMarketChangePercent')
        if var_1d_raw is not None:
            try:
                var_1d = round(float(var_1d_raw), 2)  # già in % — no * 100
            except (TypeError, ValueError):
                var_1d = None

        # Performance 1M-3M-6M-YTD-1Y da storico prezzi PURI (no aggiustamento dividendi)
        # auto_adjust=False + actions=False evita distorsioni ex-dividend per mercati asiatici
        def _safe_perf(p_new, p_old):
            if p_old and p_old != 0:
                v = round(((p_new / p_old) - 1) * 100, 2)
                return v if abs(v) < 500 else None  # sanity: > 500% è artifact dati
            return None

        try:
            hist = stock.history(period='1y', interval='1d', auto_adjust=False, actions=False)
            if hist is not None and not hist.empty:
                close_h = hist['Close'].dropna()
                nh = len(close_h)
                if nh >= 2:
                    p_now  = close_h.iloc[-1]
                    # var_1d fallback se regularMarketChangePercent non disponibile
                    if var_1d is None:
                        p_prev = close_h.iloc[-2]
                        var_1d = _safe_perf(p_now, p_prev)
                    idx_1m = max(0, nh - 22)
                    idx_3m = max(0, nh - 63)
                    idx_6m = max(0, nh - 126)
                    if idx_1m < nh - 1:
                        perf_1m = _safe_perf(p_now, close_h.iloc[idx_1m])
                    if idx_3m < nh - 1:
                        perf_3m = _safe_perf(p_now, close_h.iloc[idx_3m])
                    if idx_6m < nh - 1:
                        perf_6m = _safe_perf(p_now, close_h.iloc[idx_6m])
                    perf_1y = _safe_perf(p_now, close_h.iloc[0])
                    # YTD: robusto per indici tz-aware (mercati asiatici)
                    current_year = datetime.now().year
                    ytd_start = pd.Timestamp(f'{current_year}-01-01')
                    if hist.index.tz is not None:
                        ytd_start = ytd_start.tz_localize(hist.index.tz)
                    hist_ytd_close = hist.loc[hist.index >= ytd_start, 'Close'].dropna()
                    if not hist_ytd_close.empty:
                        perf_ytd = _safe_perf(p_now, hist_ytd_close.iloc[0])
        except Exception:
            pass

        return {
            'Ticker': ticker,
            'Nome': info.get('longName', info.get('shortName', ticker)),
            'Valuta': valuta,
            'Settore': info.get('sector', 'N/A'),
            'Industry': info.get('industry', 'N/A'),
            'Market Cap': info.get('marketCap'),
            'Currency': currency,
            'Prezzo': prezzo,
            'Var_1D_%': var_1d,
            'P/B': pb,
            'ROE': roe,
            'EV/EBITDA': info.get('enterpriseToEbitda'),
            'Free Cash Flow': fcf,
            'Enterprise Value': ev,
            'Total Debt': total_debt,
            'Total Cash': total_cash,
            'EBITDA': ebitda,
            'Dividend Yield': info.get('dividendYield'),
            'Analyst Coverage': info.get('numberOfAnalystOpinions', 0),
            'EV/FCF': ev_fcf,
            'Net Debt': net_debt,
            'Net Debt/EBITDA': nd_ebitda,
            'Perf_1M_%':  perf_1m,
            'Perf_3M_%':  perf_3m,
            'Perf_6M_%':  perf_6m,
            'Perf_YTD_%': perf_ytd,
            'Perf_1Y_%':  perf_1y,
            'Data Dati': datetime.now().strftime('%Y-%m-%d')
        }
    except Exception as e:
        return {'Ticker': ticker, 'Error': str(e)}

def calculate_score(stock):
    # legacy — sostituita da batch_percentile_score (chiamata dopo la selezione)
    return 0

def apply_plan_filters(stock, cfg):
    """Filtri specifici per piano — parametri passati via cfg dict."""
    market_cap = stock.get('Market Cap')
    ev_fcf     = stock.get('EV/FCF')
    pb         = stock.get('P/B')
    roe        = stock.get('ROE')
    missing = []
    if market_cap is None: missing.append('Market Cap')
    if ev_fcf is None:     missing.append('EV/FCF')
    if pb is None:         missing.append('P/B')
    if roe is None:        missing.append('ROE')
    if missing:
        return 'missing', 'Dati N/A: ' + ', '.join(missing)
    if stock.get('Settore') in EXCLUDED_SECTORS:
        return 'rejected', f"Settore escluso: {stock.get('Settore')}"
    currency = stock.get('Currency', 'USD')
    mc_usd = convert_market_cap_to_usd(market_cap, currency)
    if mc_usd is None or mc_usd < MARKET_CAP_MIN_USD:
        return 'rejected', f"Market Cap {mc_usd/1e6:.0f}M USD < 100M" if mc_usd else "Market Cap N/A"
    if ev_fcf <= 0:
        return 'rejected', f"FCF negativo (EV/FCF {ev_fcf:.2f})"
    if ev_fcf > cfg['ev_fcf_max']:
        return 'rejected', f"EV/FCF {ev_fcf:.2f} > {cfg['ev_fcf_max']}"
    if pb <= 0 or pb > cfg['pb_max']:
        return 'rejected', f"P/B {pb:.2f} > {cfg['pb_max']}"
    if roe < cfg['roe_min']:
        return 'rejected', f"ROE {format_percent_ita(roe,1)} < {format_percent_ita(cfg['roe_min'],1)}"
    nd = stock.get('Net Debt/EBITDA')
    if nd is not None and nd > cfg['nd_max']:
        return 'rejected', f"Net Debt/EBITDA {nd:.2f} > {cfg['nd_max']}"
    return 'ok', None

_HIDDEN_KEYS = {'Currency'}

_TOP_COLS_BASIC = ['Ticker','Nome','Mercato','Indice','Score',
                   'P/B','ROE','EV/FCF','Net Debt/EBITDA','Settore']
_TOP_COLS_FULL  = ['Ticker','Nome','Valuta','Mercato','Indice','Prezzo','Var 1D %','Score',
                   'Perf 1M %','Perf 3M %','Perf 6M %','Perf YTD %','Perf 1Y %',
                   'P/B','ROE','EV/FCF','Net Debt/EBITDA','Market Cap','Settore']

_SEL_COLS = [
    'Ticker','Nome','Valuta','Settore','Industry','Mercato','Indice',
    'Market Cap','P/B','ROE','EV/EBITDA','Free Cash Flow','Enterprise Value',
    'Total Debt','Total Cash','EBITDA','Dividend Yield','Analyst Coverage',
    'EV/FCF','Net Debt','Net Debt/EBITDA','Prezzo','Var_1D_%',
    'Perf_1M_%','Perf_3M_%','Perf_6M_%','Perf_YTD_%','Perf_1Y_%',
    'Data Dati','Score'
]

def _fmt_sel_val(stock, col):
    """Formatta valori per il foglio Azioni Selezionate — numeri leggibili."""
    v = stock.get(col)
    if v is None:
        return 'N/A'
    if col in {'Market Cap','Free Cash Flow','Enterprise Value',
               'Total Debt','Total Cash','EBITDA','Net Debt'}:
        return _fmt_market_cap(v)
    if col in {'Var_1D_%','Perf_1M_%','Perf_3M_%','Perf_6M_%','Perf_YTD_%','Perf_1Y_%'}:
        return f"{v:.2f}%".replace('.', ',') if isinstance(v, (int, float)) else 'N/A'
    if col in {'ROE','Dividend Yield'}:
        return format_percent_ita(v, 1) if isinstance(v, (int, float)) else 'N/A'
    if col in {'P/B','EV/EBITDA','EV/FCF','Net Debt/EBITDA'}:
        return round(float(v), 2) if isinstance(v, (int, float)) else 'N/A'
    if col == 'Prezzo':
        return round(float(v), 2) if isinstance(v, (int, float)) else 'N/A'
    if col == 'Score':
        return round(float(v), 1) if isinstance(v, (int, float)) else 0
    return v

def _stock_val(stock, col):
    """Estrae e formatta il valore di una colonna per il foglio Top."""
    m = {
        'Ticker':          lambda s: s.get('Ticker'),
        'Nome':            lambda s: s.get('Nome','N/A'),
        'Valuta':          lambda s: s.get('Valuta','N/A'),
        'Mercato':         lambda s: s.get('Mercato','N/A'),
        'Indice':          lambda s: s.get('Indice','N/A'),
        'Prezzo':          lambda s: round(s['Prezzo'],4) if s.get('Prezzo') else 'N/A',
        'Var 1D %':        lambda s: s.get('Var_1D_%'),
        'Score':           lambda s: s.get('Score',0),
        'Perf 1M %':       lambda s: s.get('Perf_1M_%'),
        'Perf 3M %':       lambda s: s.get('Perf_3M_%'),
        'Perf 6M %':       lambda s: s.get('Perf_6M_%'),
        'Perf YTD %':      lambda s: s.get('Perf_YTD_%'),
        'Perf 1Y %':       lambda s: s.get('Perf_1Y_%'),
        'P/B':             lambda s: round(s['P/B'],2) if s.get('P/B') else 'N/A',
        'ROE':             lambda s: format_percent_ita(s['ROE'],1) if s.get('ROE') else 'N/A',
        'EV/FCF':          lambda s: round(s['EV/FCF'],2) if s.get('EV/FCF') else 'N/A',
        'Net Debt/EBITDA': lambda s: round(s['Net Debt/EBITDA'],2) if s.get('Net Debt/EBITDA') else 'N/A',
        'Market Cap':      lambda s: _fmt_market_cap(s.get('Market Cap')),
        'Settore':         lambda s: s.get('Settore','N/A'),
    }
    return m.get(col, lambda s: s.get(col))(stock)

def _write_azioni_plan_excel(plan_name, cfg, selected, rejected, non_validi, errors, mkt_stats, timestamp):
    wb = Workbook()
    wb.remove(wb.active)
    HDR_FILL  = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    SEC_FONT  = Font(bold=True, size=11, color="2C5282")
    GRAY_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    top_n    = cfg['top_n']
    top_cols = _TOP_COLS_BASIC if plan_name == 'BASIC' else _TOP_COLS_FULL

    # ── PESI SCORE per piano (per la legenda) ────────────────────────────
    _pesi_label = {
        'BASIC': [("Dividend Yield", 35), ("Var 1D %", 25), ("ROE", 20), ("EV/FCF", 10), ("P/B", 10)],
        'PRO':   [("EV/FCF", 35), ("ROE", 25), ("P/B", 20), ("Net Debt/EBITDA", 15), ("Var 1D %", 5)],
        'VALUE': [("EV/FCF", 40), ("ROE", 25), ("Net Debt/EBITDA", 20), ("P/B", 15)],
    }

    # ── FOGLIO 0: Legenda ────────────────────────────────────────────────
    ws_leg = wb.create_sheet("📖 Legenda", 0)
    ws_leg.column_dimensions['A'].width = 26
    ws_leg.column_dimensions['B'].width = 80

    def _leg(r, a, b='', bold_a=False, fill=None, font_a=None):
        ca = ws_leg.cell(row=r, column=1, value=a)
        cb = ws_leg.cell(row=r, column=2, value=b)
        if bold_a:
            ca.font = font_a or Font(bold=True)
        if fill:
            ca.fill = fill
            cb.fill = fill
        ca.alignment = Alignment(vertical='top', wrap_text=True)
        cb.alignment = Alignment(vertical='top', wrap_text=True)
        return r + 1

    r = 1
    ws_leg.merge_cells('A1:B1')
    c = ws_leg.cell(row=1, column=1,
                    value=f"GUIDA AL REPORT — COME LEGGERE I DATI  |  Piano {plan_name}  |  Robot Trader 2026")
    c.font  = Font(size=14, bold=True, color="FFFFFF")
    c.fill  = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_leg.row_dimensions[1].height = 28
    r = 2

    r = _leg(r, '')
    ws_leg.merge_cells(f'A{r}:B{r}')
    ws_leg.cell(row=r, column=1, value="LO SCORE (0 – 100)").font = SEC_FONT
    r += 1
    r = _leg(r, 'Cos\'è lo Score',
             'Un punteggio da 0 a 100 che sintetizza la qualità dell\'azione per questo piano. '
             'PIÙ ALTO = MIGLIORE. Le azioni nel foglio "Top per Score" sono ordinate dal punteggio più alto al più basso.',
             bold_a=True)
    r = _leg(r, 'Come si calcola',
             'Per ogni indicatore, l\'azione viene confrontata con tutte le altre selezionate: '
             'riceve un percentile 0-100. I percentili vengono pesati e sommati.',
             bold_a=True)
    r += 1

    pesi = _pesi_label.get(plan_name, [])
    if pesi:
        ws_leg.merge_cells(f'A{r}:B{r}')
        ws_leg.cell(row=r, column=1, value=f"Pesi dello Score — Piano {plan_name}").font = SEC_FONT
        r += 1
        for metric, peso in pesi:
            r = _leg(r, f"  {metric}", f"{peso}% del punteggio finale", bold_a=False, fill=GRAY_FILL)
        r += 1

    ws_leg.merge_cells(f'A{r}:B{r}')
    ws_leg.cell(row=r, column=1, value="INDICATORI — COSA SIGNIFICANO").font = SEC_FONT
    r += 1

    indicatori = [
        ("P/B  (Price / Book)",
         "Prezzo dell'azione diviso il valore contabile del patrimonio netto.\n"
         "Esempio: P/B = 0.8 → l'azione vale meno del patrimonio (sottovalutata).\n"
         "MEGLIO SE BASSO. Filtro: P/B ≤ " + str(cfg['pb_max'])),
        ("ROE  (Return on Equity)",
         "Rendimento sul capitale proprio: utile netto / patrimonio.\n"
         "Esempio: ROE = 8% → per ogni 100€ di patrimonio, l'azienda guadagna 8€.\n"
         "MEGLIO SE ALTO. Filtro: ROE ≥ " + str(round(cfg['roe_min']*100,1)) + "%"),
        ("EV / FCF",
         "Enterprise Value diviso Free Cash Flow.\n"
         "Indica quanti anni di cash flow attuale vale l'intera azienda (debiti inclusi).\n"
         "Esempio: EV/FCF = 8 → l'azienda 'si ripaga' in 8 anni di cassa.\n"
         "MEGLIO SE BASSO. Filtro: EV/FCF ≤ " + str(cfg['ev_fcf_max'])),
        ("Net Debt / EBITDA",
         "Leva finanziaria: debito netto diviso EBITDA (utile operativo + ammortamenti).\n"
         "Indica quanti anni di EBITDA servono a ripagare il debito netto.\n"
         "Esempio: 1.5 → debito = 1,5 anni di EBITDA. Valori negativi = cassa netta (ottimo).\n"
         "MEGLIO SE BASSO. Filtro: Net Debt/EBITDA ≤ " + str(cfg['nd_max'])),
        ("Market Cap",
         "Capitalizzazione di borsa = Prezzo × Azioni in circolazione.\n"
         "M = milioni  |  B = miliardi  |  T = trilioni (in valuta locale).\n"
         "Filtro minimo: 100M USD equivalenti."),
        ("Var 1D %",
         "Variazione percentuale del prezzo rispetto alla chiusura del giorno precedente."),
        ("Perf 1M / 3M / 6M / YTD / 1Y %",
         "Performance del prezzo nell'ultimo mese / trimestre / semestre / da inizio anno / ultimo anno.\n"
         "Calcolata su prezzi puri senza aggiustamento dividendi."),
        ("Dividend Yield",
         "Dividendo annuo / Prezzo corrente. Es. 3% → ogni 100€ investiti, 3€ di dividendo annuo.\n"
         "Solo per il piano BASIC: peso 35%."),
        ("EV/EBITDA",
         "Enterprise Value diviso EBITDA. Multiplo di valutazione alternativo.\n"
         "Non usato nei filtri, presente nel foglio 'Azioni Selezionate' come dato aggiuntivo."),
    ]
    for nome, desc in indicatori:
        r = _leg(r, nome, desc, bold_a=True, fill=GRAY_FILL if indicatori.index((nome,desc)) % 2 == 0 else None)

    r += 1
    ws_leg.merge_cells(f'A{r}:B{r}')
    ws_leg.cell(row=r, column=1, value="COME LEGGERE I FOGLI DEL REPORT").font = SEC_FONT
    r += 1
    fogli = [
        ("📖 Legenda (questo foglio)", "Guida alla lettura del report."),
        ("Dashboard",
         "Riepilogo esecutivo: filtri applicati, conteggi, breakdown per mercato."),
        (f"Top {top_n} per Score",
         f"Le {top_n} azioni migliori ordinate per Score (dalla più alta alla più bassa). Qui si trovano le opportunità migliori del piano."),
        ("Azioni Selezionate",
         "Tutte le azioni che hanno superato i filtri con tutti i dati disponibili."),
        ("Esclusi - Settore",
         "Azioni escluse a priori perché appartengono a Financial Services (settore escluso per natura dei dati)."),
        ("Scartate - EV-FCF / P-B / ROE / Net Debt / Market Cap",
         "Azioni che non hanno superato uno specifico filtro. La colonna 'Motivo Scarto' indica il valore che ha causato lo scarto."),
        ("Dati Mancanti",
         "Azioni per cui uno o più dati fondamentali non erano disponibili su Yahoo Finance."),
    ]
    for nome_f, desc_f in fogli:
        r = _leg(r, nome_f, desc_f, bold_a=True, fill=GRAY_FILL if fogli.index((nome_f,desc_f)) % 2 == 0 else None)

    r += 1
    ws_leg.merge_cells(f'A{r}:B{r}')
    disc = ws_leg.cell(row=r, column=1,
                       value="⚠  DISCLAIMER — Questo report è generato automaticamente a scopo esclusivamente informativo. "
                             "Non costituisce consulenza finanziaria né raccomandazione di acquisto/vendita di strumenti finanziari. "
                             "Gli investimenti comportano rischi, inclusa la perdita del capitale. "
                             "Consulta un consulente finanziario abilitato prima di qualsiasi decisione.")
    disc.font = Font(italic=True, color="888888", size=9)
    disc.alignment = Alignment(wrap_text=True)
    ws_leg.row_dimensions[r].height = 48

    # ── FOGLIO 1: Dashboard ──────────────────────────────────────────────
    ws = wb.create_sheet("Dashboard", 1)
    ws['A1'] = f"VALUE SCREENER AZIONI — Piano {plan_name} — Robot Trader 2026"
    ws['A1'].font = Font(size=14, bold=True, color="FF8C42")
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"FILTRI PIANO {plan_name}"
    ws['A4'].font = Font(bold=True, size=12)
    for r, (lbl, val) in enumerate([
        ("EV/FCF max",        cfg['ev_fcf_max']),
        ("P/B max",           cfg['pb_max']),
        ("ROE min",           f"{cfg['roe_min']*100:.0f}%"),
        ("Net Debt/EBITDA max", cfg['nd_max']),
        ("Top N",             top_n),
        ("Universo",          len(selected)+len(rejected)+len(non_validi)),
        ("Selezionate",       len(selected)),
        ("Scartate",          len(rejected)),
        ("Dati N/A",          len(non_validi)),
        ("Errori fetch",      len(errors)),
    ], start=5):
        ws[f'A{r}'] = lbl; ws[f'A{r}'].font = Font(bold=True)
        ws[f'B{r}'] = val
    r += 2
    ws[f'A{r}'] = "BREAKDOWN PER MERCATO"; ws[f'A{r}'].font = Font(bold=True, size=12); r += 1
    for col, h in enumerate(['Mercato','Totali','Selezionate','Scartate','N/A'], 1):
        c = ws.cell(row=r, column=col, value=h); c.font = HDR_FONT; c.fill = HDR_FILL
    r += 1
    for mkt, ms in sorted(mkt_stats.items()):
        ws.cell(row=r, column=1, value=mkt).font = Font(bold=True)
        ws.cell(row=r, column=2, value=ms['totali'])
        ws.cell(row=r, column=3, value=ms['selezionate'])
        ws.cell(row=r, column=4, value=ms['scartate'])
        ws.cell(row=r, column=5, value=ms['non_validi'])
        r += 1

    # ── FOGLIO 2: Top N per Score ────────────────────────────────────────
    top_list = sorted(selected, key=lambda x: x.get('Score', 0), reverse=True)[:top_n]
    ws_top = wb.create_sheet(f"Top {top_n} per Score", 2)
    SCORE_FILL = PatternFill(start_color="F6AD55", end_color="F6AD55", fill_type="solid")
    SCORE_COL  = top_cols.index('Score') + 1  # colonna Score (1-based)

    if top_list:
        # Riga intestazione
        for col, h in enumerate(top_cols, 1):
            c = ws_top.cell(row=1, column=col, value=h)
            c.font = HDR_FONT
            c.fill = SCORE_FILL if h == 'Score' else HDR_FILL
        # Riga sub-intestazione con spiegazione breve
        sub = {
            'Ticker': 'Codice borsa', 'Nome': 'Ragione sociale', 'Valuta': 'Divisa',
            'Mercato': 'Mercato/Borsa', 'Indice': 'Indice di riferimento',
            'Prezzo': 'Prezzo corrente', 'Var 1D %': 'Var. giorno %',
            'Score': '0-100 ↑ meglio',
            'Perf 1M %': 'Perf. 1 mese', 'Perf 3M %': 'Perf. 3 mesi',
            'Perf 6M %': 'Perf. 6 mesi', 'Perf YTD %': 'Perf. da gen.',
            'Perf 1Y %': 'Perf. 1 anno',
            'P/B': 'Prezzo/Patrim. ↓', 'ROE': 'Rend. capital. ↑',
            'EV/FCF': 'Val./CashFlow ↓', 'Net Debt/EBITDA': 'Leva fin. ↓',
            'Market Cap': 'Cap. di borsa', 'Settore': 'Settore GICS',
        }
        sub_fill = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")
        for col, h in enumerate(top_cols, 1):
            c = ws_top.cell(row=2, column=col, value=sub.get(h, ''))
            c.font = Font(italic=True, size=8, color="555555")
            c.fill = SCORE_FILL if h == 'Score' else sub_fill
        for row, stock in enumerate(top_list, 3):
            for col, h in enumerate(top_cols, 1):
                c = ws_top.cell(row=row, column=col, value=_stock_val(stock, h))
                if h == 'Score':
                    c.fill = SCORE_FILL
                    c.font = Font(bold=True)

    # ── FOGLIO 3: Selezionate (tutti i piani) ───────────────────────────
    if selected:
        ws_sel = wb.create_sheet("Azioni Selezionate", 3)
        for col, h in enumerate(_SEL_COLS, 1):
            c = ws_sel.cell(row=1, column=col, value=h)
            c.font = HDR_FONT
            c.fill = HDR_FILL
        for row, s in enumerate(sorted(selected, key=lambda x: x.get('Score', 0), reverse=True), 2):
            for col, h in enumerate(_SEL_COLS, 1):
                ws_sel.cell(row=row, column=col, value=_fmt_sel_val(s, h))

    # ── FOGLI SCARTATE + DATI MANCANTI + ERRORI (tutti i piani) ─────────
    if True:
        def _m(s): return s.get('Motivo Scarto', '')
        sc_sett  = [s for s in rejected if 'Settore escluso' in _m(s)]
        sc_mc    = [s for s in rejected if 'Market Cap' in _m(s) and 'Settore' not in _m(s)]
        sc_evfcf = [s for s in rejected if 'EV/FCF' in _m(s)]
        sc_pb    = [s for s in rejected if 'P/B' in _m(s) and 'EV/FCF' not in _m(s)]
        sc_roe   = [s for s in rejected if 'ROE' in _m(s) and 'P/B' not in _m(s) and 'EV/FCF' not in _m(s)]
        sc_nd    = [s for s in rejected if 'Net Debt' in _m(s)]
        _cls     = set(id(s) for s in sc_sett+sc_mc+sc_evfcf+sc_pb+sc_roe+sc_nd)
        sc_altri = [s for s in rejected if id(s) not in _cls]

        for sheet_idx, (title, data, extra) in enumerate([
            (f"Esclusi - Settore ({len(sc_sett)})",   sc_sett,  ['Ticker','Nome','Mercato','Settore','Industry','Market Cap','Motivo Scarto']),
            (f"Scartate - Market Cap ({len(sc_mc)})",  sc_mc,    ['Ticker','Nome','Mercato','Settore','Market Cap','Valuta','Motivo Scarto']),
            (f"Scartate - EV-FCF ({len(sc_evfcf)})",  sc_evfcf, ['Ticker','Nome','Mercato','Settore','EV/FCF','P/B','ROE','Motivo Scarto']),
            (f"Scartate - P-B ({len(sc_pb)})",         sc_pb,    ['Ticker','Nome','Mercato','Settore','P/B','EV/FCF','ROE','Motivo Scarto']),
            (f"Scartate - ROE ({len(sc_roe)})",        sc_roe,   ['Ticker','Nome','Mercato','Settore','ROE','P/B','EV/FCF','Motivo Scarto']),
            (f"Scartate - Net Debt ({len(sc_nd)})",    sc_nd,    ['Ticker','Nome','Mercato','Settore','Net Debt/EBITDA','EV/FCF','P/B','Motivo Scarto']),
            (f"Scartate - Altri ({len(sc_altri)})",    sc_altri, ['Ticker','Nome','Mercato','Settore','Motivo Scarto','EV/FCF','P/B','ROE']),
        ], start=4):
            ws_sc = wb.create_sheet(title[:31], sheet_idx)
            if not data: continue
            for col, h in enumerate(extra, 1):
                ws_sc.cell(row=1, column=col, value=h).font = Font(bold=True)
            for row, s in enumerate(data, 2):
                for col, k in enumerate(extra, 1):
                    v = s.get(k)
                    if k == 'ROE' and isinstance(v, float): v = format_percent_ita(v, 1)
                    elif isinstance(v, float): v = round(v, 2)
                    ws_sc.cell(row=row, column=col, value=v)

        # Dati Mancanti
        ws_nv = wb.create_sheet(f"Dati Mancanti ({len(non_validi)})", 11)
        if non_validi:
            nv_hdr = ['Ticker','Nome','Mercato','Settore','Campi Mancanti','Market Cap','P/B','ROE','EV/FCF']
            for col, h in enumerate(nv_hdr, 1):
                ws_nv.cell(row=1, column=col, value=h).font = Font(bold=True)
            for row, s in enumerate(non_validi, 2):
                for col, k in enumerate(nv_hdr, 1):
                    v = s.get(k)
                    if k == 'ROE' and isinstance(v, float): v = format_percent_ita(v, 1)
                    elif isinstance(v, float): v = round(v, 2)
                    ws_nv.cell(row=row, column=col, value=v)

        # Errori
        ws_err = wb.create_sheet(f"Errori ({len(errors)})", 12)
        if errors:
            for col, h in enumerate(['Ticker','Error'], 1):
                ws_err.cell(row=1, column=col, value=h).font = Font(bold=True)
            for row, e in enumerate(errors, 2):
                ws_err.cell(row=row, column=1, value=e.get('Ticker'))
                ws_err.cell(row=row, column=2, value=e.get('Error'))

    fname = os.path.join(REPORTS_DIR, f"Azioni_Screener_{plan_name}_{timestamp}.xlsx")
    wb.save(fname)
    print(f"  ✅ Salvato: {os.path.basename(fname)}")

# ===========================================================================
# FASE 1 — RACCOLTA DATI (tutti i ticker, nessun filtro applicato)
# ===========================================================================
all_stocks = []   # dati validi raccolti, senza filtri
errors     = []

for idx, ticker in enumerate(ALL_AZIONI, 1):
    print(f"[{idx}/{len(ALL_AZIONI)}] {ticker}...", end=" ", flush=True)
    mercato = get_market_from_ticker(ticker)
    stock   = None
    for attempt in range(1, 4):  # max 3 tentativi
        try:
            _ex  = ThreadPoolExecutor(max_workers=1)
            _fut = _ex.submit(get_stock_data, ticker)
            try:
                stock = _fut.result(timeout=45)
            except FuturesTimeout:
                stock = {'Ticker': ticker, 'Error': 'timeout (>45s)'}
                _fut.cancel()
                break  # non ritentare su timeout
            finally:
                _ex.shutdown(wait=False)  # non aspettare thread yfinance bloccato
        except Exception as _e:
            stock = {'Ticker': ticker, 'Error': str(_e)}
        if 'Error' not in stock:
            break
        if not _is_network_error(stock.get('Error', '')):
            break  # errore dati strutturale — non ha senso riprovare
        if attempt < 3:
            print(f"\n  ↻ Errore rete — retry {attempt}/2 tra 5s...", flush=True)
            time.sleep(5)

    if 'Error' in stock:
        errors.append(stock)
        print(f"❌ [{stock['Error'][:60]}]", flush=True)
        time.sleep(0.1)
        continue
    stock['Mercato'] = mercato
    stock['Indice']  = TICKER_TO_INDICE.get(ticker, 'N/A')
    all_stocks.append(stock)
    print("✓", flush=True)
    time.sleep(0.5)

print(f"\n✅ Raccolta: {len(all_stocks)} validi · {len(errors)} errori")

# ── Alert admin se troppi errori di rete ─────────────────────────────────────
_net_err_count = sum(1 for e in errors if _is_network_error(e.get('Error', '')))
if _net_err_count > 100:
    print(f"\n⚠️  ALERT: {_net_err_count} errori di rete su {len(ALL_AZIONI)} azioni — invio email alert...", flush=True)
    _send_alert_email_azioni(len(ALL_AZIONI), _net_err_count, len(errors))

# ===========================================================================
# FASE 2 — FILTRI PER PIANO + SCRITTURA 3 FILE EXCEL
# ===========================================================================
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

for plan_name, cfg in PLAN_CONFIGS.items():
    print(f"\n{'='*60}")
    print(f"PIANO {plan_name} | EV/FCF≤{cfg['ev_fcf_max']} P/B≤{cfg['pb_max']} "
          f"ROE≥{cfg['roe_min']*100:.0f}% ND≤{cfg['nd_max']} Top{cfg['top_n']}")
    print(f"{'='*60}")

    pool = [s for s in all_stocks if s['Ticker'] in cfg['universe']] \
           if cfg['universe'] else all_stocks
    print(f"Universo: {len(pool)} ticker")

    p_sel, p_rej, p_nv = [], [], []
    p_mkt = {}

    for stock in pool:
        mkt = stock.get('Mercato', 'N/A')
        ms  = p_mkt.setdefault(mkt, {'totali':0,'selezionate':0,'scartate':0,'non_validi':0})
        ms['totali'] += 1
        sc = dict(stock)
        result, detail = apply_plan_filters(sc, cfg)
        if result == 'ok':
            p_sel.append(sc)
            ms['selezionate'] += 1
        elif result == 'rejected':
            sc['Motivo Scarto'] = detail
            p_rej.append(sc)
            ms['scartate'] += 1
        else:
            sc['Campi Mancanti'] = detail
            p_nv.append(sc)
            ms['non_validi'] += 1

    # Score percentile 0-100 calcolato in batch (pesi letti da config.json)
    batch_percentile_score(p_sel, 'azioni', plan_name)
    for sc in p_sel:
        print(f"  ✅ {sc['Ticker']:12} Score={sc['Score']:.1f}")

    print(f"→ {len(p_sel)} selezionate / {len(p_rej)} scartate / {len(p_nv)} dati N/A")
    _write_azioni_plan_excel(plan_name, cfg, p_sel, p_rej, p_nv, errors, p_mkt, timestamp)

print(f"\n{'='*60}")
print(f"✅ Tutti e 3 i piani completati.")
print(f"Fine: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"{'='*60}")

# ── Salva prices_cache.json (usato dalla tab DATABASE del dashboard) ──────────
try:
    _cache_path = os.path.join(BASE_DIR, "prices_cache.json")
    _existing_cache = {}
    if os.path.exists(_cache_path):
        with open(_cache_path, encoding='utf-8') as _cf:
            _existing_cache = json.load(_cf)
    _azioni_cache = {}
    for s in all_stocks:
        t = s.get('Ticker', '')
        if not t:
            continue
        _azioni_cache[t] = {
            'name':       s.get('Nome', t),
            'price':      round(float(s['Prezzo']), 4) if s.get('Prezzo') else None,
            'change_pct': s.get('Var_1D_%'),
            'currency':   s.get('Valuta', ''),
        }
    _existing_cache['azioni']     = _azioni_cache
    _existing_cache['azioni_at']  = datetime.now().isoformat()
    with open(_cache_path, 'w', encoding='utf-8') as _cf:
        json.dump(_existing_cache, _cf, ensure_ascii=False)
    print(f"[Cache] prices_cache.json aggiornato: {len(_azioni_cache)} azioni")
except Exception as _e:
    print(f"[Cache] Errore salvataggio: {_e}")
