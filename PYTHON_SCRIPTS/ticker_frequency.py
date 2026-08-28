# -*- coding: utf-8 -*-
"""
RT2026 — Contatore frequenza ticker nel Top N
Scansiona REPORTS_DAILY/ e conta quante volte ogni ticker compare nel foglio Top N per Score.
Salva risultato in ticker_frequency.json.
Chiamato da orchestrator.py dopo ogni run.
"""

import json
import os
import re
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(os.path.dirname(BASE_DIR), "REPORTS_DAILY")
FREQ_FILE   = os.path.join(BASE_DIR, "ticker_frequency.json")

_FILE_RE = re.compile(
    r'^(AZIONI|Azioni|ETF|FONDI_EU|FONDI)_Screener_(BASIC|PRO|VALUE)_(\d{8})_\d{6}\.xlsx$',
    re.IGNORECASE
)


def _ticker_col_idx(headers):
    """Trova indice colonna Ticker nel foglio."""
    for i, h in enumerate(headers or []):
        if h and str(h).strip().upper() in ('TICKER', 'SIMBOLO'):
            return i
    return None


def _nome_col_idx(headers):
    for i, h in enumerate(headers or []):
        if h and str(h).strip().upper() in ('NOME', 'NAME', 'ISIN'):
            return i
    return None


def build_frequency(lookback_days=30):
    """
    Analizza i file Excel degli ultimi `lookback_days` giorni.
    Ritorna dict: {asset_class: {piano: {_meta: {total_days}, ticker: {count, nome, dates, last_date}}}}
    _meta.total_days = numero di report elaborati → usato per calcolare FREQ%
    """
    if load_workbook is None:
        print('[ticker_freq] openpyxl non disponibile')
        return {}

    freq       = {}
    dates_seen = {}   # {asset_class: {piano: set(date_str)}}

    try:
        files = sorted(os.listdir(REPORTS_DIR))
    except Exception as e:
        print(f'[ticker_freq] errore lettura REPORTS_DAILY: {e}')
        return {}

    for fname in files:
        m = _FILE_RE.match(fname)
        if not m:
            continue

        asset_class, piano, date_str = m.group(1).upper(), m.group(2).upper(), m.group(3)

        # Filtra per lookback
        try:
            file_date = datetime.strptime(date_str, '%Y%m%d')
            days_old  = (datetime.now() - file_date).days
            if days_old > lookback_days:
                continue
        except Exception:
            pass

        # Traccia data elaborata per total_days
        dates_seen.setdefault(asset_class, {}).setdefault(piano, set()).add(date_str)

        fpath = os.path.join(REPORTS_DIR, fname)
        try:
            wb = load_workbook(fpath, read_only=True, data_only=True)
            top_sheet = next((s for s in wb.sheetnames if 'Top' in s and 'Score' in s), None)
            if not top_sheet:
                wb.close()
                continue

            ws   = wb[top_sheet]
            rows = list(ws.iter_rows(min_row=1, max_row=52, values_only=True))
            wb.close()

            if not rows:
                continue

            headers   = rows[0]
            tk_idx    = _ticker_col_idx(headers)
            nome_idx  = _nome_col_idx(headers)

            if tk_idx is None:
                continue

            data_rows = rows[2:] if len(rows) > 2 else rows[1:]

            freq.setdefault(asset_class, {}).setdefault(piano, {})
            ac_piano = freq[asset_class][piano]

            for row in data_rows:
                if not row or len(row) <= tk_idx:
                    continue
                ticker = row[tk_idx]
                if not ticker:
                    continue
                ticker = str(ticker).strip().upper()
                if not ticker:
                    continue

                nome = ''
                if nome_idx is not None and len(row) > nome_idx:
                    nome = str(row[nome_idx]).strip() if row[nome_idx] else ''

                entry = ac_piano.setdefault(ticker, {'count': 0, 'nome': nome, 'dates': [], 'last_date': ''})
                if date_str not in entry['dates']:
                    entry['count'] += 1
                    entry['dates'].append(date_str)
                    entry['dates'] = sorted(entry['dates'])[-30:]
                    entry['last_date'] = date_str
                if nome and not entry['nome']:
                    entry['nome'] = nome

        except Exception as e:
            print(f'[ticker_freq] errore {fname}: {e}')

    # Aggiungi _meta.total_days per ogni asset/piano
    for ac in freq:
        for piano in freq[ac]:
            total = len(dates_seen.get(ac, {}).get(piano, set()))
            freq[ac][piano]['_meta'] = {'total_days': max(total, 1)}

    return freq


def save_frequency(freq):
    try:
        with open(FREQ_FILE, 'w', encoding='utf-8') as f:
            json.dump(freq, f, indent=2, ensure_ascii=False)
        print(f'[ticker_freq] Salvato {FREQ_FILE}')
    except Exception as e:
        print(f'[ticker_freq] errore salvataggio: {e}')


def load_frequency():
    """Carica frequenza da file, ritorna dict vuoto se non esiste."""
    if not os.path.exists(FREQ_FILE):
        return {}
    try:
        with open(FREQ_FILE, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def get_freq_for_plan(asset_class, piano):
    """Ritorna {ticker: count} per asset_class/piano dal file corrente."""
    data = load_frequency()
    return {
        tk: info['count']
        for tk, info in data.get(asset_class, {}).get(piano, {}).items()
    }


def top_stable(asset_class, piano, min_count=3, top_n=10):
    """Ritorna lista [(ticker, nome, count)] dei ticker più stabili."""
    data = load_frequency()
    plan_data = data.get(asset_class, {}).get(piano, {})
    candidates = [
        (tk, info.get('nome', ''), info['count'])
        for tk, info in plan_data.items()
        if info['count'] >= min_count
    ]
    return sorted(candidates, key=lambda x: -x[2])[:top_n]


def run():
    print(f'[ticker_freq] Analisi REPORTS_DAILY — {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    freq = build_frequency(lookback_days=30)
    total = sum(
        len(tickers)
        for plans in freq.values()
        for tickers in plans.values()
    )
    print(f'[ticker_freq] {total} ticker unici trovati')
    save_frequency(freq)
    return freq


if __name__ == '__main__':
    run()
