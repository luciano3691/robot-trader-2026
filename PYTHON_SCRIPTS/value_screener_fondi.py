# -*- coding: utf-8 -*-
"""
Robot Trader 2026 - FONDI Screener
FIXED: Formato percentuale italiano + Parametri da parametri.json (VERAMENTE!)
"""
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import json
import time
import smtplib
from email.mime.text import MIMEText
from screener_utils import batch_percentile_score


def _is_network_error(detail):
    if not detail:
        return False
    d = detail.lower()
    return any(k in d for k in ['curl', 'resolve host', 'connection', 'timeout', 'network', 'errno 11001', 'recv failure'])


def _send_alert_email(asset_type, total, network_errors, non_validi_total):
    """Invia email di alert all'admin se troppi errori di rete durante lo screening."""
    try:
        with open(os.path.join(BASE_DIR, 'config.json'), encoding='utf-8') as f:
            cfg_email = json.load(f).get('email', {})
        host   = cfg_email.get('smtp_server', 'smtp.gmail.com')
        port   = cfg_email.get('smtp_port', 587)
        login  = cfg_email.get('smtp_login', '') or cfg_email.get('sender', '')
        pwd    = cfg_email.get('app_password', '')
        sender = cfg_email.get('sender', login)
        if not login or not pwd:
            print(f"[ALERT] SMTP non configurato — alert non inviato", flush=True)
            return
        soglia_pct = round(network_errors / total * 100, 1) if total else 0
        body = (
            f"ALERT — Robot Trader 2026 — Screener {asset_type}\n\n"
            f"Errori di rete rilevati durante lo screening:\n"
            f"  Errori rete: {network_errors} su {total} ({soglia_pct}%)\n"
            f"  Non validi totali: {non_validi_total}\n\n"
            f"Causa probabile: DNS non risolveva query2.finance.yahoo.com al momento del run.\n\n"
            f"AZIONE RICHIESTA — rilanciare lo screener:\n"
            f"  cd C:\\Users\\lucia\\Desktop\\Robot Trader 2026\\PYTHON_SCRIPTS\n"
            f"  python orchestrator.py {asset_type}\n\n"
            f"— Robot Trader 2026"
        )
        msg = MIMEText(body, 'plain', 'utf-8')
        msg['Subject'] = f"[ALERT] Screener {asset_type}: {network_errors} errori rete su {total} ({soglia_pct}%)"
        msg['From']    = sender
        msg['To']      = login
        with smtplib.SMTP(host, port, timeout=15) as srv:
            srv.ehlo(); srv.starttls()
            srv.login(login, pwd)
            srv.sendmail(sender, [login], msg.as_string())
        print(f"[ALERT] Email alert inviata a {login}", flush=True)
    except Exception as e:
        print(f"[ALERT] Errore invio alert email: {e}", flush=True)

# Path
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(os.path.dirname(BASE_DIR), "REPORTS_DAILY")
PARAMETRI_FILE = os.path.join(BASE_DIR, "parametri.json")
os.makedirs(REPORTS_DIR, exist_ok=True)

# Importa lista FONDI
sys.path.insert(0, BASE_DIR)
from ticker_lists_5000 import ALL_FONDI

# CARICA PARAMETRI DA JSON - VERSIONE CORRETTA
def _get_param(params: dict, key: str, default):
    """Gestisce sia {'value': X} che plain X in parametri.json"""
    val = params.get(key, default)
    if isinstance(val, dict):
        return val.get('value', default)
    return val if val is not None else default


def load_filters():
    """Carica filtri da parametri.json"""
    try:
        with open(PARAMETRI_FILE, 'r', encoding='utf-8') as f:
            parametri = json.load(f)

        fondi_params = parametri.get('fondi', {})

        return {
            'ter_max':            _get_param(fondi_params, 'ter_max',            1.00),
            'sharpe_min':         _get_param(fondi_params, 'sharpe_min',         0.3),
            'volume_min':         _get_param(fondi_params, 'volume_min',         50000),
            'performance_1y_min': _get_param(fondi_params, 'performance_1y_min', -0.30),
        }
    except Exception as e:
        print(f"⚠️ ERRORE caricamento parametri: {e}")
        print("⚠️ Uso valori di default")
        return {
            'ter_max': 1.00,
            'sharpe_min': 0.3,
            'volume_min': 50000,
            'performance_1y_min': -0.30,
        }

# CARICA FILTRI DA JSON - PUNTO CRITICO
FILTERS = load_filters()

def format_stelle_ms(rating):
    """Formatta rating Morningstar come stelle (es. 4 → ★★★★☆)"""
    if not rating or not isinstance(rating, (int, float)):
        return 'N/A'
    r = max(0, min(5, int(rating)))
    return '★' * r + '☆' * (5 - r)

def format_percent_ita(value, decimals=1):
    """Formatta percentuale con virgola italiana"""
    if value is None or (isinstance(value, float) and (value != value)):
        return 'N/A'
    return f"{value*100:.{decimals}f}%".replace('.', ',')

def analyze_fund(ticker):
    """Analizza singolo fondo"""
    try:
        fund = yf.Ticker(ticker)
        info = fund.info
        hist = fund.history(period="2y")

        if hist.empty or len(hist) < 50:
            return None, "Dati storici", "Dati storici insufficienti (nessun dato)"
        hist = hist.tail(252)  # usa solo l'ultimo anno per i calcoli

        # TER — Layer 1: info() (3 field name variants)
        ter = info.get('annualReportExpenseRatio',
              info.get('netExpenseRatio',
              info.get('totalExpenseRatio', None)))
        # TER — Layer 2: funds_data (endpoint separato da yfinance)
        if ter is None:
            try:
                fd = fund.funds_data
                if fd and fd.fund_operations is not None:
                    ops = fd.fund_operations
                    for field in ['Annual Report Expense Ratio', 'Net Expense Ratio', 'Expense Ratio']:
                        if field in ops.index:
                            val = ops.loc[field].iloc[0]
                            if val is not None and str(val) != '<NA>':
                                ter = float(val)
                                break
            except Exception:
                pass
        if ter is None:
            return None, "TER", "TER non disponibile (info() + funds_data)"

        # Per mutual fund: totalAssets al posto di volume (i fondi non hanno volume di scambio)
        total_assets = info.get('totalAssets', info.get('netAssets', 0)) or 0
        volume = info.get('volume', 0) or 0
        # Se è un mutual fund (volume=0), usa totalAssets come metrica di dimensione
        size_metric = volume if volume > 0 else total_assets
        name = info.get('longName', info.get('shortName', ticker))
        # Categoria: info() prima, poi funds_data.fund_overview come fallback
        category = info.get('category') or ''
        if not category or category == 'N/A':
            try:
                fd = fund.funds_data
                if fd is not None and fd.fund_overview is not None:
                    ov = fd.fund_overview
                    for field in ['Category', 'Fund Category', 'Morningstar Category']:
                        if field in ov.index:
                            val = str(ov.loc[field].iloc[0])
                            if val and val not in ('nan', 'N/A', 'None'):
                                category = val; break
            except Exception:
                pass
        if not category or str(category) in ('nan', 'None'):
            category = info.get('fundFamily', 'N/A') or 'N/A'
        
        # Performance 1Y, 6M, 3M, YTD + variazione giornaliera
        price_1y_ago = hist['Close'].iloc[0] if len(hist) > 0 else None
        price_now = hist['Close'].iloc[-1]
        perf_1y = ((price_now / price_1y_ago) - 1) if price_1y_ago else 0
        var_1d  = round(((price_now / hist['Close'].iloc[-2]) - 1) * 100, 2) if len(hist) >= 2 else None

        perf_3m = round(((price_now / hist['Close'].iloc[-63]) - 1) * 100, 2) if len(hist) >= 63 else None
        perf_6m = round(((price_now / hist['Close'].iloc[-126]) - 1) * 100, 2) if len(hist) >= 126 else None
        current_year = datetime.now().year
        hist_ytd = hist[hist.index.year == current_year]
        perf_ytd = round(((price_now / hist_ytd['Close'].iloc[0]) - 1) * 100, 2) if not hist_ytd.empty else None

        # Sharpe (approssimato)
        returns = hist['Close'].pct_change().dropna()
        sharpe = (returns.mean() / returns.std() * (252 ** 0.5)) if returns.std() > 0 else 0

        ms_rating = info.get('morningStarOverallRating') or info.get('morningStarRiskRating')

        data = {
            'Ticker': ticker,
            'Nome': name,
            'Categoria': category,
            'TER': ter,
            'Sharpe Ratio': sharpe,
            'Volume': volume,
            'AUM': total_assets,
            'Var_1D_%':  var_1d,
            'Performance 1Y': perf_1y,
            'Perf 3M %':  perf_3m,
            'Perf 6M %':  perf_6m,
            'Perf YTD %': perf_ytd,
            'Stelle MS':  ms_rating,
            'Prezzo': price_now,
            'Data Dati': datetime.now().strftime('%Y-%m-%d')
        }
        
        data['Score'] = 0  # calcolato in batch dopo la selezione per piano
        return data, 'ok', None

    except Exception as e:
        return None, "Eccezione", str(e)

# ─── 3-PLAN SYSTEM ────────────────────────────────────────────────────────────

SERVIZI_FILE = os.path.join(BASE_DIR, 'servizi_config.json')

def _build_fondi_plan_configs():
    """Legge i parametri filtro da servizi_config.json (modificabili dalla dashboard).
    ter_max in servizi è in % (es. 1 = 1%); apply_fondi_plan_filters usa cfg['ter_max']/100.
    performance_1y_min in servizi è in % intera (es. 5 = +5%) → converti in decimale (/100).
    """
    _fixed    = {'BASIC': {'top_n': 20}, 'PRO': {'top_n': 50}, 'VALUE': {'top_n': 50}}
    _fallback = {
        'BASIC': {'ter_max': 0.75, 'sharpe_min': 0.6, 'volume_min': 200_000, 'performance_1y_min': -0.10},
        'PRO':   {'ter_max': 1.00, 'sharpe_min': 0.3, 'volume_min':  50_000, 'performance_1y_min': -0.30},
        'VALUE': {'ter_max': 1.50, 'sharpe_min': 0.1, 'volume_min':  10_000, 'performance_1y_min': -0.50},
    }
    try:
        with open(SERVIZI_FILE, encoding='utf-8') as f:
            sv = json.load(f).get('fondi', {})
        configs = {}
        for plan, key in [('BASIC','basic'), ('PRO','pro'), ('VALUE','value')]:
            p = sv.get(key, {}).get('parametri', {})
            cfg = dict(_fallback[plan])
            if p:
                if 'ter_max'            in p: cfg['ter_max']            = float(p['ter_max'])
                if 'sharpe_min'         in p: cfg['sharpe_min']         = float(p['sharpe_min'])
                if 'volume_min'         in p: cfg['volume_min']         = int(p['volume_min'])
                if 'performance_1y_min' in p: cfg['performance_1y_min'] = float(p['performance_1y_min']) / 100
            cfg.update(_fixed[plan])
            configs[plan] = cfg
        return configs
    except Exception as e:
        print(f"⚠️ servizi_config.json non leggibile ({e}) — uso valori di default")
        return {pl: {**_fallback[pl], **_fixed[pl]} for pl in _fallback}

FONDI_PLAN_CONFIGS = _build_fondi_plan_configs()


def apply_fondi_plan_filters(fund, cfg):
    """Applica filtri piano-specifici a un fondo già analizzato.
    Ritorna ('selected', None) o ('rejected', motivo_stringa).
    TER dei fondi è in forma decimale (0.0035 = 0.35%), cfg['ter_max'] è in % (0.75 = 0.75%).
    """
    reasons = []
    ter          = fund.get('TER', 0) or 0
    sharpe       = fund.get('Sharpe Ratio', 0) or 0
    volume       = fund.get('Volume', 0) or 0
    total_assets = fund.get('AUM', 0) or 0
    size_metric  = volume if volume > 0 else total_assets
    perf_1y      = fund.get('Performance 1Y', 0) or 0

    if ter > cfg['ter_max'] / 100:
        reasons.append(f"TER alto ({format_percent_ita(ter, 2)})")
    if sharpe < cfg['sharpe_min']:
        reasons.append(f"Sharpe basso ({sharpe:.2f})")
    if size_metric < cfg['volume_min']:
        reasons.append(f"Dimensione bassa (Vol={volume:,} AUM={total_assets:,})")
    if perf_1y < cfg['performance_1y_min']:
        reasons.append(f"Performance negativa ({format_percent_ita(perf_1y, 1)})")

    if reasons:
        return 'rejected', '; '.join(reasons)
    return 'selected', None


def _fmt_num(n):
    if n is None:
        return 'N/A'
    try:
        n = float(n)
    except (TypeError, ValueError):
        return 'N/A'
    if n == 0:
        return '0'
    if abs(n) >= 1_000_000_000:
        return f"{n/1_000_000_000:.2f}B"
    if abs(n) >= 1_000_000:
        return f"{n/1_000_000:.2f}M"
    if abs(n) >= 1_000:
        return f"{n/1_000:.1f}K"
    return f"{n:.0f}"

def _score_fill(score):
    try:
        s = float(score)
    except (TypeError, ValueError):
        return None
    if s >= 75:
        return PatternFill(start_color="1A9850", end_color="1A9850", fill_type="solid")
    if s >= 55:
        return PatternFill(start_color="FEE08B", end_color="FEE08B", fill_type="solid")
    if s >= 35:
        return PatternFill(start_color="F46D43", end_color="F46D43", fill_type="solid")
    return PatternFill(start_color="D73027", end_color="D73027", fill_type="solid")

def _score_font(score):
    try:
        s = float(score)
    except (TypeError, ValueError):
        return Font(bold=True)
    color = "FFFFFF" if s < 55 or s >= 75 else "1A1A1A"
    return Font(bold=True, color=color)

def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _autofit(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_w), max_w)

_FONDI_SEL_COLS = [
    'Ticker','Nome','Categoria','TER','Sharpe Ratio','Volume','AUM',
    'Prezzo','Var_1D_%','Perf 3M %','Perf 6M %','Perf YTD %','Performance 1Y',
    'Stelle MS','Score','Data Dati'
]

def _fmt_fondi_sel_val(fund, col):
    """Formatta valori per il foglio Fondi Selezionati — numeri leggibili."""
    v = fund.get(col)
    if v is None:
        return 'N/A'
    if col in {'Volume', 'AUM'}:
        return _fmt_num(v)
    if col == 'TER':
        return format_percent_ita(v, 2) if isinstance(v, (int, float)) else 'N/A'
    if col == 'Stelle MS':
        return format_stelle_ms(v)
    if col == 'Performance 1Y':
        return format_percent_ita(v, 1) if isinstance(v, (int, float)) else 'N/A'
    if col in {'Var_1D_%', 'Perf 3M %', 'Perf 6M %', 'Perf YTD %'}:
        return f"{v:.2f}%".replace('.', ',') if isinstance(v, (int, float)) else 'N/A'
    if col in {'Sharpe Ratio', 'Prezzo'}:
        return round(float(v), 2) if isinstance(v, (int, float)) else 'N/A'
    if col == 'Score':
        return round(float(v), 1) if isinstance(v, (int, float)) else 0
    return v

def _write_fondi_plan_excel(plan_name, cfg, selected, rejected, non_validi, timestamp):
    """Scrive un Excel per il piano specificato — tutti i piani hanno gli stessi fogli."""
    wb = Workbook()
    wb.remove(wb.active)
    top_n = cfg['top_n']

    C_HEADER_BG = "1A3A5C"
    C_HEADER_FG = "FFFFFF"
    C_TITLE_FG  = "FF8C42"
    C_SECTION_BG = "EBF2FA"
    C_ALT_ROW   = "F7FAFD"
    C_GREEN     = "1A9850"
    C_RED       = "D73027"
    C_ORANGE    = "F46D43"

    def _hdr_cell(ws, row, col, val, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color=fg, size=size)
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _thin_border()
        return c

    def _write_sc(wb, title, sheet_name, idx, data, cols):
        ws = wb.create_sheet(sheet_name, idx)
        ws.freeze_panes = 'A3'
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=11, color="444444")
        ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
        ws['A1'].fill = PatternFill(start_color=C_SECTION_BG, end_color=C_SECTION_BG, fill_type="solid")
        for c, h in enumerate(cols, 1):
            _hdr_cell(ws, 2, c, h)
        for r, fund in enumerate(data, 3):
            fill_row = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if r % 2 == 0 else None
            for c, h in enumerate(cols, 1):
                v = fund.get(h)
                if h == 'TER':               v = format_percent_ita(v, 2) if v else 'N/A'
                elif h == 'Performance 1Y':  v = format_percent_ita(v, 1)
                elif h in ('Volume', 'AUM'): v = _fmt_num(v)
                elif isinstance(v, float):   v = round(v, 2)
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = _thin_border()
                if fill_row:
                    cell.fill = fill_row
        _autofit(ws)
        return ws

    # ── FOGLIO 1: Dashboard ─────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard", 0)
    ws_dash.sheet_view.showGridLines = False

    ws_dash.merge_cells('A1:F1')
    ws_dash['A1'] = f"FONDI SCREENER  ·  Piano {plan_name}  ·  Robot Trader 2026"
    ws_dash['A1'].font = Font(size=18, bold=True, color=C_TITLE_FG)
    ws_dash['A1'].fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    ws_dash['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[1].height = 36

    ws_dash.merge_cells('A2:F2')
    ws_dash['A2'] = f"Generato il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
    ws_dash['A2'].font = Font(size=10, italic=True, color="888888")
    ws_dash['A2'].alignment = Alignment(horizontal='center')

    ws_dash['A4'] = "CRITERI APPLICATI"
    ws_dash['A4'].font = Font(bold=True, size=11, color=C_HEADER_BG)
    criteri = [
        ("TER massimo",        f"≤ {cfg['ter_max']:.2f}%"),
        ("Sharpe minimo",      f"≥ {cfg['sharpe_min']}"),
        ("Dimensione minima",  f"≥ {_fmt_num(cfg['volume_min'])}"),
        ("Performance 1Y min", f"≥ {cfg['performance_1y_min']*100:.0f}%"),
    ]
    for i, (k, v) in enumerate(criteri, 5):
        ws_dash.cell(row=i, column=1, value=k).font = Font(bold=True, color="555555")
        ws_dash.cell(row=i, column=2, value=v).font = Font(color="1A3A5C")

    totale = len(selected) + len(rejected) + len(non_validi)
    ws_dash['D4'] = "RISULTATI SCREENING"
    ws_dash['D4'].font = Font(bold=True, size=11, color=C_HEADER_BG)
    stat_rows = [
        ("Universo Fondi",        len(ALL_FONDI),   "444444"),
        ("✅ Selezionati",         len(selected),    C_GREEN),
        ("❌ Scartati (filtri)",   len(rejected),    C_ORANGE),
        ("⚠️  Non validi",         len(non_validi),  C_RED),
        ("Totale verificato",     totale,           "444444"),
        ("Tasso selezione",       f"{len(selected)/totale*100:.1f}%" if totale else "0%", C_GREEN),
    ]
    for i, (label, value, color) in enumerate(stat_rows, 5):
        ws_dash.cell(row=i, column=4, value=label).font = Font(bold=True, color="555555")
        ws_dash.cell(row=i, column=5, value=value).font = Font(bold=True, color=color)

    top5 = sorted(selected, key=lambda x: x.get('Score', 0), reverse=True)[:5]
    if top5:
        r_start = 13
        ws_dash.cell(row=r_start, column=1, value="TOP 5 FONDI PER SCORE").font = Font(bold=True, size=11, color=C_HEADER_BG)
        ws_dash.merge_cells(f'A{r_start}:F{r_start}')
        for c, h in enumerate(['#', 'Ticker', 'Nome', 'Score', 'Perf 1Y %', 'TER'], 1):
            _hdr_cell(ws_dash, r_start+1, c, h)
        for rank, fund in enumerate(top5, 1):
            row = r_start + 1 + rank
            score = fund.get('Score', 0)
            perf1y = fund.get('Performance 1Y')
            ws_dash.cell(row=row, column=1, value=rank).alignment = Alignment(horizontal='center')
            ws_dash.cell(row=row, column=2, value=fund.get('Ticker')).font = Font(bold=True, color=C_HEADER_BG)
            ws_dash.cell(row=row, column=3, value=fund.get('Nome'))
            sc = ws_dash.cell(row=row, column=4, value=round(score, 1))
            sc.fill = _score_fill(score)
            sc.font = _score_font(score)
            sc.alignment = Alignment(horizontal='center')
            p1 = ws_dash.cell(row=row, column=5, value=round(perf1y*100, 2) if perf1y else 'N/A')
            p1.font = Font(color=C_GREEN if (perf1y or 0) >= 0 else C_RED)
            ws_dash.cell(row=row, column=6, value=format_percent_ita(fund.get('TER'), 2) if fund.get('TER') else 'N/A')
            for c in range(1, 7):
                ws_dash.cell(row=row, column=c).border = _thin_border()

    r_leg = 22
    ws_dash.cell(row=r_leg, column=1, value="LEGENDA SCORE").font = Font(bold=True, size=9, color="888888")
    for col, (label, color) in enumerate([("≥75 Ottimo","1A9850"),("55-74 Buono","FEE08B"),("35-54 Medio","F46D43"),("<35 Basso","D73027")], 1):
        c = ws_dash.cell(row=r_leg+1, column=col, value=label)
        c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        c.font = Font(bold=True, size=9, color="FFFFFF" if color in ("1A9850","D73027","F46D43") else "333333")
        c.alignment = Alignment(horizontal='center')

    for col, w in zip('ABCDEF', [28, 16, 28, 22, 16, 12]):
        ws_dash.column_dimensions[col].width = w

    # ── FOGLIO 2: Top N per Score ──────────────────────────────
    ws_top = wb.create_sheet(f"Top {top_n} per Score", 1)
    ws_top.freeze_panes = 'A2'
    ws_top.sheet_view.showGridLines = False
    top_funds = sorted(selected, key=lambda x: x.get('Score', 0), reverse=True)[:top_n]
    if top_funds:
        headers = ['#', 'Ticker', 'Nome', 'Score', 'Stelle MS', 'Perf 1Y %', 'Perf 6M %',
                   'Perf 3M %', 'Perf YTD %', 'TER', 'Sharpe', 'AUM', 'Categoria']
        for col, header in enumerate(headers, 1):
            _hdr_cell(ws_top, 1, col, header)
        for row, fund in enumerate(top_funds, 2):
            score = fund.get('Score', 0)
            perf1y = fund.get('Performance 1Y')
            fill_alt = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if row % 2 == 0 else None
            cells_vals = [
                (1,  row-1,                        None,                                         Alignment(horizontal='center')),
                (2,  fund.get('Ticker'),            Font(bold=True, color=C_HEADER_BG),           None),
                (3,  fund.get('Nome'),              None,                                         None),
                (4,  round(score, 1),               _score_font(score),                           Alignment(horizontal='center')),
                (5,  format_stelle_ms(fund.get('Stelle MS')), None,                               Alignment(horizontal='center')),
                (6,  round(perf1y*100,2) if perf1y else 'N/A',
                     Font(color=C_GREEN if (perf1y or 0)>=0 else C_RED, bold=True),               Alignment(horizontal='right')),
                (7,  fund.get('Perf 6M %'),         None,                                         Alignment(horizontal='right')),
                (8,  fund.get('Perf 3M %'),         None,                                         Alignment(horizontal='right')),
                (9,  fund.get('Perf YTD %'),        None,                                         Alignment(horizontal='right')),
                (10, format_percent_ita(fund.get('TER'),2) if fund.get('TER') else 'N/A', None,  Alignment(horizontal='center')),
                (11, round(fund.get('Sharpe Ratio',0),2), None,                                   Alignment(horizontal='right')),
                (12, _fmt_num(fund.get('AUM',0)),   None,                                         Alignment(horizontal='right')),
                (13, fund.get('Categoria','N/A'),   None,                                         None),
            ]
            for col, val, font, align in cells_vals:
                cell = ws_top.cell(row=row, column=col, value=val)
                if col == 4:
                    cell.fill = _score_fill(score)
                elif fill_alt:
                    cell.fill = fill_alt
                if font:
                    cell.font = font
                if align:
                    cell.alignment = align
                cell.border = _thin_border()
        ws_top.row_dimensions[1].height = 22
        _autofit(ws_top, min_w=6, max_w=45)

    # ── FOGLIO 3: Fondi Selezionati ────────────────────────────
    if selected:
        ws_sel = wb.create_sheet("Fondi Selezionati", 2)
        ws_sel.freeze_panes = 'A2'
        for col, header in enumerate(_FONDI_SEL_COLS, 1):
            _hdr_cell(ws_sel, 1, col, header)
        for row, fund in enumerate(sorted(selected, key=lambda x: x.get('Score', 0), reverse=True), 2):
            fill_alt = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if row % 2 == 0 else None
            for col, key in enumerate(_FONDI_SEL_COLS, 1):
                cell = ws_sel.cell(row=row, column=col, value=_fmt_fondi_sel_val(fund, key))
                cell.border = _thin_border()
                if fill_alt:
                    cell.fill = fill_alt
        _autofit(ws_sel)

    # ── FOGLI SCARTATI ─────────────────────────────────────────
    def _m(f): return f.get('Motivo Scarto', '')
    sc_ter    = [f for f in rejected if 'TER alto' in _m(f)]
    sc_sharpe = [f for f in rejected if 'Sharpe basso' in _m(f) and 'TER' not in _m(f)]
    sc_dim    = [f for f in rejected if 'Dimensione bassa' in _m(f) and 'TER' not in _m(f) and 'Sharpe' not in _m(f)]
    sc_perf   = [f for f in rejected if 'Performance negativa' in _m(f) and 'TER' not in _m(f)]
    class_ids = set(id(f) for f in sc_ter + sc_sharpe + sc_dim + sc_perf)
    sc_altri  = [f for f in rejected if id(f) not in class_ids]

    _write_sc(wb, f"SCARTATI PER TER > {cfg['ter_max']:.2f}% ({len(sc_ter)})",
              "Scartati - TER Alto", 3, sc_ter,
              ['Ticker','Nome','TER','Sharpe Ratio','Performance 1Y','Categoria','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI PER SHARPE < {cfg['sharpe_min']} ({len(sc_sharpe)})",
              "Scartati - Sharpe Basso", 4, sc_sharpe,
              ['Ticker','Nome','Sharpe Ratio','TER','Performance 1Y','Categoria','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI PER DIMENSIONE < {cfg['volume_min']:,} ({len(sc_dim)})",
              "Scartati - Dimensione", 5, sc_dim,
              ['Ticker','Nome','Volume','AUM','TER','Sharpe Ratio','Categoria','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI PER PERF1Y < {cfg['performance_1y_min']*100:.0f}% ({len(sc_perf)})",
              "Scartati - Performance", 6, sc_perf,
              ['Ticker','Nome','Performance 1Y','TER','Sharpe Ratio','Categoria','Motivo Scarto'])
    _write_sc(wb, f"ALTRI MOTIVI ({len(sc_altri)})",
              "Scartati - Altri Motivi", 7, sc_altri,
              ['Ticker','Nome','Motivo Scarto','TER','Sharpe Ratio'])

    # ── FOGLIO Non Validi ──────────────────────────────────────
    ws_nv = wb.create_sheet("Non Validi", 8)
    ws_nv.freeze_panes = 'A3'
    ws_nv.merge_cells('A1:C1')
    ws_nv['A1'] = f"NON VALIDI — DATI MANCANTI ({len(non_validi)})"
    ws_nv['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_nv['A1'].fill = PatternFill(start_color=C_ORANGE, end_color=C_ORANGE, fill_type="solid")
    ws_nv['A1'].alignment = Alignment(horizontal='center')
    for col, h in enumerate(['Ticker', 'Campo Mancante', 'Motivo'], 1):
        _hdr_cell(ws_nv, 2, col, h)
    for row, item in enumerate(non_validi, 3):
        ws_nv.cell(row=row, column=1, value=item.get('Ticker')).border = _thin_border()
        ws_nv.cell(row=row, column=2, value=item.get('Campo Mancante', '—')).border = _thin_border()
        ws_nv.cell(row=row, column=3, value=item.get('Motivo', '—')).border = _thin_border()
    _autofit(ws_nv)

    filename = os.path.join(REPORTS_DIR, f"FONDI_Screener_{plan_name}_{timestamp}.xlsx")
    wb.save(filename)
    print(f"✅ [{plan_name}] Report salvato: {filename}")


print("="*70)
print("FONDI SCREENER - Robot Trader 2026")
print("="*70)
print(f"Fondi totali da analizzare: {len(ALL_FONDI)}")
print("="*70)

# PHASE 1: raccolta dati su tutti i fondi (senza filtri piano)
all_fondi = []
non_validi_global = []

total = len(ALL_FONDI)
for i, ticker in enumerate(ALL_FONDI, 1):
    print(f"[{i}/{total}] {ticker}...", flush=True)
    data, status, detail = None, None, None
    for attempt in range(1, 4):  # max 3 tentativi
        data, status, detail = analyze_fund(ticker)
        if data is not None:
            break
        if not _is_network_error(detail):
            break  # errore dati strutturale — non ha senso riprovare
        if attempt < 3:
            print(f"  ↻ Errore rete — retry {attempt}/2 tra 5s...", flush=True)
            time.sleep(5)

    if data:
        all_fondi.append(data)
        if attempt > 1:
            print(f"  ✅ Raccolto al tentativo {attempt}", flush=True)
        else:
            print(f"  ✅ Raccolto", flush=True)
    else:
        campo  = status or 'Sconosciuto'
        motivo = detail or 'Dati insufficienti'
        non_validi_global.append({'Ticker': ticker, 'Campo Mancante': campo, 'Motivo': motivo})
        print(f"  ⚠️  Non valido [{campo}]: {motivo}", flush=True)

print("="*70)
print(f"📊 Raccolti: {len(all_fondi)} | Non validi: {len(non_validi_global)}")
print("="*70)

# ── Alert admin se troppi errori di rete ─────────────────────────────────────
_net_err_count = sum(1 for x in non_validi_global if _is_network_error(x.get('Motivo', '')))
if _net_err_count > 50:
    print(f"\n⚠️  ALERT: {_net_err_count} errori di rete su {total} fondi — invio email alert...", flush=True)
    _send_alert_email('FONDI', total, _net_err_count, len(non_validi_global))

# PHASE 2: filtro e output per ciascun piano
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

for plan_name, cfg in FONDI_PLAN_CONFIGS.items():
    print(f"\n{'─'*60}")
    print(f"▶ Piano {plan_name}: TER≤{cfg['ter_max']:.2f}% | Sharpe≥{cfg['sharpe_min']} | "
          f"Dim≥{cfg['volume_min']:,} | Perf1Y≥{cfg['performance_1y_min']*100:.0f}% | Top{cfg['top_n']}")

    p_sel = []
    p_rej = []

    for fund in all_fondi:
        sc = fund.copy()
        result, motivo = apply_fondi_plan_filters(sc, cfg)
        if result == 'selected':
            p_sel.append(sc)
        else:
            sc['Motivo Scarto'] = motivo
            p_rej.append(sc)

    # Score percentile 0-100 calcolato in batch (pesi letti da config.json)
    batch_percentile_score(p_sel, 'fondi', plan_name)
    print(f"→ {len(p_sel)} selezionati / {len(p_rej)} scartati / {len(non_validi_global)} non validi")
    _write_fondi_plan_excel(plan_name, cfg, p_sel, p_rej, non_validi_global, timestamp)

print(f"\n{'='*60}")
print(f"✅ Tutti e 3 i piani FONDI completati.")
print(f"Fine: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"{'='*60}")

# ── Salva prices_cache.json (usato dalla tab DATABASE del dashboard) ──────────
try:
    _cache_path = os.path.join(BASE_DIR, "prices_cache.json")
    _existing_cache = {}
    if os.path.exists(_cache_path):
        with open(_cache_path, encoding='utf-8') as _cf:
            _existing_cache = json.load(_cf)
    _fondi_cache = {}
    for s in all_fondi:
        t = s.get('Ticker', '')
        if not t:
            continue
        p = s.get('Prezzo')
        _fondi_cache[t] = {
            'name':       s.get('Nome', t),
            'price':      round(float(p), 4) if p else None,
            'change_pct': s.get('Var_1D_%'),
            'currency':   s.get('Valuta', s.get('Currency', '')),
        }
    _existing_cache['fondi']    = _fondi_cache
    _existing_cache['fondi_at'] = datetime.now().isoformat()
    with open(_cache_path, 'w', encoding='utf-8') as _cf:
        json.dump(_existing_cache, _cf, ensure_ascii=False)
    print(f"[Cache] prices_cache.json aggiornato: {len(_fondi_cache)} fondi")
except Exception as _e:
    print(f"[Cache] Errore salvataggio: {_e}")
