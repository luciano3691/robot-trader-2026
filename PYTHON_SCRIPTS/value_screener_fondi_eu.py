# -*- coding: utf-8 -*-
"""
Robot Trader 2026 - FONDI EU Screener
Analizza fondi UCITS europei usando fondi_eu_universe_cache.json
(costruito da fetch_fondi_eu_universe.py) + dati Yahoo Finance.
"""
import socket
socket.setdefaulttimeout(20)
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import json
import time
import smtplib
from email.mime.text import MIMEText
from screener_utils import batch_percentile_score

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(os.path.dirname(BASE_DIR), "REPORTS_DAILY")
CACHE_FILE  = os.path.join(BASE_DIR, 'fondi_eu_universe_cache.json')
SERVIZI_FILE = os.path.join(BASE_DIR, 'servizi_config.json')
os.makedirs(REPORTS_DIR, exist_ok=True)


def _is_network_error(detail):
    if not detail:
        return False
    d = detail.lower()
    return any(k in d for k in ['curl', 'resolve host', 'connection', 'timeout',
                                 'network', 'errno 11001', 'recv failure'])


def _send_alert_email(total, network_errors, non_validi_total):
    try:
        with open(os.path.join(BASE_DIR, 'config.json'), encoding='utf-8') as f:
            cfg_email = json.load(f).get('email', {})
        host   = cfg_email.get('smtp_server', 'smtp.gmail.com')
        port   = cfg_email.get('smtp_port', 587)
        login  = cfg_email.get('smtp_login', '') or cfg_email.get('sender', '')
        pwd    = cfg_email.get('app_password', '')
        sender = cfg_email.get('sender', login)
        if not login or not pwd:
            return
        pct = round(network_errors / total * 100, 1) if total else 0
        body = (
            f"ALERT — Robot Trader 2026 — Screener FONDI EU\n\n"
            f"Errori rete: {network_errors}/{total} ({pct}%)\n"
            f"Non validi: {non_validi_total}\n\n"
            f"Azione: python orchestrator.py FONDI_EU\n\n— Robot Trader 2026"
        )
        msg = MIMEText(body, 'plain', 'utf-8')
        msg['Subject'] = f"[ALERT] Screener FONDI EU: {network_errors} errori rete ({pct}%)"
        msg['From'] = sender
        msg['To']   = login
        with smtplib.SMTP(host, port, timeout=15) as srv:
            srv.ehlo(); srv.starttls()
            srv.login(login, pwd)
            srv.sendmail(sender, [login], msg.as_string())
        print(f"[ALERT] Email inviata a {login}", flush=True)
    except Exception as e:
        print(f"[ALERT] Errore invio: {e}", flush=True)


# ── Carica universo da cache ──────────────────────────────────────────────────
def load_eu_universe():
    if not os.path.exists(CACHE_FILE):
        print(f"⚠️  Cache non trovata: {CACHE_FILE}")
        print("   Esegui prima: python fetch_fondi_eu_universe.py --seed")
        return []
    with open(CACHE_FILE, encoding='utf-8') as f:
        cache = json.load(f)
    # Solo fondi con ticker Yahoo Finance e non in errore
    valid = [v for v in cache.values()
             if not v.get('error') and v.get('yahoo_ticker')]
    print(f"Cache: {len(cache)} ISIN totali → {len(valid)} con ticker Yahoo Finance", flush=True)
    return valid


def format_stelle_ms(rating):
    if not rating or not isinstance(rating, (int, float)):
        return 'N/A'
    r = max(0, min(5, int(rating)))
    return '★' * r + '☆' * (5 - r)


def format_percent_ita(value, decimals=1):
    if value is None or (isinstance(value, float) and value != value):
        return 'N/A'
    return f"{value*100:.{decimals}f}%".replace('.', ',')


def analyze_fund_eu(entry):
    """Analizza un fondo EU usando Yahoo Finance (ticker 0P... o exchange)."""
    ticker    = entry.get('yahoo_ticker', '')
    isin      = entry.get('isin', ticker)
    ter_cache = entry.get('ter')       # TER da justETF (più affidabile)
    ms_cache  = entry.get('ms_rating')
    category  = entry.get('category', '') or entry.get('manager', 'N/A') or 'N/A'
    manager   = entry.get('manager', 'N/A') or 'N/A'

    try:
        fund = yf.Ticker(ticker)
        info = fund.info
        hist = fund.history(period="2y")

        if hist.empty or len(hist) < 50:
            return None, "Dati storici", f"Storico insufficiente per {ticker}"

        hist = hist.tail(252)

        # TER: priorità alla cache justETF, fallback yfinance
        ter = ter_cache
        if ter is None:
            ter = info.get('annualReportExpenseRatio',
                  info.get('netExpenseRatio',
                  info.get('totalExpenseRatio', None)))
            if ter is None:
                return None, "TER", f"TER non disponibile (cache+yfinance) per {ticker}"

        name = entry.get('name') or info.get('longName', info.get('shortName', ticker))
        currency = entry.get('currency') or info.get('currency', 'EUR')

        total_assets = info.get('totalAssets', info.get('netAssets', 0)) or 0
        volume = info.get('volume', 0) or 0
        size_metric = volume if volume > 0 else total_assets

        # Performance
        price_now   = hist['Close'].iloc[-1]
        price_1y    = hist['Close'].iloc[0]
        perf_1y     = ((price_now / price_1y) - 1) if price_1y else 0
        var_1d      = round(((price_now / hist['Close'].iloc[-2]) - 1) * 100, 2) if len(hist) >= 2 else None
        perf_3m     = round(((price_now / hist['Close'].iloc[-63]) - 1) * 100, 2) if len(hist) >= 63 else None
        perf_6m     = round(((price_now / hist['Close'].iloc[-126]) - 1) * 100, 2) if len(hist) >= 126 else None
        current_year = datetime.now().year
        hist_ytd    = hist[hist.index.year == current_year]
        perf_ytd    = round(((price_now / hist_ytd['Close'].iloc[0]) - 1) * 100, 2) if not hist_ytd.empty else None

        # Sharpe
        returns = hist['Close'].pct_change().dropna()
        sharpe  = (returns.mean() / returns.std() * (252 ** 0.5)) if returns.std() > 0 else 0

        # MS rating: priorità cache, fallback yfinance
        if not ms_cache:
            ms_cache = info.get('morningStarOverallRating') or info.get('morningStarRiskRating')
        if not category or category == 'N/A':
            category = info.get('category') or info.get('fundFamily') or manager or 'N/A'

        data = {
            'ISIN':          isin,
            'Ticker':        ticker,
            'Nome':          name,
            'Gestore':       manager,
            'Categoria':     category,
            'Valuta':        currency,
            'TER':           ter,
            'Sharpe Ratio':  sharpe,
            'AUM':           total_assets,
            'Var_1D_%':      var_1d,
            'Performance 1Y': perf_1y,
            'Perf 3M %':     perf_3m,
            'Perf 6M %':     perf_6m,
            'Perf YTD %':    perf_ytd,
            'Stelle MS':     ms_cache,
            'Prezzo':        price_now,
            'Data Dati':     datetime.now().strftime('%Y-%m-%d'),
            'Score':         0,
        }
        return data, 'ok', None

    except Exception as e:
        return None, "Eccezione", str(e)


# ── Piano configs ─────────────────────────────────────────────────────────────
def _build_plan_configs():
    _fixed    = {'BASIC': {'top_n': 20}, 'PRO': {'top_n': 50}, 'VALUE': {'top_n': 50}}
    _fallback = {
        'BASIC': {'ter_max': 1.00, 'sharpe_min': 0.4, 'aum_min': 50_000_000,  'performance_1y_min': -0.10},
        'PRO':   {'ter_max': 1.50, 'sharpe_min': 0.2, 'aum_min': 10_000_000,  'performance_1y_min': -0.30},
        'VALUE': {'ter_max': 2.00, 'sharpe_min': 0.1, 'aum_min':  1_000_000,  'performance_1y_min': -0.50},
    }
    try:
        with open(SERVIZI_FILE, encoding='utf-8') as f:
            sv = json.load(f).get('fondi_eu', sv.get('fondi', {}))
    except Exception:
        sv = {}
    configs = {}
    for plan, key in [('BASIC', 'basic'), ('PRO', 'pro'), ('VALUE', 'value')]:
        p = sv.get(key, {}).get('parametri', {}) if sv else {}
        cfg = dict(_fallback[plan])
        if p:
            for k in ('ter_max', 'sharpe_min', 'aum_min', 'performance_1y_min'):
                if k in p:
                    cfg[k] = float(p[k]) if k != 'aum_min' else int(p[k])
            if 'performance_1y_min' in p:
                cfg['performance_1y_min'] = float(p['performance_1y_min']) / 100
        cfg.update(_fixed[plan])
        configs[plan] = cfg
    return configs


def _build_plan_configs_safe():
    _fixed    = {'BASIC': {'top_n': 20}, 'PRO': {'top_n': 50}, 'VALUE': {'top_n': 50}}
    _fallback = {
        'BASIC': {'ter_max': 1.00, 'sharpe_min': 0.4, 'aum_min': 50_000_000,  'performance_1y_min': -0.10},
        'PRO':   {'ter_max': 1.50, 'sharpe_min': 0.2, 'aum_min': 10_000_000,  'performance_1y_min': -0.30},
        'VALUE': {'ter_max': 2.00, 'sharpe_min': 0.1, 'aum_min':  1_000_000,  'performance_1y_min': -0.50},
    }
    try:
        with open(SERVIZI_FILE, encoding='utf-8') as f:
            sv_all = json.load(f)
        sv = sv_all.get('fondi_eu', sv_all.get('fondi', {}))
        configs = {}
        for plan, key in [('BASIC', 'basic'), ('PRO', 'pro'), ('VALUE', 'value')]:
            p = sv.get(key, {}).get('parametri', {}) if sv else {}
            cfg = dict(_fallback[plan])
            if p:
                for k in ('ter_max', 'sharpe_min', 'aum_min'):
                    if k in p:
                        cfg[k] = float(p[k]) if k != 'aum_min' else int(p[k])
                if 'performance_1y_min' in p:
                    cfg['performance_1y_min'] = float(p['performance_1y_min']) / 100
            cfg.update(_fixed[plan])
            configs[plan] = cfg
        return configs
    except Exception as e:
        print(f"⚠️  servizi_config.json: {e} — uso default")
        return {pl: {**_fallback[pl], **_fixed[pl]} for pl in _fallback}


PLAN_CONFIGS = _build_plan_configs_safe()


def apply_filters(fund, cfg):
    reasons = []
    ter    = fund.get('TER', 0) or 0
    sharpe = fund.get('Sharpe Ratio', 0) or 0
    aum    = fund.get('AUM', 0) or 0
    perf1y = fund.get('Performance 1Y', 0) or 0

    if ter > cfg['ter_max'] / 100:
        reasons.append(f"TER alto ({format_percent_ita(ter, 2)})")
    if sharpe < cfg['sharpe_min']:
        reasons.append(f"Sharpe basso ({sharpe:.2f})")
    if aum < cfg['aum_min']:
        reasons.append(f"AUM basso ({aum:,.0f})")
    if perf1y < cfg['performance_1y_min']:
        reasons.append(f"Perf negativa ({format_percent_ita(perf1y, 1)})")
    return ('rejected', '; '.join(reasons)) if reasons else ('selected', None)


# ── Excel helpers ─────────────────────────────────────────────────────────────
def _fmt_num(n):
    if n is None:
        return 'N/A'
    try:
        n = float(n)
    except (TypeError, ValueError):
        return 'N/A'
    if abs(n) >= 1_000_000_000:
        return f"{n/1_000_000_000:.2f}B"
    if abs(n) >= 1_000_000:
        return f"{n/1_000_000:.2f}M"
    if abs(n) >= 1_000:
        return f"{n/1_000:.1f}K"
    return f"{n:.0f}"


def _score_fill(score):
    try:
        s = float(score)
    except (TypeError, ValueError):
        return None
    if s >= 75:
        return PatternFill(start_color="1A9850", end_color="1A9850", fill_type="solid")
    if s >= 55:
        return PatternFill(start_color="FEE08B", end_color="FEE08B", fill_type="solid")
    if s >= 35:
        return PatternFill(start_color="F46D43", end_color="F46D43", fill_type="solid")
    return PatternFill(start_color="D73027", end_color="D73027", fill_type="solid")


def _score_font(score):
    try:
        s = float(score)
    except (TypeError, ValueError):
        return Font(bold=True)
    color = "FFFFFF" if s < 55 or s >= 75 else "1A1A1A"
    return Font(bold=True, color=color)


def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def _autofit(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_w), max_w)


def _hdr_cell(ws, row, col, val, bg="1A3A5C", fg="FFFFFF", size=10):
    c = ws.cell(row=row, column=col, value=val)
    c.font  = Font(bold=True, color=fg, size=size)
    c.fill  = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = _thin_border()
    return c


def _write_plan_excel(plan_name, cfg, selected, rejected, non_validi, timestamp, universe_total):
    wb = Workbook()
    wb.remove(wb.active)

    C_HEADER_BG = "1A3A5C"
    C_TITLE_FG  = "FF8C42"
    C_SECTION_BG = "EBF2FA"
    C_ALT_ROW   = "F7FAFD"
    C_GREEN     = "1A9850"
    C_RED       = "D73027"
    C_ORANGE    = "F46D43"
    top_n = cfg['top_n']

    def _write_sc(wb, title, sheet_name, idx, data, cols):
        ws = wb.create_sheet(sheet_name, idx)
        ws.freeze_panes = 'A3'
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=11, color="444444")
        ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
        ws['A1'].fill = PatternFill(start_color=C_SECTION_BG, end_color=C_SECTION_BG, fill_type="solid")
        for c, h in enumerate(cols, 1):
            _hdr_cell(ws, 2, c, h)
        for r, fund in enumerate(data, 3):
            fill_row = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if r % 2 == 0 else None
            for c, h in enumerate(cols, 1):
                v = fund.get(h)
                if h == 'TER':              v = format_percent_ita(v, 2) if v else 'N/A'
                elif h == 'Performance 1Y': v = format_percent_ita(v, 1)
                elif h == 'AUM':            v = _fmt_num(v)
                elif isinstance(v, float):  v = round(v, 2)
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = _thin_border()
                if fill_row:
                    cell.fill = fill_row
        _autofit(ws)

    # ── Dashboard ──────────────────────────────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard", 0)
    ws_dash.sheet_view.showGridLines = False

    ws_dash.merge_cells('A1:F1')
    ws_dash['A1'] = f"FONDI EU SCREENER  ·  Piano {plan_name}  ·  Robot Trader 2026"
    ws_dash['A1'].font = Font(size=18, bold=True, color=C_TITLE_FG)
    ws_dash['A1'].fill = PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type="solid")
    ws_dash['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[1].height = 36

    ws_dash.merge_cells('A2:F2')
    ws_dash['A2'] = f"Generato il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
    ws_dash['A2'].font = Font(size=10, italic=True, color="888888")
    ws_dash['A2'].alignment = Alignment(horizontal='center')

    ws_dash['A4'] = "CRITERI APPLICATI"
    ws_dash['A4'].font = Font(bold=True, size=11, color=C_HEADER_BG)
    criteri = [
        ("TER massimo",        f"≤ {cfg['ter_max']:.2f}%"),
        ("Sharpe minimo",      f"≥ {cfg['sharpe_min']}"),
        ("AUM minimo",         f"≥ {_fmt_num(cfg['aum_min'])}"),
        ("Performance 1Y min", f"≥ {cfg['performance_1y_min']*100:.0f}%"),
    ]
    for i, (k, v) in enumerate(criteri, 5):
        ws_dash.cell(row=i, column=1, value=k).font = Font(bold=True, color="555555")
        ws_dash.cell(row=i, column=2, value=v).font = Font(color=C_HEADER_BG)

    totale = len(selected) + len(rejected) + len(non_validi)
    ws_dash['D4'] = "RISULTATI SCREENING"
    ws_dash['D4'].font = Font(bold=True, size=11, color=C_HEADER_BG)
    stat_rows = [
        ("Universo Fondi EU",     universe_total,   "444444"),
        ("Con dati Yahoo",        totale,            "444444"),
        ("✅ Selezionati",         len(selected),    C_GREEN),
        ("❌ Scartati (filtri)",   len(rejected),    C_ORANGE),
        ("⚠️  Non validi",         len(non_validi),  C_RED),
        ("Tasso selezione",       f"{len(selected)/totale*100:.1f}%" if totale else "0%", C_GREEN),
    ]
    for i, (label, value, color) in enumerate(stat_rows, 5):
        ws_dash.cell(row=i, column=4, value=label).font = Font(bold=True, color="555555")
        ws_dash.cell(row=i, column=5, value=value).font = Font(bold=True, color=color)

    top5 = sorted(selected, key=lambda x: x.get('Score', 0), reverse=True)[:5]
    if top5:
        r_start = 13
        ws_dash.cell(row=r_start, column=1, value="TOP 5 FONDI EU PER SCORE").font = Font(bold=True, size=11, color=C_HEADER_BG)
        ws_dash.merge_cells(f'A{r_start}:G{r_start}')
        for c, h in enumerate(['#', 'ISIN', 'Nome', 'Gestore', 'Score', 'Perf 1Y', 'TER'], 1):
            _hdr_cell(ws_dash, r_start+1, c, h)
        for rank, fund in enumerate(top5, 1):
            row = r_start + 1 + rank
            score = fund.get('Score', 0)
            perf1y = fund.get('Performance 1Y')
            ws_dash.cell(row=row, column=1, value=rank).alignment = Alignment(horizontal='center')
            ws_dash.cell(row=row, column=2, value=fund.get('ISIN')).font = Font(bold=True, size=9, color=C_HEADER_BG)
            ws_dash.cell(row=row, column=3, value=fund.get('Nome'))
            ws_dash.cell(row=row, column=4, value=fund.get('Gestore'))
            sc = ws_dash.cell(row=row, column=5, value=round(score, 1))
            sc.fill = _score_fill(score)
            sc.font = _score_font(score)
            sc.alignment = Alignment(horizontal='center')
            p1 = ws_dash.cell(row=row, column=6, value=round(perf1y*100, 2) if perf1y else 'N/A')
            p1.font = Font(color=C_GREEN if (perf1y or 0) >= 0 else C_RED)
            ws_dash.cell(row=row, column=7, value=format_percent_ita(fund.get('TER'), 2) if fund.get('TER') else 'N/A')
            for c in range(1, 8):
                ws_dash.cell(row=row, column=c).border = _thin_border()

    r_leg = 22
    ws_dash.cell(row=r_leg, column=1, value="LEGENDA SCORE").font = Font(bold=True, size=9, color="888888")
    for col, (label, color) in enumerate([("≥75 Ottimo","1A9850"),("55-74 Buono","FEE08B"),("35-54 Medio","F46D43"),("<35 Basso","D73027")], 1):
        c = ws_dash.cell(row=r_leg+1, column=col, value=label)
        c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        c.font = Font(bold=True, size=9, color="FFFFFF" if color in ("1A9850","D73027","F46D43") else "333333")
        c.alignment = Alignment(horizontal='center')

    for col, w in zip('ABCDEFG', [30, 18, 32, 22, 18, 14, 12]):
        ws_dash.column_dimensions[col].width = w

    # ── Top N per Score ────────────────────────────────────────────────────────
    ws_top = wb.create_sheet(f"Top {top_n} per Score", 1)
    ws_top.freeze_panes = 'A2'
    ws_top.sheet_view.showGridLines = False
    top_funds = sorted(selected, key=lambda x: x.get('Score', 0), reverse=True)[:top_n]
    if top_funds:
        headers = ['#', 'ISIN', 'Ticker', 'Nome', 'Gestore', 'Score', 'Stelle MS',
                   'Perf 1Y %', 'Perf 6M %', 'Perf 3M %', 'Perf YTD %',
                   'TER', 'Sharpe', 'AUM', 'Valuta', 'Categoria']
        for col, header in enumerate(headers, 1):
            _hdr_cell(ws_top, 1, col, header)
        for row, fund in enumerate(top_funds, 2):
            score  = fund.get('Score', 0)
            perf1y = fund.get('Performance 1Y')
            fill_alt = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if row % 2 == 0 else None
            vals = [
                (1,  row-1,                        None,                   Alignment(horizontal='center')),
                (2,  fund.get('ISIN'),              Font(size=9, color=C_HEADER_BG), None),
                (3,  fund.get('Ticker'),            Font(bold=True, color=C_HEADER_BG), None),
                (4,  fund.get('Nome'),              None,                   None),
                (5,  fund.get('Gestore'),           None,                   None),
                (6,  round(score, 1),               _score_font(score),     Alignment(horizontal='center')),
                (7,  format_stelle_ms(fund.get('Stelle MS')), None,         Alignment(horizontal='center')),
                (8,  round(perf1y*100,2) if perf1y else 'N/A',
                     Font(color=C_GREEN if (perf1y or 0)>=0 else C_RED, bold=True), Alignment(horizontal='right')),
                (9,  fund.get('Perf 6M %'),         None,                   Alignment(horizontal='right')),
                (10, fund.get('Perf 3M %'),         None,                   Alignment(horizontal='right')),
                (11, fund.get('Perf YTD %'),        None,                   Alignment(horizontal='right')),
                (12, format_percent_ita(fund.get('TER'),2) if fund.get('TER') else 'N/A', None, Alignment(horizontal='center')),
                (13, round(fund.get('Sharpe Ratio',0),2), None,             Alignment(horizontal='right')),
                (14, _fmt_num(fund.get('AUM',0)),   None,                   Alignment(horizontal='right')),
                (15, fund.get('Valuta','EUR'),       None,                   Alignment(horizontal='center')),
                (16, fund.get('Categoria','N/A'),   None,                   None),
            ]
            for col, val, font, align in vals:
                cell = ws_top.cell(row=row, column=col, value=val)
                if col == 6:
                    cell.fill = _score_fill(score)
                elif fill_alt:
                    cell.fill = fill_alt
                if font:
                    cell.font = font
                if align:
                    cell.alignment = align
                cell.border = _thin_border()
        ws_top.row_dimensions[1].height = 22
        _autofit(ws_top, min_w=6, max_w=45)

    # ── Fondi Selezionati (raw) ────────────────────────────────────────────────
    if selected:
        ws_sel = wb.create_sheet("Fondi Selezionati", 2)
        ws_sel.freeze_panes = 'A2'
        hdr = list(selected[0].keys())
        for col, header in enumerate(hdr, 1):
            _hdr_cell(ws_sel, 1, col, header)
        for row, fund in enumerate(selected, 2):
            fill_alt = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if row % 2 == 0 else None
            for col, key in enumerate(hdr, 1):
                value = fund.get(key)
                if key == 'TER' and isinstance(value, float):
                    value = format_percent_ita(value, 2)
                elif key == 'Performance 1Y' and isinstance(value, float):
                    value = format_percent_ita(value, 1)
                elif key == 'Stelle MS':
                    value = format_stelle_ms(value)
                elif key == 'AUM':
                    value = _fmt_num(value)
                elif isinstance(value, float):
                    value = round(value, 2)
                cell = ws_sel.cell(row=row, column=col, value=value)
                cell.border = _thin_border()
                if fill_alt:
                    cell.fill = fill_alt
        _autofit(ws_sel)

    # ── Scartati ──────────────────────────────────────────────────────────────
    def _m(f):
        return f.get('Motivo Scarto', '')
    sc_ter    = [f for f in rejected if 'TER alto' in _m(f)]
    sc_sharpe = [f for f in rejected if 'Sharpe basso' in _m(f) and 'TER' not in _m(f)]
    sc_aum    = [f for f in rejected if 'AUM basso' in _m(f) and 'TER' not in _m(f) and 'Sharpe' not in _m(f)]
    sc_perf   = [f for f in rejected if 'Perf negativa' in _m(f) and 'TER' not in _m(f)]
    class_ids = set(id(f) for f in sc_ter + sc_sharpe + sc_aum + sc_perf)
    sc_altri  = [f for f in rejected if id(f) not in class_ids]

    _write_sc(wb, f"TER > {cfg['ter_max']:.2f}% ({len(sc_ter)})",
              "Scartati - TER Alto", 3, sc_ter,
              ['ISIN','Nome','Gestore','TER','Sharpe Ratio','Performance 1Y','Valuta','Motivo Scarto'])
    _write_sc(wb, f"Sharpe < {cfg['sharpe_min']} ({len(sc_sharpe)})",
              "Scartati - Sharpe Basso", 4, sc_sharpe,
              ['ISIN','Nome','Gestore','Sharpe Ratio','TER','Performance 1Y','Valuta','Motivo Scarto'])
    _write_sc(wb, f"AUM < {_fmt_num(cfg['aum_min'])} ({len(sc_aum)})",
              "Scartati - AUM Basso", 5, sc_aum,
              ['ISIN','Nome','Gestore','AUM','TER','Sharpe Ratio','Valuta','Motivo Scarto'])
    _write_sc(wb, f"Perf1Y < {cfg['performance_1y_min']*100:.0f}% ({len(sc_perf)})",
              "Scartati - Performance", 6, sc_perf,
              ['ISIN','Nome','Gestore','Performance 1Y','TER','Sharpe Ratio','Valuta','Motivo Scarto'])
    _write_sc(wb, f"Altri motivi ({len(sc_altri)})",
              "Scartati - Altri", 7, sc_altri,
              ['ISIN','Nome','Gestore','Motivo Scarto','TER','Sharpe Ratio'])

    # ── Non Validi ────────────────────────────────────────────────────────────
    ws_nv = wb.create_sheet("Non Validi", 8)
    ws_nv.freeze_panes = 'A3'
    ws_nv.merge_cells('A1:D1')
    ws_nv['A1'] = f"NON VALIDI — DATI MANCANTI ({len(non_validi)})"
    ws_nv['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_nv['A1'].fill = PatternFill(start_color=C_ORANGE, end_color=C_ORANGE, fill_type="solid")
    ws_nv['A1'].alignment = Alignment(horizontal='center')
    for col, h in enumerate(['ISIN', 'Ticker', 'Campo Mancante', 'Motivo'], 1):
        _hdr_cell(ws_nv, 2, col, h)
    for row, item in enumerate(non_validi, 3):
        ws_nv.cell(row=row, column=1, value=item.get('ISIN', '')).border = _thin_border()
        ws_nv.cell(row=row, column=2, value=item.get('Ticker', '')).border = _thin_border()
        ws_nv.cell(row=row, column=3, value=item.get('Campo Mancante', '—')).border = _thin_border()
        ws_nv.cell(row=row, column=4, value=item.get('Motivo', '—')).border = _thin_border()
    _autofit(ws_nv)

    filename = os.path.join(REPORTS_DIR, f"FONDI_EU_Screener_{plan_name}_{timestamp}.xlsx")
    wb.save(filename)
    print(f"✅ [{plan_name}] Report salvato: {filename}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
eu_universe = load_eu_universe()
universe_total = len(eu_universe)

print("=" * 70)
print("FONDI EU SCREENER - Robot Trader 2026")
print("=" * 70)
print(f"Fondi EU da analizzare: {universe_total}")
print("=" * 70)

if universe_total == 0:
    print("⚠️  Nessun fondo EU disponibile. Esegui prima:")
    print("   python fetch_fondi_eu_universe.py --seed")
    sys.exit(0)

# PHASE 1: raccolta dati
all_fondi_eu  = []
non_validi_global = []
total = universe_total

for i, entry in enumerate(eu_universe, 1):
    ticker = entry.get('yahoo_ticker', '')
    isin   = entry.get('isin', ticker)
    print(f"[{i}/{total}] {isin} ({ticker})...", flush=True)
    data, status, detail = None, None, None
    for attempt in range(1, 4):
        data, status, detail = analyze_fund_eu(entry)
        if data is not None:
            break
        if not _is_network_error(detail):
            break
        if attempt < 3:
            print(f"  ↻ Errore rete — retry {attempt}/2 tra 5s...", flush=True)
            time.sleep(5)

    if data:
        all_fondi_eu.append(data)
        print(f"  ✅ OK{' (tentativo '+str(attempt)+')' if attempt > 1 else ''}", flush=True)
    else:
        non_validi_global.append({
            'ISIN':          isin,
            'Ticker':        ticker,
            'Campo Mancante': status or 'Sconosciuto',
            'Motivo':        detail or 'Dati insufficienti',
        })
        print(f"  ⚠️  [{status}]: {detail}", flush=True)

print("=" * 70)
print(f"📊 Raccolti: {len(all_fondi_eu)} | Non validi: {len(non_validi_global)}")
print("=" * 70)

_net_err = sum(1 for x in non_validi_global if _is_network_error(x.get('Motivo', '')))
if _net_err > 20:
    print(f"\n⚠️  ALERT: {_net_err} errori di rete — invio email alert...", flush=True)
    _send_alert_email(total, _net_err, len(non_validi_global))

# PHASE 2: filtro e output per piano
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

for plan_name, cfg in PLAN_CONFIGS.items():
    print(f"\n{'─'*60}")
    print(f"▶ Piano {plan_name}: TER≤{cfg['ter_max']:.2f}% | Sharpe≥{cfg['sharpe_min']} | "
          f"AUM≥{_fmt_num(cfg['aum_min'])} | Perf1Y≥{cfg['performance_1y_min']*100:.0f}% | Top{cfg['top_n']}")

    p_sel = []
    p_rej = []
    for fund in all_fondi_eu:
        sc = fund.copy()
        result, motivo = apply_filters(sc, cfg)
        if result == 'selected':
            p_sel.append(sc)
        else:
            sc['Motivo Scarto'] = motivo
            p_rej.append(sc)

    batch_percentile_score(p_sel, 'fondi', plan_name)
    print(f"→ {len(p_sel)} selezionati / {len(p_rej)} scartati / {len(non_validi_global)} non validi")
    _write_plan_excel(plan_name, cfg, p_sel, p_rej, non_validi_global, timestamp, universe_total)

print(f"\n{'='*60}")
print(f"✅ Tutti e 3 i piani FONDI EU completati.")
print(f"Fine: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"{'='*60}")

# ── Aggiorna prices_cache.json ────────────────────────────────────────────────
try:
    _cache_path = os.path.join(BASE_DIR, "prices_cache.json")
    _existing = {}
    if os.path.exists(_cache_path):
        with open(_cache_path, encoding='utf-8') as _cf:
            _existing = json.load(_cf)
    _eu_cache = {}
    for s in all_fondi_eu:
        k = s.get('ISIN') or s.get('Ticker', '')
        if not k:
            continue
        p = s.get('Prezzo')
        _eu_cache[k] = {
            'name':       s.get('Nome', k),
            'price':      round(float(p), 4) if p else None,
            'change_pct': s.get('Var_1D_%'),
            'currency':   s.get('Valuta', 'EUR'),
        }
    _existing['fondi_eu']    = _eu_cache
    _existing['fondi_eu_at'] = datetime.now().isoformat()
    with open(_cache_path, 'w', encoding='utf-8') as _cf:
        json.dump(_existing, _cf, ensure_ascii=False)
    print(f"[Cache] prices_cache.json aggiornato: {len(_eu_cache)} fondi EU")
except Exception as _e:
    print(f"[Cache] Errore salvataggio: {_e}")
