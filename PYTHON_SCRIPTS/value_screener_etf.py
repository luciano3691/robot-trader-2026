# -*- coding: utf-8 -*-
"""
Robot Trader 2026 - ETF Screener
FIXED: Formato percentuale italiano + Parametri dinamici da parametri.json
"""
import socket
socket.setdefaulttimeout(20)
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import json
import time
import smtplib
from email.mime.text import MIMEText
from screener_utils import batch_percentile_score


def _is_network_error(detail):
    if not detail:
        return False
    d = detail.lower()
    return any(k in d for k in ['curl', 'resolve host', 'connection', 'timeout', 'network', 'errno 11001', 'recv failure'])


def _send_alert_email(asset_type, total, network_errors, non_validi_total):
    """Invia email di alert all'admin se troppi errori di rete durante lo screening."""
    try:
        with open(os.path.join(BASE_DIR, 'config.json'), encoding='utf-8') as f:
            cfg_email = json.load(f).get('email', {})
        host   = cfg_email.get('smtp_server', 'smtp.gmail.com')
        port   = cfg_email.get('smtp_port', 587)
        login  = cfg_email.get('smtp_login', '') or cfg_email.get('sender', '')
        pwd    = cfg_email.get('app_password', '')
        sender = cfg_email.get('sender', login)
        if not login or not pwd:
            print(f"[ALERT] SMTP non configurato — alert non inviato", flush=True)
            return
        soglia_pct = round(network_errors / total * 100, 1) if total else 0
        body = (
            f"ALERT — Robot Trader 2026 — Screener {asset_type}\n\n"
            f"Errori di rete rilevati durante lo screening:\n"
            f"  Errori rete: {network_errors} su {total} ({soglia_pct}%)\n"
            f"  Non validi totali: {non_validi_total}\n\n"
            f"Causa probabile: DNS non risolveva query2.finance.yahoo.com al momento del run.\n\n"
            f"AZIONE RICHIESTA — rilanciare lo screener:\n"
            f"  cd C:\\Users\\lucia\\Desktop\\Robot Trader 2026\\PYTHON_SCRIPTS\n"
            f"  python orchestrator.py {asset_type}\n\n"
            f"— Robot Trader 2026"
        )
        msg = MIMEText(body, 'plain', 'utf-8')
        msg['Subject'] = f"[ALERT] Screener {asset_type}: {network_errors} errori rete su {total} ({soglia_pct}%)"
        msg['From']    = sender
        msg['To']      = login
        with smtplib.SMTP(host, port, timeout=15) as srv:
            srv.ehlo(); srv.starttls()
            srv.login(login, pwd)
            srv.sendmail(sender, [login], msg.as_string())
        print(f"[ALERT] Email alert inviata a {login}", flush=True)
    except Exception as e:
        print(f"[ALERT] Errore invio alert email: {e}", flush=True)

# Path
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(os.path.dirname(BASE_DIR), "REPORTS_DAILY")
PARAMETRI_FILE = os.path.join(BASE_DIR, "parametri.json")
os.makedirs(REPORTS_DIR, exist_ok=True)

# Importa lista ETF
sys.path.insert(0, BASE_DIR)
from ticker_lists_5000 import ALL_ETF

# ── ETF Universe: carica da etf_universe_cache.json se disponibile ───────────
_JUSTETF_CACHE_FILE = os.path.join(BASE_DIR, 'etf_universe_cache.json')
_EU_SUFFIX_SET = frozenset(['L', 'DE', 'AS', 'MI', 'PA', 'SW', 'ST', 'HE',
                             'CO', 'OL', 'LI', 'BR', 'LS', 'MC', 'VI'])
_etf_meta_cache = {}  # base_ticker → {ter, distribution, replication, inception_date, isin}

if os.path.exists(_JUSTETF_CACHE_FILE):
    with open(_JUSTETF_CACHE_FILE, encoding='utf-8') as _jf:
        _raw_ju = json.load(_jf)
    _eu_tickers_cache = []
    for _isin, _entry in _raw_ju.items():
        if _entry.get('error') or not _entry.get('preferred_ticker'):
            continue
        _tk = _entry['preferred_ticker']
        _base = _tk.split('.')[0]
        _etf_meta_cache[_base] = {
            'isin':           _isin,
            'ter':            _entry.get('ter'),
            'distribution':   'Accumulating' if _entry.get('is_acc') is True
                              else ('Distributing' if _entry.get('is_acc') is False else None),
            'replication':    'Physical' if _entry.get('is_physical') is True
                              else ('Synthetic' if _entry.get('is_physical') is False else None),
            'inception_date': _entry.get('inception'),
        }
        _eu_tickers_cache.append(_tk)
    # US ETF = tickers senza suffisso EU
    _us_only = [t for t in ALL_ETF
                if '.' not in t or t.split('.')[-1] not in _EU_SUFFIX_SET]
    ETF_UNIVERSE = list(dict.fromkeys(_us_only + _eu_tickers_cache))
    print(f"[Cache] etf_universe_cache.json: {len(_eu_tickers_cache)} ETF UCITS caricati")
    print(f"[Cache] Universo totale: {len(ETF_UNIVERSE)} ETF "
          f"({len(_us_only)} US + {len(_eu_tickers_cache)} UCITS)")
else:
    ETF_UNIVERSE = ALL_ETF
    print(f"[Cache] etf_universe_cache.json non trovata — universo: {len(ALL_ETF)} ETF")
    print("        Esegui: python fetch_justetf_universe.py  per espandere a ~4800 UCITS")

# CARICA PARAMETRI DA JSON
def _get_param(params: dict, key: str, default):
    """Gestisce sia {'value': X} che plain X in parametri.json"""
    val = params.get(key, default)
    if isinstance(val, dict):
        return val.get('value', default)
    return val if val is not None else default


def load_filters():
    """Carica filtri da parametri.json"""
    try:
        with open(PARAMETRI_FILE, 'r', encoding='utf-8') as f:
            parametri = json.load(f)

        etf_params = parametri.get('etf', {})

        return {
            'ter_max':            _get_param(etf_params, 'ter_max',            0.50),
            'sharpe_min':         _get_param(etf_params, 'sharpe_min',         0.5),
            'volume_min':         _get_param(etf_params, 'volume_min',         100000),
            'performance_1y_min': _get_param(etf_params, 'performance_1y_min', -0.20),
            'min_age_years':      _get_param(etf_params, 'min_age_years',      3),
            'only_accumulating':  _get_param(etf_params, 'only_accumulating',  True),
            'only_physical':      _get_param(etf_params, 'only_physical',      True),
        }
    except Exception as e:
        print(f"⚠️ ERRORE caricamento parametri: {e}")
        print("⚠️ Uso valori di default")
        return {
            'ter_max': 0.50, 'sharpe_min': 0.5,
            'volume_min': 100000, 'performance_1y_min': -0.20,
            'min_age_years': 3, 'only_accumulating': True, 'only_physical': True,
        }

FILTERS = load_filters()

# ISIN map per ETF europei (yfinance non espone TER per listing EU)
# TER fetchato live da justETF al momento dello screening
EUROPEAN_ETF_ISIN = {
    # ── Vanguard ──────────────────────────────────────────────────────────────
    'VWRL': 'IE00B3RBWM25',  # Vanguard FTSE All-World distributing
    'VWCE': 'IE00BK5BQT80',  # Vanguard FTSE All-World accumulating
    'VWRP': 'IE00BK5BQT80',  # alias VWCE su LSE
    'VGWL': 'IE00BK5BQT80',  # alias VWCE su .DE
    'VUSA': 'IE00B3XXRP09',  # Vanguard S&P 500
    'VEUR': 'IE00B945VV12',  # Vanguard FTSE Developed Europe
    'VUKE': 'IE00B810Q511',  # Vanguard FTSE 100
    'VFEM': 'IE00B3VVMM84',  # Vanguard FTSE Emerging Markets
    'VAGP': 'IE00BG47KB92',  # Vanguard Global Aggregate Bond EUR Hedged
    'VAPX': 'IE00B9F5YL18',  # Vanguard FTSE Developed Asia Pacific ex Japan
    'VJPN': 'IE00B95PGT31',  # Vanguard FTSE Japan
    'VJPA': 'IE00B95PGT31',  # alias VJPN su .DE
    'VERX': 'IE00BKX55S42',  # Vanguard FTSE Developed Europe ex UK
    # ── iShares ETC metalli preziosi ──────────────────────────────────────────
    'IGLN': 'IE00B4ND3602',  # iShares Physical Gold ETC
    'SGLD': 'IE00B4ND3602',  # alias IGLN
    # ── WisdomTree Physical ───────────────────────────────────────────────────
    'PHAU': 'DE000A0N62G0',  # WisdomTree Physical Gold
    'PHAG': 'JE00B1VS3333',  # WisdomTree Physical Silver
    'PHPT': 'JE00B1VS3770',  # WisdomTree Physical Platinum
    # ── iShares Xetra (DAX / STOXX settori) ──────────────────────────────────
    'EXS1': 'DE0005933931',  # iShares Core DAX UCITS ETF
    'EXW1': 'DE0002635307',  # iShares STOXX Europe 600 UCITS ETF
    'EXSA': 'DE0005933956',  # iShares Core EURO STOXX 50 UCITS ETF (DE)
    'EXHD': 'DE000A0F5UG3',  # iShares DivDAX UCITS ETF
    'EXV3': 'DE000A0H0785',  # iShares STOXX Europe 600 Basic Resources
    'EXV6': 'DE0006289382',  # iShares STOXX Europe 600 Technology
    'EXV7': 'DE000A0H0799',  # iShares STOXX Europe 600 Utilities
    'EXV8': 'DE000A0Q4R36',  # iShares STOXX Europe 600 Health Care
    # ── iShares obbligazionari e factor ───────────────────────────────────────
    'IAGG': 'IE00B3F81409',  # iShares Core Global Aggregate Bond UCITS ETF
    'IQLT': 'IE00BP3QZD73',  # iShares Edge MSCI World Quality Factor UCITS ETF
    # ── HSBC ──────────────────────────────────────────────────────────────────
    'HGEM': 'IE00B5SSQL53',  # HSBC MSCI Emerging Markets UCITS ETF
    'HJPN': 'IE00B5VX7566',  # HSBC MSCI Japan UCITS ETF
    'HPEU': 'IE00BFBRDM48',  # HSBC Core MSCI Europe UCITS ETF
    'HPSS': 'IE00BFBRDL31',  # HSBC Core MSCI Pacific ex Japan UCITS ETF
    # HPUS: ISIN non funziona su justETF — gestito da cross-exchange fallback
    # ── Invesco ───────────────────────────────────────────────────────────────
    # PBUS: ISIN non funziona su justETF — gestito da cross-exchange fallback
    'PBEE': 'IE00BD0NC250',  # Invesco FTSE All World Emerging Markets UCITS ETF
    # ── SPDR / State Street ───────────────────────────────────────────────────
    'WDIV': 'IE00B9CQXS71',  # SPDR S&P Global Dividend Aristocrats UCITS ETF
    'SPGB': 'IE00B43QJJ40',  # SPDR Bloomberg Global Aggregate Bond UCITS ETF
    # ── WisdomTree equity/dividend ────────────────────────────────────────────
    'WQCD': 'IE00BZ56RN96',  # WisdomTree Global Quality Dividend Growth UCITS ETF
    # ── Amundi / Lyxor ────────────────────────────────────────────────────────
    'MXWU': 'LU1681043599',  # Amundi MSCI World UCITS ETF USD Acc
    # ── BNP Paribas Easy ──────────────────────────────────────────────────────
    'ESE':  'FR0011550185',  # BNP Paribas Easy S&P 500 EUR
}

# TER statici per ETF europei senza ISIN su justETF o con dati strutturalmente assenti
EUROPEAN_ETF_TER_STATIC = {
    'VMID': 0.10,  # Vanguard FTSE 250 UCITS ETF
    'VAGS': 0.15,  # Vanguard Global Aggregate Bond GBP Hedged Acc
}

# Cache locale: base_ticker → {'ter': float, 'distribution': str, 'replication': str}
_justetf_cache = {}

# Cache cross-exchange TER: base_ticker → ter (float) oppure None
_ter_crossexchange_cache = {}
_EU_EXTENSIONS = ['DE', 'L', 'AS', 'PA', 'MI', 'SW']

import urllib.request
import re as _re
from datetime import datetime as _dt

def get_justetf_data(ticker_full):
    """Fetch TER + distribution policy + replication da justETF.
    Ritorna dict {'ter', 'distribution', 'replication'} oppure None.
    """
    base = ticker_full.split('.')[0]

    if base in _justetf_cache:
        return _justetf_cache[base]

    # Fallback statico TER (VMID, VAGS) — replica fisica, nessun campo distribution
    if base in EUROPEAN_ETF_TER_STATIC:
        result = {'ter': EUROPEAN_ETF_TER_STATIC[base], 'distribution': None,
                  'replication': 'Physical', 'inception_date': None}
        _justetf_cache[base] = result
        return result

    isin = EUROPEAN_ETF_ISIN.get(base)
    if not isin:
        _justetf_cache[base] = None
        return None

    try:
        url = f"https://www.justetf.com/en/etf-profile.html?isin={isin}"
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        with urllib.request.urlopen(req, timeout=12) as resp:
            body = resp.read().decode('utf-8', errors='ignore')

        ter_m  = _re.search(r'etf-profile-header_ter-value[^>]*>([0-9.,]+)%', body)
        dist_m = _re.search(r'distribution-policy-value[^>]*>([^<]+)', body)
        rep_m  = _re.search(r'etf-profile-header_replication-value[^>]*>([^<]+)', body)
        # Inception date: justETF mostra DD.MM.YYYY — prova più pattern
        incep_m = _re.search(r'fund-inception[^>]*>([0-9]{2}\.[0-9]{2}\.[0-9]{4})', body, _re.IGNORECASE)
        if not incep_m:
            incep_m = _re.search(r'launch.?date[^>]*>([0-9]{2}\.[0-9]{2}\.[0-9]{4})', body, _re.IGNORECASE)

        result = {
            'ter':            float(ter_m.group(1).replace(',', '.')) if ter_m else None,
            'distribution':   dist_m.group(1).strip() if dist_m else None,
            'replication':    rep_m.group(1).strip() if rep_m else None,
            'inception_date': incep_m.group(1) if incep_m else None,  # "DD.MM.YYYY"
        }
        _justetf_cache[base] = result
        return result
    except Exception:
        pass

    _justetf_cache[base] = None
    return None


def detect_accumulating(info, justetf_data):
    """Rileva se l'ETF è ad accumulazione.
    Priorità: justETF (più affidabile) → nome → dividendi.
    """
    # 1. justETF distribution policy (per ETF europei)
    if justetf_data and justetf_data.get('distribution'):
        return 'accumulating' in justetf_data['distribution'].lower()

    # 2. Keywords nel nome
    name = (info.get('longName') or info.get('shortName') or '').lower()
    acc_kw = ['acc', 'accumul', '(c)', 'capitaliz', 'thesaurierend', 'reinvest']
    dist_kw = ['dist', 'distribut', '(d)', 'income', 'ausschüttend']
    if any(k in name for k in acc_kw):
        return True
    if any(k in name for k in dist_kw):
        return False

    # 3. Dividendi: se distribuisce → distributing
    div = (info.get('trailingAnnualDividendRate') or
           info.get('dividendRate') or
           info.get('dividendYield') or 0)
    if div and float(div) > 0:
        return False

    return None  # sconosciuto


def detect_physical(info, justetf_data):
    """Rileva se l'ETF usa replica fisica.
    Priorità: justETF → nome → assumi True per ETF USA.
    """
    # 1. justETF replication field
    if justetf_data and justetf_data.get('replication'):
        rep = justetf_data['replication'].lower()
        return 'synthetic' not in rep and 'swap' not in rep

    # 2. Keywords nel nome
    name = (info.get('longName') or info.get('shortName') or '').lower()
    if 'synthetic' in name or 'swap' in name or 'unfunded' in name:
        return False

    return True  # ETF USA e senza dati: assumi fisico

def format_percent_ita(value, decimals=1):
    """Formatta percentuale con virgola italiana (per Sharpe, Perf, ecc.)"""
    if value is None or (isinstance(value, float) and (value != value)):
        return 'N/A'
    return f"{value*100:.{decimals}f}%".replace('.', ',')

def format_stelle_ms(rating):
    """Formatta rating Morningstar come stelle (es. 4 → ★★★★☆)"""
    if not rating or not isinstance(rating, (int, float)):
        return 'N/A'
    r = max(0, min(5, int(rating)))
    return '★' * r + '☆' * (5 - r)

def format_ter_etf(ter, decimals=4):
    """Formatta TER ETF: yfinance restituisce già in % (0.0945 = 0.0945%)"""
    if ter is None or not isinstance(ter, (int, float)) or ter <= 0:
        return 'N/A'
    return f"{ter:.{decimals}f}%".replace('.', ',')

def analyze_etf(ticker):
    """Analizza singolo ETF"""
    try:
        etf = yf.Ticker(ticker)
        info = etf.info
        hist = etf.history(period="2y")

        if hist.empty or len(hist) < 50:
            return None, "Dati storici", "Dati storici insufficienti (nessun dato)"
        hist = hist.tail(252)  # usa solo l'ultimo anno per i calcoli
        
        # Metriche
        # TER: prova netExpenseRatio (ETF USA), poi annualReportExpenseRatio, poi funds_data
        ter = info.get('netExpenseRatio', info.get('annualReportExpenseRatio', info.get('totalExpenseRatio', None)))
        if ter is None:
            try:
                fd = etf.funds_data
                if fd and fd.fund_operations is not None:
                    ops = fd.fund_operations
                    if 'Annual Report Expense Ratio' in ops.index:
                        val = ops.loc['Annual Report Expense Ratio'].iloc[0]
                        if val is not None and str(val) != '<NA>':
                            ter = float(val)
            except Exception:
                pass
        base_ticker = ticker.split('.')[0]

        # Cross-exchange TER fallback: se TER mancante, prova altri exchange EU per lo stesso fondo
        # (es. LCWD.AS → None, LCWD.PA → 0.12). Cache per evitare ripetizioni su listing multipli.
        if (ter is None or ter <= 0) and '.' in ticker:
            if base_ticker in _ter_crossexchange_cache:
                ter = _ter_crossexchange_cache[base_ticker]
            else:
                for alt_ext in _EU_EXTENSIONS:
                    alt_ticker = f"{base_ticker}.{alt_ext}"
                    if alt_ticker == ticker:
                        continue
                    try:
                        alt_info = yf.Ticker(alt_ticker).info
                        ter_alt = (alt_info.get('netExpenseRatio') or
                                   alt_info.get('annualReportExpenseRatio') or
                                   alt_info.get('totalExpenseRatio'))
                        if ter_alt and ter_alt > 0:
                            ter = ter_alt
                            break
                    except Exception:
                        pass
                _ter_crossexchange_cache[base_ticker] = ter  # None se nessun exchange ha TER

        # justETF: usa cache pre-scaricata se disponibile, altrimenti fallback live
        justetf = None
        if base_ticker in _etf_meta_cache:
            # Dati già in cache da fetch_justetf_universe.py — nessuna richiesta HTTP
            _cm = _etf_meta_cache[base_ticker]
            justetf = {
                'ter':            _cm.get('ter'),
                'distribution':   _cm.get('distribution'),
                'replication':    _cm.get('replication'),
                'inception_date': _cm.get('inception_date'),
            }
        elif base_ticker in EUROPEAN_ETF_ISIN or base_ticker in EUROPEAN_ETF_TER_STATIC:
            justetf = get_justetf_data(ticker)

        # TER — Layer 1: yfinance + cross-exchange, Layer 2: justETF (cache o live)
        if ter is None or ter <= 0:
            if not justetf:
                justetf = get_justetf_data(ticker)
            if justetf:
                ter = justetf.get('ter')
            if ter is None or ter <= 0:
                return None, "TER", "TER non disponibile (yfinance + justETF)"

        volume   = info.get('volume', 0)
        name     = info.get('longName', info.get('shortName', ticker))
        category = info.get('category', 'N/A')

        # Età ETF — Layer 1: fundInceptionDate (yfinance)
        #            Layer 2: inception_date da justETF (DD.MM.YYYY)
        #            Layer 3: prima data price history (lower bound)
        inception = info.get('fundInceptionDate')
        if inception:
            age_years = (_dt.now() - _dt.fromtimestamp(inception)).days / 365.25
        else:
            age_years = None
            if justetf and justetf.get('inception_date'):
                try:
                    incep_dt = _dt.strptime(justetf['inception_date'], '%d.%m.%Y')
                    age_years = (_dt.now() - incep_dt).days / 365.25
                except ValueError:
                    pass
            if age_years is None and not hist.empty:
                # Lower bound: l'ETF esiste almeno da questa data
                first_date = hist.index[0].to_pydatetime().replace(tzinfo=None)
                age_years = (_dt.now() - first_date).days / 365.25

        # Tipo (ACC/DIST) e Replica
        is_acc      = detect_accumulating(info, justetf)
        is_physical = detect_physical(info, justetf)

        # Performance 1Y, 6M, 3M, YTD + variazione giornaliera
        price_1y_ago = hist['Close'].iloc[0] if len(hist) > 0 else None
        price_now    = hist['Close'].iloc[-1]
        perf_1y      = ((price_now / price_1y_ago) - 1) if price_1y_ago else 0
        var_1d       = round(((price_now / hist['Close'].iloc[-2]) - 1) * 100, 2) if len(hist) >= 2 else None

        perf_3m = round(((price_now / hist['Close'].iloc[-63]) - 1) * 100, 2) if len(hist) >= 63 else None
        perf_6m = round(((price_now / hist['Close'].iloc[-126]) - 1) * 100, 2) if len(hist) >= 126 else None
        current_year = datetime.now().year
        hist_ytd = hist[hist.index.year == current_year]
        perf_ytd = round(((price_now / hist_ytd['Close'].iloc[0]) - 1) * 100, 2) if not hist_ytd.empty else None

        # Sharpe (approssimato)
        returns = hist['Close'].pct_change().dropna()
        sharpe  = (returns.mean() / returns.std() * (252 ** 0.5)) if returns.std() > 0 else 0

        net_assets = info.get('totalAssets', 0) or 0
        ms_rating  = info.get('morningStarOverallRating') or info.get('morningStarRiskRating')

        isin = (_etf_meta_cache.get(base_ticker, {}).get('isin') or
                EUROPEAN_ETF_ISIN.get(base_ticker) or info.get('isin') or '')

        data = {
            'Ticker':        ticker,
            'ISIN':          isin,
            'Nome':          name,
            'Categoria':     category,
            'TER':           ter,
            'Sharpe Ratio':  sharpe,
            'Volume':        volume,
            'Net Assets':    net_assets,
            'Var_1D_%':      var_1d,
            'Performance 1Y': perf_1y,
            'Perf 3M %':     perf_3m,
            'Perf 6M %':     perf_6m,
            'Perf YTD %':    perf_ytd,
            'Età (anni)':    round(age_years, 1) if age_years else 'N/A',
            'Tipo':          'ACC' if is_acc else ('DIST' if is_acc is False else 'N/A'),
            'Replica':       'Fisica' if is_physical else 'Sintetica',
            'Stelle MS':     ms_rating,
            'Prezzo':        price_now,
            'Data Dati':     datetime.now().strftime('%Y-%m-%d'),
        }

        data['Score'] = 0  # calcolato in batch dopo la selezione per piano
        return data, 'ok', None

    except Exception as e:
        return None, "Eccezione", str(e)

# ─── 3-PLAN SYSTEM ────────────────────────────────────────────────────────────

SERVIZI_FILE = os.path.join(BASE_DIR, 'servizi_config.json')

def _build_etf_plan_configs():
    """Legge i parametri filtro da servizi_config.json (modificabili dalla dashboard).
    I parametri non configurabili (età, acc, fisica, top_n) restano hardcoded.
    performance_1y_min in servizi è in % intera (es. 5 = +5%) → converti in decimale (/100).
    """
    _fixed = {
        'BASIC': {'min_age_years': 5, 'only_accumulating': True, 'only_physical': True, 'top_n': 20},
        'PRO':   {'min_age_years': 3, 'only_accumulating': True, 'only_physical': True, 'top_n': 50},
        'VALUE': {'min_age_years': 2, 'only_accumulating': True, 'only_physical': True, 'top_n': 50},
    }
    _fallback = {
        'BASIC': {'ter_max': 0.35, 'sharpe_min': 0.8, 'volume_min': 500_000, 'performance_1y_min': -0.10},
        'PRO':   {'ter_max': 0.50, 'sharpe_min': 0.5, 'volume_min': 100_000, 'performance_1y_min': -0.20},
        'VALUE': {'ter_max': 0.75, 'sharpe_min': 0.3, 'volume_min':  50_000, 'performance_1y_min': -0.30},
    }
    try:
        with open(SERVIZI_FILE, encoding='utf-8') as f:
            sv = json.load(f).get('etf', {})
        configs = {}
        for plan, key in [('BASIC','basic'), ('PRO','pro'), ('VALUE','value')]:
            p = sv.get(key, {}).get('parametri', {})
            cfg = dict(_fallback[plan])
            if p:
                if 'ter_max'            in p: cfg['ter_max']            = float(p['ter_max'])
                if 'sharpe_min'         in p: cfg['sharpe_min']         = float(p['sharpe_min'])
                if 'volume_min'         in p: cfg['volume_min']         = int(p['volume_min'])
                if 'performance_1y_min' in p: cfg['performance_1y_min'] = float(p['performance_1y_min']) / 100
            cfg.update(_fixed[plan])
            configs[plan] = cfg
        return configs
    except Exception as e:
        print(f"⚠️ servizi_config.json non leggibile ({e}) — uso valori di default")
        return {pl: {**_fallback[pl], **_fixed[pl]} for pl in _fallback}

ETF_PLAN_CONFIGS = _build_etf_plan_configs()


def apply_etf_plan_filters(etf, cfg):
    """Applica filtri piano-specifici a un ETF già analizzato.
    Ritorna ('selected', None) o ('rejected', motivo_stringa).
    """
    reasons = []
    ter       = etf.get('TER', 0) or 0
    sharpe    = etf.get('Sharpe Ratio', 0) or 0
    volume    = etf.get('Volume', 0) or 0
    perf_1y   = etf.get('Performance 1Y', 0) or 0
    age_years = etf.get('Età (anni)')
    tipo      = etf.get('Tipo', 'N/A')
    replica   = etf.get('Replica', '')

    if ter > cfg['ter_max']:
        reasons.append(f"TER alto ({ter:.4f}%)")
    if sharpe < cfg['sharpe_min']:
        reasons.append(f"Sharpe basso ({sharpe:.2f})")
    if volume < cfg['volume_min']:
        reasons.append(f"Volume basso ({volume:,})")
    if perf_1y < cfg['performance_1y_min']:
        reasons.append(f"Performance negativa ({format_percent_ita(perf_1y, 1)})")
    if isinstance(age_years, (int, float)) and age_years < cfg['min_age_years']:
        reasons.append(f"ETF troppo giovane ({age_years:.1f} anni < {cfg['min_age_years']})")
    if cfg['only_accumulating'] and tipo != 'ACC':
        label = "DIST" if tipo == 'DIST' else "tipo non rilevato"
        reasons.append(f"Non accumulazione ({label})")
    if cfg['only_physical'] and replica != 'Fisica':
        reasons.append("Replica sintetica")

    if reasons:
        return 'rejected', '; '.join(reasons)
    return 'selected', None


def _fmt_num(n):
    """Formatta numeri grandi in forma leggibile: 1234567 → 1.23M"""
    if n is None:
        return 'N/A'
    try:
        n = float(n)
    except (TypeError, ValueError):
        return 'N/A'
    if n == 0:
        return '0'
    if abs(n) >= 1_000_000_000:
        return f"{n/1_000_000_000:.2f}B"
    if abs(n) >= 1_000_000:
        return f"{n/1_000_000:.2f}M"
    if abs(n) >= 1_000:
        return f"{n/1_000:.1f}K"
    return f"{n:.0f}"

def _score_fill(score):
    """Colore cella in base allo score: verde→giallo→arancio→rosso"""
    try:
        s = float(score)
    except (TypeError, ValueError):
        return None
    if s >= 75:
        return PatternFill(start_color="1A9850", end_color="1A9850", fill_type="solid")  # verde
    if s >= 55:
        return PatternFill(start_color="FEE08B", end_color="FEE08B", fill_type="solid")  # giallo
    if s >= 35:
        return PatternFill(start_color="F46D43", end_color="F46D43", fill_type="solid")  # arancio
    return PatternFill(start_color="D73027", end_color="D73027", fill_type="solid")      # rosso

def _score_font(score):
    try:
        s = float(score)
    except (TypeError, ValueError):
        return Font(bold=True)
    color = "FFFFFF" if s < 55 or s >= 75 else "1A1A1A"
    return Font(bold=True, color=color)

def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _autofit(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, min_w), max_w)

_EU_ISIN_PREFIXES = ('IE','LU','FR','DE','GB','CH','AT','BE','NL','ES','IT','DK','SE','FI','NO','PL','CZ','HU')

def _ticker_url(ticker, isin=None):
    if isin and isin[:2] in _EU_ISIN_PREFIXES:
        return f"https://www.justetf.com/en/etf-profile.html?isin={isin}"
    return f"https://finance.yahoo.com/quote/{ticker}"

def _set_ticker_link(cell, ticker, isin=None, bold=True, color="0563C1"):
    if not ticker:
        return
    cell.hyperlink = _ticker_url(ticker, isin)
    cell.font = Font(bold=bold, color=color, underline="single")

_ETF_SEL_COLS = [
    'Ticker','ISIN','Nome','Categoria','Età (anni)','Tipo','Replica',
    'Stelle MS','TER','Sharpe Ratio','Volume','Net Assets',
    'Prezzo','Var_1D_%','Performance 1Y',
    'Perf 3M %','Perf 6M %','Perf YTD %','Score','Data Dati'
]

def _fmt_etf_sel_val(etf, col):
    """Formatta valori per il foglio ETF Selezionati — numeri leggibili."""
    v = etf.get(col)
    if v is None:
        return 'N/A'
    if col in {'Volume', 'Net Assets'}:
        return _fmt_num(v)
    if col == 'TER':
        return format_ter_etf(v)
    if col == 'Stelle MS':
        return format_stelle_ms(v)
    if col == 'Performance 1Y':
        return format_percent_ita(v, 1) if isinstance(v, (int, float)) else 'N/A'
    if col in {'Var_1D_%', 'Perf 3M %', 'Perf 6M %', 'Perf YTD %'}:
        return f"{v:.2f}%".replace('.', ',') if isinstance(v, (int, float)) else 'N/A'
    if col in {'Sharpe Ratio', 'Prezzo'}:
        return round(float(v), 2) if isinstance(v, (int, float)) else 'N/A'
    if col == 'Score':
        return round(float(v), 1) if isinstance(v, (int, float)) else 0
    return v

def _write_etf_plan_excel(plan_name, cfg, selected, rejected, non_validi, timestamp):
    """Scrive un Excel per il piano specificato — tutti i piani hanno gli stessi fogli."""
    wb = Workbook()
    wb.remove(wb.active)
    top_n = cfg['top_n']

    # ── colori palette ──────────────────────────────────────────
    C_HEADER_BG  = "1A3A5C"   # blu scuro intestazioni principali
    C_HEADER_FG  = "FFFFFF"
    C_TITLE_FG   = "FF8C42"   # arancio titolo
    C_SECTION_BG = "EBF2FA"   # azzurro chiaro sezioni
    C_ALT_ROW    = "F7FAFD"   # grigio-blu alternato
    C_GREEN      = "1A9850"
    C_RED        = "D73027"
    C_ORANGE     = "F46D43"

    def _hdr_cell(ws, row, col, val, bg=C_HEADER_BG, fg=C_HEADER_FG, size=10):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color=fg, size=size)
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _thin_border()
        return c

    def _write_sc(wb, title, sheet_name, idx, data, cols):
        ws = wb.create_sheet(sheet_name, idx)
        ws.freeze_panes = 'A3'
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=11, color="444444")
        ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
        ws['A1'].fill = PatternFill(start_color=C_SECTION_BG, end_color=C_SECTION_BG, fill_type="solid")
        for c, h in enumerate(cols, 1):
            _hdr_cell(ws, 2, c, h)
        for r, etf in enumerate(data, 3):
            fill_row = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if r % 2 == 0 else None
            for c, h in enumerate(cols, 1):
                v = etf.get(h)
                if h == 'TER':               v = format_ter_etf(v)
                elif h == 'Performance 1Y':  v = format_percent_ita(v, 1)
                elif h == 'Volume':          v = _fmt_num(v)
                elif h == 'Net Assets':      v = _fmt_num(v)
                elif isinstance(v, float):   v = round(v, 2)
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = _thin_border()
                if fill_row:
                    cell.fill = fill_row
                if h == 'Ticker':
                    _set_ticker_link(cell, etf.get('Ticker'), etf.get('ISIN'))
        _autofit(ws)
        return ws

    # ── FOGLIO 1: Dashboard ─────────────────────────────────────
    ws_dash = wb.create_sheet("Dashboard", 0)
    ws_dash.sheet_view.showGridLines = False

    # Titolo
    ws_dash.merge_cells('A1:F1')
    ws_dash['A1'] = f"ETF SCREENER  ·  Piano {plan_name}  ·  Robot Trader 2026"
    ws_dash['A1'].font = Font(size=18, bold=True, color=C_TITLE_FG)
    ws_dash['A1'].fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
    ws_dash['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_dash.row_dimensions[1].height = 36

    ws_dash.merge_cells('A2:F2')
    ws_dash['A2'] = f"Generato il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
    ws_dash['A2'].font = Font(size=10, italic=True, color="888888")
    ws_dash['A2'].alignment = Alignment(horizontal='center')

    # Criteri filtro applicati
    ws_dash['A4'] = "CRITERI APPLICATI"
    ws_dash['A4'].font = Font(bold=True, size=11, color=C_HEADER_BG)
    criteri = [
        ("TER massimo",         f"≤ {cfg['ter_max']:.2f}%"),
        ("Sharpe minimo",       f"≥ {cfg['sharpe_min']}"),
        ("Volume minimo",       f"≥ {_fmt_num(cfg['volume_min'])}"),
        ("Performance 1Y min",  f"≥ {cfg['performance_1y_min']*100:.0f}%"),
        ("Età minima ETF",      f"≥ {cfg['min_age_years']} anni"),
        ("Tipo",                "Solo Accumulazione"),
        ("Replica",             "Solo Fisica"),
    ]
    for i, (k, v) in enumerate(criteri, 5):
        ws_dash.cell(row=i, column=1, value=k).font = Font(bold=True, color="555555")
        cell_v = ws_dash.cell(row=i, column=2, value=v)
        cell_v.font = Font(color="1A3A5C")

    # Statistiche
    totale = len(selected) + len(rejected) + len(non_validi)
    ws_dash['D4'] = "RISULTATI SCREENING"
    ws_dash['D4'].font = Font(bold=True, size=11, color=C_HEADER_BG)
    stat_rows = [
        ("Universo ETF",          len(ETF_UNIVERSE), "444444"),
        ("✅ Selezionati",         len(selected),   C_GREEN),
        ("❌ Scartati (filtri)",   len(rejected),   C_ORANGE),
        ("⚠️  Non validi",         len(non_validi), C_RED),
        ("Totale verificato",     totale,          "444444"),
        ("Tasso selezione",       f"{len(selected)/totale*100:.1f}%" if totale else "0%", C_GREEN),
    ]
    for i, (label, value, color) in enumerate(stat_rows, 5):
        lbl = ws_dash.cell(row=i, column=4, value=label)
        lbl.font = Font(bold=True, color="555555")
        val = ws_dash.cell(row=i, column=5, value=value)
        val.font = Font(bold=True, color=color)

    # Top 5 direttamente in Dashboard
    top5 = sorted(selected, key=lambda x: x.get('Score', 0), reverse=True)[:5]
    if top5:
        r_start = 13
        ws_dash.cell(row=r_start, column=1, value="TOP 5 ETF PER SCORE").font = Font(bold=True, size=11, color=C_HEADER_BG)
        ws_dash.merge_cells(f'A{r_start}:F{r_start}')
        hdrs5 = ['#', 'Ticker', 'Nome', 'Score', 'Perf 1Y %', 'TER']
        for c, h in enumerate(hdrs5, 1):
            _hdr_cell(ws_dash, r_start+1, c, h)
        for rank, etf in enumerate(top5, 1):
            row = r_start + 1 + rank
            score = etf.get('Score', 0)
            perf1y = etf.get('Performance 1Y')
            ws_dash.cell(row=row, column=1, value=rank).alignment = Alignment(horizontal='center')
            _tk_cell = ws_dash.cell(row=row, column=2, value=etf.get('Ticker'))
            _set_ticker_link(_tk_cell, etf.get('Ticker'), etf.get('ISIN'))
            ws_dash.cell(row=row, column=3, value=etf.get('Nome'))
            sc = ws_dash.cell(row=row, column=4, value=round(score, 1))
            sc.fill = _score_fill(score)
            sc.font = _score_font(score)
            sc.alignment = Alignment(horizontal='center')
            p1 = ws_dash.cell(row=row, column=5, value=round(perf1y*100, 2) if perf1y else 'N/A')
            p1.font = Font(color=C_GREEN if (perf1y or 0) >= 0 else C_RED)
            ws_dash.cell(row=row, column=6, value=format_ter_etf(etf.get('TER')))
            for c in range(1, 7):
                ws_dash.cell(row=row, column=c).border = _thin_border()

    # Legenda score
    r_leg = r_start + 8 if top5 else 14
    ws_dash.cell(row=r_leg, column=1, value="LEGENDA SCORE").font = Font(bold=True, size=9, color="888888")
    for col, (label, color) in enumerate([("≥75 Ottimo","1A9850"),("55-74 Buono","FEE08B"),("35-54 Medio","F46D43"),("<35 Basso","D73027")], 1):
        c = ws_dash.cell(row=r_leg+1, column=col, value=label)
        c.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        c.font = Font(bold=True, size=9, color="FFFFFF" if color in ("1A9850","D73027","F46D43") else "333333")
        c.alignment = Alignment(horizontal='center')

    for col, w in zip('ABCDEF', [28, 16, 28, 22, 16, 12]):
        ws_dash.column_dimensions[col].width = w
    ws_dash.row_dimensions[r_start+1].height = 20

    # ── FOGLIO 2: Top N per Score ──────────────────────────────
    ws_top = wb.create_sheet(f"Top {top_n} per Score", 1)
    ws_top.freeze_panes = 'A2'
    ws_top.sheet_view.showGridLines = False
    top_etfs = sorted(selected, key=lambda x: x.get('Score', 0), reverse=True)[:top_n]
    if top_etfs:
        headers = ['#', 'Ticker', 'Nome', 'Score', 'Stelle MS', 'Perf 1Y %', 'Perf 6M %',
                   'Perf 3M %', 'Perf YTD %', 'TER', 'Sharpe', 'Volume', 'AUM (Net Assets)']
        for col, header in enumerate(headers, 1):
            _hdr_cell(ws_top, 1, col, header)
        for row, etf in enumerate(top_etfs, 2):
            score = etf.get('Score', 0)
            perf1y = etf.get('Performance 1Y')
            fill_alt = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if row % 2 == 0 else None

            cells_vals = [
                (1,  row - 1,                None,                                         Alignment(horizontal='center')),
                (2,  etf.get('Ticker'),       Font(bold=True, color=C_HEADER_BG),           None),
                (3,  etf.get('Nome'),         None,                                         None),
                (4,  round(score, 1),         _score_font(score),                           Alignment(horizontal='center')),
                (5,  format_stelle_ms(etf.get('Stelle MS')), None,                          Alignment(horizontal='center')),
                (6,  round(perf1y*100, 2) if perf1y else 'N/A',
                     Font(color=C_GREEN if (perf1y or 0) >= 0 else C_RED, bold=True),       Alignment(horizontal='right')),
                (7,  etf.get('Perf 6M %'),   None,                                         Alignment(horizontal='right')),
                (8,  etf.get('Perf 3M %'),   None,                                         Alignment(horizontal='right')),
                (9,  etf.get('Perf YTD %'),  None,                                         Alignment(horizontal='right')),
                (10, format_ter_etf(etf.get('TER')), None,                                 Alignment(horizontal='center')),
                (11, round(etf.get('Sharpe Ratio', 0), 2), None,                           Alignment(horizontal='right')),
                (12, _fmt_num(etf.get('Volume', 0)),  None,                                Alignment(horizontal='right')),
                (13, _fmt_num(etf.get('Net Assets', 0)), None,                             Alignment(horizontal='right')),
            ]
            for col, val, font, align in cells_vals:
                cell = ws_top.cell(row=row, column=col, value=val)
                if col == 4:
                    cell.fill = _score_fill(score)
                elif fill_alt:
                    cell.fill = fill_alt
                if col == 2:
                    _set_ticker_link(cell, etf.get('Ticker'), etf.get('ISIN'))
                elif font:
                    cell.font = font
                if align:
                    cell.alignment = align
                cell.border = _thin_border()

        ws_top.row_dimensions[1].height = 22
        _autofit(ws_top, min_w=6, max_w=45)

    # ── FOGLIO 3: ETF Selezionati ──────────────────────────────
    if selected:
        ws_sel = wb.create_sheet("ETF Selezionati", 2)
        ws_sel.freeze_panes = 'A2'
        for col, header in enumerate(_ETF_SEL_COLS, 1):
            _hdr_cell(ws_sel, 1, col, header)
        for row, etf in enumerate(sorted(selected, key=lambda x: x.get('Score', 0), reverse=True), 2):
            fill_alt = PatternFill(start_color=C_ALT_ROW, end_color=C_ALT_ROW, fill_type="solid") if row % 2 == 0 else None
            for col, key in enumerate(_ETF_SEL_COLS, 1):
                cell = ws_sel.cell(row=row, column=col, value=_fmt_etf_sel_val(etf, key))
                cell.border = _thin_border()
                if fill_alt:
                    cell.fill = fill_alt
                if key == 'Ticker':
                    _set_ticker_link(cell, etf.get('Ticker'), etf.get('ISIN'))
        _autofit(ws_sel)

    # ── FOGLI SCARTATI ─────────────────────────────────────────
    def _m(e): return e.get('Motivo Scarto', '')
    sc_ter    = [e for e in rejected if 'TER alto' in _m(e)]
    sc_sharpe = [e for e in rejected if 'Sharpe basso' in _m(e) and 'TER' not in _m(e)]
    sc_vol    = [e for e in rejected if 'Volume basso' in _m(e) and 'TER' not in _m(e) and 'Sharpe' not in _m(e)]
    sc_perf   = [e for e in rejected if 'Performance negativa' in _m(e) and 'TER' not in _m(e)]
    sc_acc    = [e for e in rejected if 'Non accumulazione' in _m(e) and 'TER' not in _m(e) and 'Sharpe' not in _m(e) and 'Volume' not in _m(e)]
    sc_eta    = [e for e in rejected if 'troppo giovane' in _m(e) and 'TER' not in _m(e) and 'Non acc' not in _m(e)]
    sc_sint   = [e for e in rejected if 'sintetica' in _m(e).lower() and 'TER' not in _m(e)]
    class_ids = set(id(e) for e in sc_ter+sc_sharpe+sc_vol+sc_perf+sc_acc+sc_eta+sc_sint)
    sc_altri  = [e for e in rejected if id(e) not in class_ids]

    _write_sc(wb, f"SCARTATI PER TER > {cfg['ter_max']:.2f}% ({len(sc_ter)})",
              "Scartati - TER Alto", 3, sc_ter,
              ['Ticker','Nome','TER','Sharpe Ratio','Performance 1Y','Categoria','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI PER SHARPE < {cfg['sharpe_min']} ({len(sc_sharpe)})",
              "Scartati - Sharpe Basso", 4, sc_sharpe,
              ['Ticker','Nome','Sharpe Ratio','TER','Performance 1Y','Categoria','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI PER VOLUME < {cfg['volume_min']:,} ({len(sc_vol)})",
              "Scartati - Volume Basso", 5, sc_vol,
              ['Ticker','Nome','Volume','TER','Sharpe Ratio','Categoria','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI PER PERF1Y < {cfg['performance_1y_min']*100:.0f}% ({len(sc_perf)})",
              "Scartati - Performance", 6, sc_perf,
              ['Ticker','Nome','Performance 1Y','TER','Sharpe Ratio','Categoria','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI — NON ACCUMULAZIONE ({len(sc_acc)})",
              "Scartati - Non Accumul.", 7, sc_acc,
              ['Ticker','Nome','Tipo','Replica','TER','Sharpe Ratio','Performance 1Y','Età (anni)','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI — ETF GIOVANE < {cfg['min_age_years']}y ({len(sc_eta)})",
              "Scartati - ETF Giovani", 8, sc_eta,
              ['Ticker','Nome','Età (anni)','TER','Sharpe Ratio','Performance 1Y','Motivo Scarto'])
    _write_sc(wb, f"SCARTATI — REPLICA SINTETICA ({len(sc_sint)})",
              "Scartati - Sintetica", 9, sc_sint,
              ['Ticker','Nome','Replica','TER','Sharpe Ratio','Performance 1Y','Motivo Scarto'])
    _write_sc(wb, f"ALTRI MOTIVI ({len(sc_altri)})",
              "Scartati - Altri Motivi", 10, sc_altri,
              ['Ticker','Nome','Motivo Scarto','TER','Sharpe Ratio'])

    # ── FOGLIO Non Validi ──────────────────────────────────────
    ws_nv = wb.create_sheet("Non Validi", 11)
    ws_nv.freeze_panes = 'A3'
    ws_nv.merge_cells('A1:C1')
    ws_nv['A1'] = f"NON VALIDI — DATI MANCANTI ({len(non_validi)})"
    ws_nv['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_nv['A1'].fill = PatternFill(start_color=C_ORANGE, end_color=C_ORANGE, fill_type="solid")
    ws_nv['A1'].alignment = Alignment(horizontal='center')
    for col, h in enumerate(['Ticker', 'Campo Mancante', 'Motivo'], 1):
        _hdr_cell(ws_nv, 2, col, h)
    for row, item in enumerate(non_validi, 3):
        _nv_tk = ws_nv.cell(row=row, column=1, value=item.get('Ticker'))
        _nv_tk.border = _thin_border()
        _set_ticker_link(_nv_tk, item.get('Ticker'), item.get('ISIN'))
        ws_nv.cell(row=row, column=2, value=item.get('Campo Mancante', '—')).border = _thin_border()
        ws_nv.cell(row=row, column=3, value=item.get('Motivo', '—')).border = _thin_border()
    _autofit(ws_nv)

    filename = os.path.join(REPORTS_DIR, f"ETF_Screener_{plan_name}_{timestamp}.xlsx")
    wb.save(filename)
    print(f"✅ [{plan_name}] Report salvato: {filename}")


print("="*70)
print("ETF SCREENER - Robot Trader 2026")
print("="*70)
print(f"ETF totali da analizzare: {len(ETF_UNIVERSE)}")
print("="*70)

# PHASE 1: raccolta dati su tutti gli ETF (senza filtri piano)
all_etfs = []
non_validi_global = []

for i, ticker in enumerate(ETF_UNIVERSE, 1):
    print(f"[{i}/{len(ETF_UNIVERSE)}] {ticker}...", flush=True)
    result, status, detail = None, None, None
    for attempt in range(1, 4):  # max 3 tentativi
        result, status, detail = analyze_etf(ticker)
        if result is not None:
            break
        if not _is_network_error(detail):
            break  # errore dati strutturale — non ha senso riprovare
        if attempt < 3:
            print(f"  ↻ Errore rete — retry {attempt}/2 tra 5s...", flush=True)
            time.sleep(5)

    if result:
        all_etfs.append(result)
        if attempt > 1:
            print(f"  ✅ Raccolto al tentativo {attempt}", flush=True)
        else:
            print(f"  ✅ Raccolto", flush=True)
    else:
        campo  = status or 'Sconosciuto'
        motivo = detail or 'Dati insufficienti'
        non_validi_global.append({'Ticker': ticker, 'Campo Mancante': campo, 'Motivo': motivo})
        print(f"  ⚠️  Non valido [{campo}]: {motivo}", flush=True)

print("="*70)
print(f"📊 Raccolti: {len(all_etfs)} | Non validi: {len(non_validi_global)}")
print("="*70)

# ── Alert admin se troppi errori di rete ─────────────────────────────────────
_net_err_count = sum(1 for x in non_validi_global if _is_network_error(x.get('Motivo', '')))
if _net_err_count > 50:
    print(f"\n⚠️  ALERT: {_net_err_count} errori di rete su {len(ETF_UNIVERSE)} ETF — invio email alert...", flush=True)
    _send_alert_email('ETF', len(ETF_UNIVERSE), _net_err_count, len(non_validi_global))

# DEDUPLICAZIONE per ISIN — stesso fondo su più exchange conta una volta sola
def _dedup_by_isin(etfs):
    """Per ogni ISIN noto, preferisce il listing ACC (più liquido tra ACC).
    Se nessun listing è ACC, tiene il più liquido in assoluto.
    ETF senza ISIN (US) passano invariati.
    """
    grouped = {}
    no_isin = []
    for etf in etfs:
        isin = (etf.get('ISIN') or '').strip()
        if isin:
            grouped.setdefault(isin, []).append(etf)
        else:
            no_isin.append(etf)
    deduped = list(no_isin)
    for isin, group in grouped.items():
        acc_group = [e for e in group if e.get('Tipo') == 'ACC']
        best_pool = acc_group if acc_group else group
        best = max(best_pool, key=lambda e: e.get('Volume', 0) or 0)
        if len(group) > 1:
            dupes = [e['Ticker'] for e in group if e is not best]
            tipo = best.get('Tipo', '?')
            print(f"  [DEDUP] {isin}: mantenuto {best['Ticker']} ({tipo}, vol={best.get('Volume',0):,}), rimossi: {dupes}")
        deduped.append(best)
    return deduped

pre_dedup = len(all_etfs)
all_etfs  = _dedup_by_isin(all_etfs)
print(f"📊 Dopo deduplicazione ISIN: {len(all_etfs)} ETF unici (rimossi {pre_dedup - len(all_etfs)} duplicati)")
print("="*70)

# PHASE 2: filtro e output per ciascun piano
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

for plan_name, cfg in ETF_PLAN_CONFIGS.items():
    print(f"\n{'─'*60}")
    print(f"▶ Piano {plan_name}: TER≤{cfg['ter_max']:.2f}% | Sharpe≥{cfg['sharpe_min']} | "
          f"Vol≥{cfg['volume_min']:,} | Perf1Y≥{cfg['performance_1y_min']*100:.0f}% | Top{cfg['top_n']}")

    p_sel = []
    p_rej = []

    for etf in all_etfs:
        sc = etf.copy()
        result, motivo = apply_etf_plan_filters(sc, cfg)
        if result == 'selected':
            p_sel.append(sc)
        else:
            sc['Motivo Scarto'] = motivo
            p_rej.append(sc)

    # Score percentile 0-100 calcolato in batch (pesi letti da config.json)
    batch_percentile_score(p_sel, 'etf', plan_name)
    print(f"→ {len(p_sel)} selezionati / {len(p_rej)} scartati / {len(non_validi_global)} non validi")
    _write_etf_plan_excel(plan_name, cfg, p_sel, p_rej, non_validi_global, timestamp)

print(f"\n{'='*60}")
print(f"✅ Tutti e 3 i piani ETF completati.")
print(f"Fine: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(f"{'='*60}")

# ── Salva prices_cache.json (usato dalla tab DATABASE del dashboard) ──────────
try:
    _cache_path = os.path.join(BASE_DIR, "prices_cache.json")
    _existing_cache = {}
    if os.path.exists(_cache_path):
        with open(_cache_path, encoding='utf-8') as _cf:
            _existing_cache = json.load(_cf)
    _etf_cache = {}
    for s in all_etfs:
        t = s.get('Ticker', '')
        if not t:
            continue
        p = s.get('Prezzo')
        _etf_cache[t] = {
            'name':       s.get('Nome', t),
            'price':      round(float(p), 4) if p else None,
            'change_pct': s.get('Var_1D_%'),
            'currency':   s.get('Valuta', ''),
        }
    _existing_cache['etf']    = _etf_cache
    _existing_cache['etf_at'] = datetime.now().isoformat()
    with open(_cache_path, 'w', encoding='utf-8') as _cf:
        json.dump(_existing_cache, _cf, ensure_ascii=False)
    print(f"[Cache] prices_cache.json aggiornato: {len(_etf_cache)} ETF")
except Exception as _e:
    print(f"[Cache] Errore salvataggio: {_e}")
